import streamlit as st
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
import io
import math
import json
from datetime import datetime, timezone, timedelta
from supabase import create_client

# --- CONFIGURACIÓN VISUAL DE LA APP ---
st.set_page_config(
    page_title="ObraCubic - Grandes Cosas Comienzan Aquí",
    page_icon="https://raw.githubusercontent.com/luissalazarbastias-star/obracubic/refs/heads/main/Foto%202.png",
    initial_sidebar_state="collapsed"
)

st.markdown("""
    <style>
    [data-testid="stToolbar"] {visibility: hidden !important;}
    </style>
""", unsafe_allow_html=True)

# ============================
# CONEXIÓN A SUPABASE
# ============================
@st.cache_resource
def conectar_supabase():
    """Crea el cliente de Supabase usando los secrets. Devuelve None si falla."""
    try:
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
        return create_client(url, key)
    except Exception as e:
        st.session_state["_supabase_error"] = str(e)
        return None

supabase = conectar_supabase()

# ============================
# GUARDADO DE PROYECTOS (Supabase)
# ============================
LIMITE_GRATIS = 5

# Claves de sistema que NO se guardan como parte de la cubicación
CLAVES_SISTEMA = {
    "nav_option", "vista_cuenta", "usuario", "usuario_nombre", "usuario_plan",
    "_goto", "_supabase_error", "ir_a_cubicacion",
    "login_email", "login_pass", "reg_nombre", "reg_email", "reg_pass",
    "reg_pass2", "reg_terminos", "nombre_nuevo_proyecto_guardar",
    "precios_usuario_cargados", "precios_usuario_dict", "_precios_actuales",
    "presupuesto_actual", "usuario_plan_vence",
    "usuario_empresa", "usuario_rut", "usuario_telefono", "usuario_logo_url",
}

def _es_serializable(valor):
    try:
        json.dumps(valor)
        return True
    except Exception:
        return False

def capturar_cubicacion():
    """Toma una foto de los datos de cubicación en session_state."""
    datos = {}
    for k, v in st.session_state.items():
        if k in CLAVES_SISTEMA:
            continue
        if k.startswith("btn_") or k.startswith("FormSubmitter"):
            continue
        if _es_serializable(v):
            datos[k] = v
    return datos

def listar_proyectos(usuario_id):
    try:
        res = supabase.table("proyectos").select("id, nombre, actualizado_en") \
            .eq("usuario_id", usuario_id).order("actualizado_en", desc=True).execute()
        return res.data or []
    except Exception:
        return []

def guardar_proyecto(usuario_id, nombre, datos):
    supabase.table("proyectos").insert({
        "usuario_id": usuario_id, "nombre": nombre, "datos": datos,
    }).execute()

def cargar_proyecto(proyecto_id):
    res = supabase.table("proyectos").select("datos").eq("id", proyecto_id).single().execute()
    return res.data["datos"] if res.data else {}

def eliminar_proyecto(proyecto_id):
    supabase.table("proyectos").delete().eq("id", proyecto_id).execute()


def listar_bitacora(usuario_id, proyecto):
    """Devuelve las entradas de bitácora de un proyecto, más recientes primero."""
    try:
        res = supabase.table("bitacora").select("id, nota, foto, creado_en") \
            .eq("usuario_id", usuario_id).eq("proyecto", proyecto) \
            .order("creado_en", desc=True).execute()
        return res.data or []
    except Exception:
        return []


def guardar_bitacora(usuario_id, proyecto, nota, foto_base64=None):
    """Inserta una entrada de bitácora (nota y, opcionalmente, una foto en base64)."""
    supabase.table("bitacora").insert({
        "usuario_id": usuario_id,
        "proyecto": proyecto,
        "nota": nota,
        "foto": foto_base64,
    }).execute()


def eliminar_bitacora(entrada_id):
    supabase.table("bitacora").delete().eq("id", entrada_id).execute()


def cargar_precios_usuario(usuario_id):
    """Devuelve un dict {material: precio} con los precios guardados del usuario."""
    try:
        res = supabase.table("precios_usuario").select("material, precio").eq("usuario_id", usuario_id).execute()
        return {fila["material"]: fila["precio"] for fila in (res.data or [])}
    except Exception:
        return {}


def guardar_precios_usuario(usuario_id, precios):
    """Guarda/actualiza un dict {material: precio} para el usuario (upsert)."""
    if not precios:
        return
    filas = [
        {"usuario_id": usuario_id, "material": mat, "precio": int(p)}
        for mat, p in precios.items() if p and p > 0
    ]
    if filas:
        supabase.table("precios_usuario").upsert(filas, on_conflict="usuario_id,material").execute()


def cargar_datos_profesional(usuario_id):
    """Carga empresa, rut y teléfono del usuario desde Supabase a la sesión."""
    try:
        perfil = supabase.table("perfiles").select("empresa, rut, telefono").eq("id", usuario_id).single().execute()
        st.session_state["usuario_empresa"] = perfil.data.get("empresa")
        st.session_state["usuario_rut"] = perfil.data.get("rut")
        st.session_state["usuario_telefono"] = perfil.data.get("telefono")
    except Exception:
        pass


def guardar_datos_profesional(usuario_id, empresa, rut, telefono):
    """Guarda los datos profesionales del usuario en Supabase."""
    supabase.table("perfiles").update({
        "empresa": empresa or None,
        "rut": rut or None,
        "telefono": telefono or None,
    }).eq("id", usuario_id).execute()


def refrescar_plan_usuario():
    """Vuelve a leer el plan y su vencimiento desde Supabase y actualiza la sesión.
    Aplica el vencimiento automático (si venció, queda en gratis)."""
    usuario = st.session_state.get("usuario")
    if not usuario:
        return
    try:
        perfil = supabase.table("perfiles").select("plan, plan_vence").eq("id", usuario["id"]).single().execute()
        plan_bd = perfil.data.get("plan", "gratis")
        vence = perfil.data.get("plan_vence")
        st.session_state["usuario_plan_vence"] = vence
        st.session_state["usuario_plan"] = _plan_vigente(plan_bd, vence)
    except Exception:
        pass


# ============================
# SISTEMA DE PLANES
# ============================
# Planes: "gratis", "pro_basico", "pro_elite"
# (el caso "sin cuenta" = sin sesión iniciada)

PLANES_INFO = {
    "gratis":     {"nombre": "Plan Gratis", "emoji": "☕", "precio": "$0"},
    "pro_basico": {"nombre": "Plan Pro Básico", "emoji": "🚀", "precio": "$5.990 / mes"},
    "pro_elite":  {"nombre": "Plan Pro Élite", "emoji": "👑", "precio": "$14.990 / mes"},
}

# Jerarquía de planes (para comparar "al menos este nivel")
NIVEL_PLAN = {"sin_cuenta": 0, "gratis": 1, "pro_basico": 2, "pro_elite": 3}

def _plan_vigente(plan_bd, plan_vence):
    """Devuelve el plan efectivo considerando la fecha de vencimiento.
    Si el plan de pago ya venció, devuelve 'gratis'."""
    plan_pago = plan_bd in ("pro_basico", "pro_elite", "premium")
    if plan_pago and plan_vence:
        try:
            from datetime import datetime, timezone
            # plan_vence viene como string ISO desde Supabase
            if isinstance(plan_vence, str):
                fecha_vence = datetime.fromisoformat(plan_vence.replace("Z", "+00:00"))
            else:
                fecha_vence = plan_vence
            if fecha_vence.tzinfo is None:
                fecha_vence = fecha_vence.replace(tzinfo=timezone.utc)
            ahora = datetime.now(timezone.utc)
            if ahora > fecha_vence:
                return "gratis"  # el plan venció
        except Exception:
            pass
    return plan_bd

def plan_actual():
    """Devuelve el plan del usuario actual: 'sin_cuenta', 'gratis', 'pro_basico' o 'pro_elite'."""
    if not st.session_state.get("usuario"):
        return "sin_cuenta"
    plan = st.session_state.get("usuario_plan", "gratis")
    # Compatibilidad: el plan antiguo "premium" se trata como Pro Élite
    if plan == "premium":
        return "pro_elite"
    if plan not in NIVEL_PLAN:
        return "gratis"
    return plan

def tiene_nivel(plan_minimo):
    """True si el plan actual es al menos el nivel indicado."""
    return NIVEL_PLAN.get(plan_actual(), 0) >= NIVEL_PLAN.get(plan_minimo, 99)

# Permisos por funcionalidad (qué plan mínimo se requiere)
def puede_presupuesto():
    """Generar presupuestos: Pro Básico hacia arriba."""
    return tiene_nivel("pro_basico")

def puede_pdf_con_logo():
    """PDF personalizado con logo del usuario: Pro Básico hacia arriba."""
    return tiene_nivel("pro_basico")

def datos_usuario_pdf():
    """Arma el diccionario de datos profesionales para el PDF, desde la sesión.
    Devuelve None si el usuario no es Pro (así el PDF usa el branding de ObraCubic)."""
    if not puede_pdf_con_logo():
        return None
    datos = {
        "empresa": st.session_state.get("usuario_empresa") or st.session_state.get("usuario_nombre"),
        "rut": st.session_state.get("usuario_rut"),
        "email": (st.session_state.get("usuario") or {}).get("email"),
        "telefono": st.session_state.get("usuario_telefono"),
        "logo_url": st.session_state.get("usuario_logo_url"),
    }
    # Si no hay ningún dato útil, devolver None
    if not any(datos.values()):
        return None
    return datos

def puede_cubicaciones_ilimitadas():
    """Cubicaciones ilimitadas: Pro Básico hacia arriba."""
    return tiene_nivel("pro_basico")

def puede_exportar_excel():
    """Exportar a Excel: solo Pro Élite."""
    return tiene_nivel("pro_elite")

def puede_apu():
    """Presupuesto avanzado (APU): solo Pro Élite."""
    return tiene_nivel("pro_elite")

def puede_bitacora():
    """Bitácora de obra: solo Pro Élite."""
    return tiene_nivel("pro_elite")

def usuario_es_premium():
    """Compatibilidad: True si el usuario tiene algún plan de pago (Pro Básico o Élite)."""
    return tiene_nivel("pro_basico")


def parsear_cantidad(valor):
    """Extrae (cantidad, unidad) de un texto tipo '43 sacos'. Devuelve (None, None) si no hay número."""
    import re
    v = str(valor).strip()
    m = re.match(r'^([\d.,]+)\s*(.*)$', v)
    if not m:
        return None, None
    num_str = m.group(1)
    resto = m.group(2).strip()
    try:
        if num_str.count('.') == 1 and len(num_str.split('.')[1]) == 3 and ',' not in num_str:
            num = float(num_str.replace('.', ''))   # miles: 4.987 -> 4987
        else:
            num = float(num_str.replace('.', '').replace(',', '.')) if num_str.count(',') == 1 else float(num_str.replace(',', ''))
    except Exception:
        try:
            num = float(num_str.replace(',', '.'))
        except Exception:
            return None, None
    return num, resto


def fmt_clp(valor):
    """Formatea un número como pesos chilenos: 1234567 -> $1.234.567"""
    try:
        return "$" + f"{int(round(valor)):,}".replace(",", ".")
    except Exception:
        return "$0"


# Etiquetas que NO son materiales comprables (datos informativos) — usado por el APU
APU_NO_MATERIALES = [
    "área", "area", "superficie", "volumen", "dosificación", "dosificacion",
    "espesor", "medida", "tipo", "manos", "capas", "traslape", "tramos",
    "dirección x", "dirección y", "direccion x", "direccion y", "altura",
    "pendiente", "par inclinado", "pendolón", "pendolon", "largo", "ancho",
    "metros lineales netos", "acero total",
]


def materiales_de_partida_apu(bloque):
    """Devuelve [(material, cantidad, unidad)] comprables de un bloque de cubicación."""
    out = []
    for etiqueta, valor in bloque.get("items", []):
        eb = etiqueta.lower()
        if any(p in eb for p in APU_NO_MATERIALES):
            continue
        cant, uni = parsear_cantidad(valor)
        if cant is None or cant <= 0:
            continue
        # Conversiones básicas para que la unidad tenga sentido de compra
        if uni and "kg" in uni.lower():
            if "gravilla" in eb:
                cant, uni = round(cant / 1500, 2), "m³"
            elif "arena" in eb:
                cant, uni = round(cant / 1600, 2), "m³"
        out.append((etiqueta, round(cant, 2), uni or ""))
    return out


def medida_de_partida_apu(bloque):
    """Detecta (cantidad, unidad) de la medida base de la partida: volumen, área,
    superficie o largo. Devuelve (0, '') si no la encuentra."""
    claves = ["volumen", "área", "area", "superficie", "largo total", "metros lineales"]
    for etiqueta, valor in bloque.get("items", []):
        eb = etiqueta.lower()
        if any(c in eb for c in claves):
            limpio = str(valor).replace("≈", "").strip()
            cant, uni = parsear_cantidad(limpio)
            if cant and cant > 0:
                return round(cant, 2), (uni or "")
    return 0.0, ""


# Precios referenciales NETOS (sin IVA). Valores de ejemplo, el usuario los ajusta.
# La coincidencia es por palabra clave dentro del nombre del material.
PRECIOS_REFERENCIALES = [
    # (palabra_clave_en_minusculas, precio_neto). Orden: más específico primero.
    # --- Hormigón ---
    ("fibrocemento", 8395),  # antes que "cemento" para evitar colisión
    ("cemento", 3992),       # saco 25kg
    ("gravilla", 26891),     # m³
    ("arena mortero", 23950),
    ("arena", 23950),        # m³
    # --- Acero ---
    ("fierro 8", 3269),
    ("fierro 10", 5034),
    ("fierro 12", 7269),
    ("fierro 16", 13017),
    ("fierro", 7269),        # barra 6m (12mm por defecto)
    ("alambre", 2429),       # kg
    ("acero", 7269),
    ("barra", 7269),
    # --- Moldajes ---
    ("terciado ranurado", 23941),
    ("terciado estructural", 18479),
    ("terciado", 18479),
    ("tabla moldaje", 18479),
    ("pino dimensionado", 4109),
    # --- Muros ---
    ("ladrillo", 454),       # unidad
    # --- Metalcon ---
    ("zócalo", 2429),        # antes que "canal"/"cal"
    ("zocalo", 2429),
    ("guardapolvo", 2429),
    ("zinc", 10496),         # antes que "canal" (acanalado)
    ("canal", 4613),         # barra 3m
    ("montantes perf", 5286),
    ("montantes normal", 4193),
    ("montante perf", 5286),
    ("montante normal", 4193),
    ("montante", 4193),
    ("diagonal", 3529),
    ("lana de vidrio", 16378),  # rollo
    # --- Madera tabique ---
    ("pino tabique", 3571),
    ("pino", 3571),
    ("soleras", 3571),
    ("cadenetas", 3571),
    ("total listones", 3571),
    ("aislante", 16378),
    # --- Revestimientos ---
    ("yeso cartón", 6126),
    ("yeso carton", 6126),
    ("siding", 4109),
    ("osb", 9655),
    ("planchas", 8395),      # genérico revestimiento
    ("tablas", 4109),
    ("fijaciones", 2933),
    # --- Pisos ---
    ("cerámico", 5874),
    ("ceramico", 5874),
    ("pegamento", 5454),
    ("fragüe", 2092),
    ("fragüe", 2092),
    ("piso flotante", 14773),
    ("flotante", 14773),
    ("baldosa", 12597),
    ("deck", 15958),
    ("tornillos galvanizados", 4445),
    # --- Terminaciones ---
    ("pintura", 21000),      # galón
    ("sellador", 12597),
    ("pasta muro", 18990),   # tarro 25 kg (ref. Tajamar/Sodimac)
    ("cinta", 3521),
    ("estuco", 3521),
    ("cal", 4950),
    ("cielo", 5874),
    ("perfiles at", 2681),
    ("portante", 2681),
    # --- Cubierta madera ---
    ("costanera metálica", 15538),
    ("costanera metalica", 15538),
    ("costanera", 2092),     # madera 2x2 unidad
    ("madera cerchas", 3571),
    # --- Cubierta metálica ---
    ("omega 92", 9655),
    ("omega 70", 7555),
    ("omega", 9655),
    ("perfil c", 19319),
    # --- Planchas cubierta ---
    ("teja asf", 29403),     # paquete
    ("teja de arcilla", 521),
    ("teja arcilla", 521),
    ("teja", 521),
    ("panel sándwich", 18067),
    ("panel sandwich", 18067),
    ("sándwich", 18067),
    ("sandwich", 18067),
    # --- Aislación ---
    ("fieltro", 12513),      # rollo
    ("aislación térmica", 16378),
    ("aislacion termica", 16378),
    # --- Cierres perimetrales y faena ---
    # (van antes del "malla" genérico para que ganen las mallas específicas)
    ("malla eslabonada", 42000),      # rollo simple torsión
    ("malla rachel", 24990),          # rollo malla sombra / raschel
    ("malla raschel", 24990),
    ("rachel", 24990),
    ("malla naranja", 9990),          # rollo 50m señalización faena
    ("malla señalización", 9990),
    ("malla senalizacion", 9990),
    ("señalización naranja", 9990),
    ("acmafor", 32000),               # panel provisional móvil
    ("panel provisional", 32000),
    ("base hormig", 8990),            # base móvil prefabricada
    ("polines", 3500),                # polín impregnado 4" 3,2m
    ("polin", 3500),
    ("polín", 3500),
    ("poste", 7990),                  # poste metálico / madera de soporte
    ("perfiles met", 9655),
    ("perfil metál", 9655),
    ("perfil metal", 9655),
    ("riel", 4613),                   # riel / solera horizontal (perfil 3m)
    ("pie derecho", 3571),
    ("abrazadera", 1490),
    ("tubo phs", 11990),
    ("phs", 11990),
    ("zinc alum", 10496),
    # --- Tornillos / clavos (genéricos al final) ---
    ("tornillos autoperf", 10916),  # caja 1.000 metalcon
    ("tornillos yeso", 8395),
    ("tornillo", 10916),
    ("clavos", 2681),        # kg
    ("clavo", 2681),
    ("malla", 45000),
]

def precio_referencial(material):
    """Devuelve un precio neto referencial según el nombre del material, o 0 si no hay."""
    nombre = material.lower()
    for clave, precio in PRECIOS_REFERENCIALES:
        if clave in nombre:
            return precio
    return 0


# Mano de obra referencial NETA por unidad de medida de la partida (CLP).
# OJO: son valores de ejemplo. Ajústalos a los de tu zona (igual que los materiales).
# La unidad corresponde a la medida de la partida: hormigones por m³, terminaciones
# por m², cierres/zócalos por ml. La coincidencia es por palabra clave en la partida.
MANO_OBRA_REFERENCIAL = [
    # (palabra_clave_en_minusculas, precio_neto_por_unidad). Más específico primero.
    # --- Moldajes (por m²) — antes que losa/viga/pilar para no confundir ---
    ("moldaje", 7000),
    # --- Hormigones (por m³) ---
    ("emplantillado", 15000),
    ("sobrecimiento", 22000),
    ("cimiento", 20000),
    ("radier", 14000),
    ("losa", 25000),
    ("pilar", 35000),
    ("viga", 32000),
    ("muro horm", 28000),
    ("hormig", 22000),
    # --- Albañilería y tabiques (por m²) ---
    ("ladrillo", 16000),
    ("bloque", 14000),
    ("tabique", 12000),
    ("metalcon", 12000),
    # --- Terminaciones (por m²) ---
    ("estuco", 6500),
    ("pintura", 2800),
    ("cielo", 8000),
    ("revestimiento", 7000),
    ("cerámic", 9000),
    ("ceramic", 9000),
    ("piso", 6500),
    ("pavimento", 6500),
    # --- Cubierta (por m²) ---
    ("cubierta", 9000),
    ("techumbre", 9000),
    # --- Cierres y zócalos (por ml) ---
    ("cierre", 8000),
    ("zócalo", 2000),
    ("zocalo", 2000),
]


def mano_obra_referencial(partida):
    """Precio referencial de mano de obra por unidad de medida según el nombre de la partida."""
    nombre = (partida or "").lower()
    for clave, precio in MANO_OBRA_REFERENCIAL:
        if clave in nombre:
            return precio
    return 0


# Opciones de marca/tipo por material (precios NETOS). El usuario elige una o usa "Otro".
# La coincidencia es por palabra clave en el nombre del material.
OPCIONES_PRECIO = {
    "cemento": [
        ("Polpaico (25kg)", 4881),
        ("Melón (25kg)", 4353),
        ("Bío Bío (25kg)", 4269),
        ("Transex (25kg)", 4101),
    ],
    "gravilla": [
        ("Chancada 3/4\"", 27601),
        ("Fina 3/8\"", 27118),
        ("Mayorista/camionada", 22900),
    ],
    "arena": [
        ("Gruesa (mayorista)", 20250),
        ("Gruesa (detalle)", 29113),
        ("Fina/estuco", 25920),
    ],
    "fierro": [
        ("8mm (barra 6m)", 2500),
        ("10mm (barra 6m)", 3655),
        ("12mm (barra 6m)", 5336),
        ("16mm (barra 6m)", 9664),
    ],
}

def opciones_para(material):
    """Devuelve la lista de opciones (etiqueta, precio) para un material, o None."""
    nombre = material.lower()
    for clave, opciones in OPCIONES_PRECIO.items():
        if clave in nombre:
            return opciones
    return None


# Catálogo de materiales por rubro para el apartado "Mis precios" (Mi cuenta).
# Permite ponerle precio a cada material sin necesidad de una cubicación abierta.
CATALOGO_MATERIALES = {
    "Hormigón": [
        ("Cemento", "saco 25kg"),
        ("Gravilla", "m³"),
        ("Arena", "m³"),
    ],
    "Acero estructural": [
        ("Fierro 8mm", "barra 6m"),
        ("Fierro 10mm", "barra 6m"),
        ("Fierro 12mm", "barra 6m"),
        ("Fierro 16mm", "barra 6m"),
        ("Alambre de amarre", "kg"),
    ],
    "Moldajes": [
        ("Terciado moldaje", "plancha"),
        ("Pino dimensionado", "unidad"),
    ],
    "Muros": [
        ("Ladrillo", "unidad"),
        ("Canales", "barra 3m"),
        ("Montantes normales", "barra 3m"),
        ("Montantes perforados", "barra 3m"),
        ("Diagonales", "barra 3m"),
        ("Tornillos autoperforantes", "caja 1.000"),
        ("Lana de vidrio", "rollo"),
        ("Pino tabique", "unidad 3.2m"),
        ("Clavos", "kg"),
    ],
    "Revestimientos": [
        ("Yeso cartón (Volcanita)", "plancha"),
        ("Tornillos yeso-cartón", "caja 1.000"),
        ("Fibrocemento", "plancha"),
        ("Siding fibrocemento", "tabla"),
        ("Terciado estructural", "plancha"),
        ("Terciado ranurado", "plancha"),
        ("OSB", "plancha"),
        ("Fijaciones", "caja 100u"),
    ],
    "Pisos y Pavimentos": [
        ("Cerámico / Porcelanato", "m²"),
        ("Pegamento", "saco 25kg"),
        ("Fragüe", "bolsa 1kg"),
        ("Piso flotante", "caja"),
        ("Baldosa", "m²"),
        ("Deck de madera", "m²"),
        ("Tornillos galvanizados", "caja 100u"),
    ],
    "Terminaciones": [
        ("Pintura", "galón"),
        ("Sellador", "galón"),
        ("Pasta muro", "tineta"),
        ("Cinta de juntas", "rollo"),
        ("Estuco", "saco"),
        ("Cal", "saco"),
        ("Cielo (plancha)", "plancha"),
        ("Perfiles cielo (Portante 40R)", "barra 3m"),
        ("Zócalos / Guardapolvos", "unidad 2.4m"),
    ],
    "Cubierta / Techumbre": [
        ("Costanera madera", "unidad"),
        ("Madera cerchas", "unidad"),
        ("Perfil Omega 92x50", "barra 6m"),
        ("Perfil Omega 70x40", "barra 6m"),
        ("Costanera metálica", "barra 6m"),
        ("Perfil C (cercha)", "barra 6m"),
        ("Zinc acanalado", "plancha"),
        ("Teja asfáltica", "paquete"),
        ("Teja de arcilla", "unidad"),
        ("Panel sándwich", "m²"),
        ("Fieltro asfáltico", "rollo"),
        ("Aislación térmica", "rollo"),
    ],
}


def aviso_premium(funcion="Esta función"):
    """Muestra un aviso de función de pago para usuarios sin plan Pro."""
    st.warning(f"⭐ {funcion} es parte de los **Planes Pro**.")
    st.caption("Mejora a Plan Pro para desbloquear el presupuesto con precios, PDF con tu marca y más.")


def mostrar_terminos():
    """Muestra el texto de los términos y condiciones de ObraCubic.
    BORRADOR base: debe ser revisado por un abogado antes de uso definitivo."""
    st.markdown("""
## Términos y Condiciones de Uso — ObraCubic

**Última actualización:** revise y ajuste esta fecha.

Al crear una cuenta y utilizar ObraCubic ("la aplicación", "el servicio"), usted acepta los siguientes términos. Le recomendamos leerlos con atención.

### 1. Descripción del servicio
ObraCubic es una herramienta digital que permite calcular cantidades de materiales de construcción (cubicación) y generar presupuestos referenciales. El servicio se ofrece en un plan gratuito y un plan premium con funciones adicionales.

### 2. Naturaleza referencial de los cálculos y precios
Los resultados de cubicación, las cantidades de materiales, los precios y los presupuestos generados por ObraCubic son **estimaciones de carácter referencial**. No constituyen un cálculo estructural, una cotización oficial ni una garantía de cantidades o costos exactos.

Es responsabilidad exclusiva del usuario **verificar** todos los resultados con un profesional competente (arquitecto, ingeniero, constructor) y con proveedores reales antes de tomar decisiones de compra, construcción o contratación. ObraCubic no se hace responsable por diferencias entre los valores estimados y los valores reales de obra.

### 3. Uso aceptable
El usuario se compromete a utilizar la aplicación de forma lícita y a no:
- Usar el servicio para fines ilegales o no autorizados.
- Intentar vulnerar la seguridad de la aplicación o de otros usuarios.
- Compartir su cuenta con terceros de forma que infrinja estos términos.

### 4. Cuenta de usuario
El usuario es responsable de mantener la confidencialidad de su contraseña y de toda actividad realizada desde su cuenta. Los datos que ingrese (proyectos, precios) son de su responsabilidad.

### 5. Planes y pagos
El plan gratuito ofrece funciones básicas de cubicación. El plan premium ofrece funciones adicionales y puede tener un costo. Las condiciones de cobro, renovación y cancelación se informarán al momento de contratar el plan premium.

### 6. Datos personales
ObraCubic recopila los datos necesarios para prestar el servicio (correo, nombre, proyectos guardados). Estos datos se almacenan de forma segura y no se comparten con terceros salvo obligación legal. El usuario puede solicitar la eliminación de su cuenta y datos.

### 7. Limitación de responsabilidad
ObraCubic se ofrece "tal cual", sin garantías de disponibilidad ininterrumpida ni de ausencia de errores. En la máxima medida permitida por la ley, ObraCubic no será responsable por daños directos o indirectos derivados del uso o la imposibilidad de uso del servicio, ni por decisiones tomadas en base a los cálculos referenciales entregados.

### 8. Propiedad intelectual
La aplicación, su diseño, código y contenidos son propiedad de ObraCubic. El usuario conserva la propiedad de los datos que ingrese.

### 9. Modificaciones
ObraCubic puede actualizar estos términos en cualquier momento. Los cambios se informarán a través de la aplicación.

### 10. Contacto
Para consultas sobre estos términos, escriba a: *(indique su correo de contacto)*.

---

*Este documento es un borrador base de referencia y no constituye asesoría legal. Antes de su uso definitivo, especialmente si el servicio tendrá costo, debe ser revisado por un abogado para asegurar su cumplimiento con la legislación chilena vigente (incluyendo la Ley 19.496 de Protección al Consumidor y la Ley 19.628 sobre Protección de la Vida Privada).*
""")


# ============================
# DATOS DE DOSIFICACIÓN (CBB)
# ============================
DOSIFICACIONES = {
    "G-5": {
        "descripcion": "Hormigón de muy baja resistencia",
        "cemento_kg": 170,
        "gravilla_kg": 1025,
        "arena_kg": 910,
        "agua_lt": 195,
    },
    "G-10": {
        "descripcion": "Hormigón de baja resistencia",
        "cemento_kg": 230,
        "gravilla_kg": 1055,
        "arena_kg": 835,
        "agua_lt": 195,
    },
    "G-15": {
        "descripcion": "Emplantillado, sobrecimientos simples",
        "cemento_kg": 275,
        "gravilla_kg": 1070,
        "arena_kg": 800,
        "agua_lt": 195,
    },
    "G-20": {
        "descripcion": "Radier, cimientos normales",
        "cemento_kg": 340,
        "gravilla_kg": 1095,
        "arena_kg": 715,
        "agua_lt": 200,
    },
    "G-25": {
        "descripcion": "Losas estructurales, pilares",
        "cemento_kg": 380,
        "gravilla_kg": 1120,
        "arena_kg": 645,
        "agua_lt": 200,
    },
    "G-30": {
        "descripcion": "Obras especiales, alta resistencia",
        "cemento_kg": 440,
        "gravilla_kg": 1145,
        "arena_kg": 585,
        "agua_lt": 200,
    },
}

# Peso de un saco de cemento (kg). En Chile el saco estándar es de 25 kg.
KG_POR_SACO_CEMENTO = 25

def calcular_materiales(volumen_m3, dosificacion):
    dos = DOSIFICACIONES[dosificacion]
    vol = volumen_m3
    # El cemento se calcula con el valor exacto en kg/m³ y se redondea solo
    # el total de sacos al final (más preciso que redondear sacos por m³).
    return {
        "cemento_sacos": round(vol * dos["cemento_kg"] / KG_POR_SACO_CEMENTO),
        "gravilla_kg":   round(vol * dos["gravilla_kg"]),
        "arena_kg":      round(vol * dos["arena_kg"]),
        "agua_lt":       round(vol * dos["agua_lt"]),
    }

def mostrar_materiales(materiales, rubro=None, partida=None):
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Cemento",   f"{materiales['cemento_sacos']} sacos")
    m2.metric("Gravilla",  f"{materiales['gravilla_kg']} kg")
    m3.metric("Arena",     f"{materiales['arena_kg']} kg")
    m4.metric("Agua",      f"{materiales['agua_lt']} lt")
    # Si se indica la partida, registrar los materiales para PDF y presupuesto.
    # Solo se registra si la partida tiene al menos un material > 0; de lo
    # contrario se elimina del registro (evita partidas "fantasma" en el PDF).
    if rubro and partida:
        tiene_datos = any(
            (materiales.get(k, 0) or 0) > 0
            for k in ("cemento_sacos", "gravilla_kg", "arena_kg", "agua_lt")
        )
        if tiene_datos:
            registrar_pdf(rubro, partida, [
                ("Cemento", f"{materiales['cemento_sacos']} sacos"),
                ("Gravilla", f"{materiales['gravilla_kg']} kg"),
                ("Arena", f"{materiales['arena_kg']} kg"),
                ("Agua", f"{materiales['agua_lt']} lt"),
            ])
        else:
            quitar_pdf(rubro, partida)


def _items_todos_cero(items):
    """Devuelve True si todos los valores numéricos de los items son cero.
    Sirve para detectar partidas vacías que no deben aparecer en el PDF."""
    import re
    encontro_numero = False
    for _etiqueta, valor in items:
        for num in re.findall(r"-?\d+[.,]?\d*", str(valor)):
            encontro_numero = True
            try:
                if float(num.replace(",", ".")) != 0:
                    return False
            except ValueError:
                continue
    # Si no había ningún número, no la consideramos "todo cero" (puede ser texto)
    return encontro_numero


def quitar_pdf(rubro, partida):
    """Elimina una partida del registro persistente (PDF y presupuesto)."""
    clave = f"{rubro}||{partida}"
    st.session_state.setdefault("materiales_persistente", {})
    st.session_state["materiales_persistente"].pop(clave, None)


def registrar_pdf(rubro, partida, items):
    """Registra el resultado final de una partida para el PDF.
    items: lista de tuplas (etiqueta, valor).
    Si todos los valores son cero, no la registra (y la elimina si existía)."""
    # No registrar partidas completamente vacías (todo en cero)
    if _items_todos_cero(items):
        quitar_pdf(rubro, partida)
        return
    st.session_state.setdefault("pdf_extra", [])
    st.session_state["pdf_extra"].append(
        {"rubro": rubro, "partida": partida, "items": items}
    )
    # Acumulador PERSISTENTE por (rubro, partida) para el presupuesto.
    # No se reinicia al cambiar de sección; se sobrescribe solo esa partida.
    st.session_state.setdefault("materiales_persistente", {})
    st.session_state["materiales_persistente"][f"{rubro}||{partida}"] = {
        "rubro": rubro, "partida": partida, "items": items,
    }


def secciones_input(key_prefix, campos, etiqueta="elemento"):
    """Renderiza un grupo de secciones con botones agregar/quitar.
    campos: lista de tuplas (clave, label). Cada sección es un dict con esas claves.
    Devuelve la lista de secciones (cada una un dict). Sirve para cubicar
    elementos con medidas distintas (tramos de cierre, cimientos, vigas, etc.)."""
    lista_key = f"sec_{key_prefix}"
    if lista_key not in st.session_state:
        st.session_state[lista_key] = [{c[0]: 0.0 for c in campos}]

    for i, sec in enumerate(st.session_state[lista_key]):
        cols = st.columns([3] * len(campos) + [1])
        for j, (clave, label) in enumerate(campos):
            with cols[j]:
                sec[clave] = st.number_input(
                    f"{label} {i+1}", value=float(sec.get(clave, 0.0)),
                    min_value=0.0, step=0.1, key=f"{key_prefix}_{clave}_{i}")
        with cols[-1]:
            st.write("")
            st.write("")
            if len(st.session_state[lista_key]) > 1 and st.button("🗑️", key=f"del_{key_prefix}_{i}"):
                st.session_state[lista_key].pop(i)
                st.rerun()

    if st.button(f"➕ Agregar {etiqueta}", key=f"add_{key_prefix}"):
        st.session_state[lista_key].append({c[0]: 0.0 for c in campos})
        st.rerun()

    return st.session_state[lista_key]


PESO_BARRAS = {
    "Ø8mm":  0.395,
    "Ø10mm": 0.617,
    "Ø12mm": 0.888,
    "Ø16mm": 1.578,
    "Ø20mm": 2.466,
    "Ø25mm": 3.854,
}
RATIO_ACERO = {
    "Losa":    {"ratio": 8,   "unidad": "kg/m2"},
    "Viga":    {"ratio": 120, "unidad": "kg/m3"},
    "Pilar":   {"ratio": 150, "unidad": "kg/m3"},
    "Radier":  {"ratio": 5,   "unidad": "kg/m2"},
    "Cimiento":{"ratio": 80,  "unidad": "kg/m3"},
}
# ============================
# MALLA ACMA
# ============================
MALLA_ACMA = {
    "C84  (Ø5,0mm c/15cm)": {"kg_m2": 1.39},
    "C92  (Ø5,5mm c/15cm)": {"kg_m2": 1.54},
    "C128 (Ø6,0mm c/15cm)": {"kg_m2": 2.13},
    "C188 (Ø6,0mm c/10cm)": {"kg_m2": 3.16},
    "C257 (Ø7,0mm c/10cm)": {"kg_m2": 4.27},
}

# ============================
# FUNCIÓN EXPORTAR PDF
# ============================
NARANJA    = colors.HexColor("#FF6B00")
GRIS_OSCURO = colors.HexColor("#1E1E1E")
GRIS_CLARO  = colors.HexColor("#F5F5F5")
BLANCO     = colors.white

LOGO_OBRACUBIC_URL = "https://raw.githubusercontent.com/luissalazarbastias-star/obracubic/refs/heads/main/Foto%202.png"


def _obtener_logo_pdf(datos_usuario, estilo_fallback, ancho=2*cm, alto=2*cm):
    """Devuelve un flowable de logo para el PDF.
    - Si el usuario es Pro y tiene logo personalizado, usa ese.
    - Si no, usa el logo de ObraCubic.
    - Si algo falla (red, formato), devuelve un espacio vacío sin romper el PDF.
    """
    import urllib.request
    from reportlab.platypus import Image as RLImage

    # Determinar qué URL de logo usar
    logo_url = LOGO_OBRACUBIC_URL
    try:
        usa_logo_personalizado = (
            datos_usuario
            and datos_usuario.get("logo_url")
            and "puede_pdf_con_logo" in globals()
            and puede_pdf_con_logo()
        )
        if usa_logo_personalizado:
            logo_url = datos_usuario["logo_url"]
    except Exception:
        logo_url = LOGO_OBRACUBIC_URL

    # Intentar descargar el logo elegido; si falla, intentar el de ObraCubic; si todo falla, vacío
    for url_intento in (logo_url, LOGO_OBRACUBIC_URL):
        try:
            req = urllib.request.Request(url_intento, headers={"User-Agent": "Mozilla/5.0"})
            logo_data = urllib.request.urlopen(req, timeout=8).read()
            return RLImage(io.BytesIO(logo_data), width=ancho, height=alto)
        except Exception:
            continue
    # Si ni el personalizado ni el de ObraCubic cargaron, espacio vacío (no rompe el PDF)
    return Paragraph("", estilo_fallback)


def _filas_datos_profesional(datos_usuario):
    """Devuelve filas (etiqueta, valor) con los datos profesionales del usuario,
    para insertarlas en la tabla de encabezado del PDF. Solo incluye lo que venga."""
    filas = []
    if not datos_usuario:
        return filas
    if datos_usuario.get("empresa"):
        filas.append(["Profesional/Empresa:", datos_usuario["empresa"]])
    if datos_usuario.get("rut"):
        filas.append(["RUT / Matrícula:", datos_usuario["rut"]])
    contacto = []
    if datos_usuario.get("email"):
        contacto.append(datos_usuario["email"])
    if datos_usuario.get("telefono"):
        contacto.append(datos_usuario["telefono"])
    if contacto:
        filas.append(["Contacto:", "  ·  ".join(contacto)])
    return filas


def generar_pdf_cubicacion(
    nombre_proyecto,
    vol_emp, dos_emp, mat_emp,
    vol_pilares, dos_cim, mat_cim,
    vol_sc_neto, dos_sc, mat_sc,
    vol_radier, dos_rad, mat_rad,
    total_hormigon,
    total_sacos, total_gravilla, total_arena, total_agua,
    acero_losa_kg=None,
    acero_pilar_kg=None,
    acero_viga_kg=None,
    acero_radier_kg=None,
    canal_tipo=None, cant_piezas_canal=0, ml_canal=0, largo_canal=0,
    montante_tipo=None, total_montantes=0, largo_montante=0,
    esq_tipo=None, cant_esquinas=0, largo_esq=0,
    pdf_extra=None,
    con_marca_agua=True,
    datos_usuario=None,
):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    styles = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle("Titulo", parent=styles["Title"],
        fontSize=22, textColor=NARANJA, spaceAfter=4)
    estilo_subtitulo = ParagraphStyle("Subtitulo", parent=styles["Normal"],
        fontSize=10, textColor=GRIS_OSCURO, spaceAfter=2)
    estilo_seccion = ParagraphStyle("Seccion", parent=styles["Heading2"],
        fontSize=13, textColor=NARANJA, spaceBefore=14, spaceAfter=6)
    estilo_normal = ParagraphStyle("Normal2", parent=styles["Normal"],
        fontSize=10, textColor=GRIS_OSCURO)
    estilo_pie = ParagraphStyle("Pie", parent=styles["Normal"],
        fontSize=8, textColor=colors.grey, alignment=1)

    story = []
    zona_chile = timezone(timedelta(hours=-4))
    fecha_hoy = datetime.now(zona_chile).strftime("%d/%m/%Y %H:%M")

    # Encabezado con logo arriba a la derecha (logo personalizado si es Pro)
    logo = _obtener_logo_pdf(datos_usuario, estilo_normal)

    # Título del encabezado: nombre de la empresa del usuario si la tiene
    if datos_usuario and datos_usuario.get("empresa") and "puede_pdf_con_logo" in globals() and puede_pdf_con_logo():
        titulo_enc = f"{datos_usuario['empresa']}<br/><font size=9 color='grey'>Generado con ObraCubic</font>"
    else:
        titulo_enc = "ObraCubic<br/><font size=9 color='grey'>Grandes Estructuras se Levantan con Decisiones Precisas</font>"

    encabezado = Table(
        [[
            Paragraph(titulo_enc, estilo_titulo),
            logo
        ]],
        colWidths=[14.5*cm, 2.5*cm]
    )
    encabezado.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    story.append(encabezado)
    story.append(HRFlowable(width="100%", thickness=2, color=NARANJA, spaceAfter=6))

    datos_header = [
        ["Proyecto:", nombre_proyecto or "Sin nombre"],
        ["Fecha:", fecha_hoy],
        ["Volumen Total:", f"{total_hormigon:.2f} m3"],
    ]
    # Agregar datos profesionales del usuario (Pro) si vienen
    datos_header = _filas_datos_profesional(datos_usuario) + datos_header
    tabla_header = Table(datos_header, colWidths=[4*cm, 13*cm])
    tabla_header.setStyle(TableStyle([
        ("TEXTCOLOR", (0, 0), (0, -1), NARANJA),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tabla_header)
    story.append(Spacer(1, 10))

    # Sección Hormigón
    if total_hormigon and total_hormigon > 0:
        story.append(Paragraph("CUBICACION DE HORMIGON", estilo_seccion))

        partidas = [
            ("Emplantillado",      vol_emp,      dos_emp, mat_emp),
            ("Cimiento / Pilares", vol_pilares,  dos_cim, mat_cim),
            ("Sobrecimiento",      vol_sc_neto,  dos_sc,  mat_sc),
            ("Radier",             vol_radier,   dos_rad, mat_rad),
        ]

        for nombre_partida, vol, dos, mat in partidas:
            if vol and vol > 0:
                story.append(Paragraph(f"<b>{nombre_partida}</b>  —  Dosificacion: {dos}", estilo_normal))
                datos_p = [
                    ["Volumen", "Cemento", "Gravilla", "Arena", "Agua"],
                    [
                        f"{vol:.2f} m3",
                        f"{mat['cemento_sacos']} sacos",
                        f"{mat['gravilla_kg']} kg",
                        f"{mat['arena_kg']} kg",
                        f"{mat['agua_lt']} lt",
                    ],
                ]
                tabla_p = Table(datos_p, colWidths=[3.4*cm]*5)
                tabla_p.setStyle(TableStyle([
                    ("BACKGROUND",  (0, 0), (-1, 0), NARANJA),
                    ("TEXTCOLOR",   (0, 0), (-1, 0), BLANCO),
                    ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE",    (0, 0), (-1, -1), 9),
                    ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [GRIS_CLARO, BLANCO]),
                    ("GRID",        (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ("TOPPADDING",  (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
                ]))
                story.append(tabla_p)
                story.append(Spacer(1, 8))

        story.append(HRFlowable(width="100%", thickness=1, color=NARANJA, spaceAfter=6))
        story.append(Paragraph("RESUMEN TOTAL DE MATERIALES", estilo_seccion))

        datos_resumen = [
            ["Volumen Total", "Cemento Total", "Gravilla Total", "Arena Total", "Agua Total"],
            [
                f"{total_hormigon:.2f} m3",
                f"{total_sacos} sacos",
                f"{total_gravilla} kg",
                f"{total_arena} kg",
                f"{total_agua} lt",
            ],
        ]
        tabla_res = Table(datos_resumen, colWidths=[3.4*cm]*5)
        tabla_res.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), GRIS_OSCURO),
            ("TEXTCOLOR",   (0, 0), (-1, 0), BLANCO),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND",  (0, 1), (-1, 1), NARANJA),
            ("TEXTCOLOR",   (0, 1), (-1, 1), BLANCO),
            ("FONTNAME",    (0, 1), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 10),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("TOPPADDING",  (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 7),
        ]))
        story.append(tabla_res)
        story.append(Spacer(1, 12))

    # Sección acero (si hay datos)
    acero_datos = {
        "Losa": acero_losa_kg,
        "Pilar": acero_pilar_kg,
        "Viga": acero_viga_kg,
        "Radier": acero_radier_kg,
    }
    acero_validos = {k: v for k, v in acero_datos.items() if v and v > 0}

    if acero_validos:
        story.append(HRFlowable(width="100%", thickness=1, color=NARANJA, spaceAfter=6))
        story.append(Paragraph("ACERO ESTRUCTURAL", estilo_seccion))
        total_acero = sum(acero_validos.values())
        filas_acero = [["Elemento", "Acero (kg)"]]
        for elem, kg in acero_validos.items():
            filas_acero.append([elem, f"{kg:.1f} kg"])
        filas_acero.append(["TOTAL", f"{total_acero:.1f} kg"])
        tabla_acero = Table(filas_acero, colWidths=[9*cm, 8*cm])
        tabla_acero.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), GRIS_OSCURO),
            ("TEXTCOLOR",   (0, 0), (-1, 0), BLANCO),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND",  (0, -1), (-1, -1), NARANJA),
            ("TEXTCOLOR",   (0, -1), (-1, -1), BLANCO),
            ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 10),
            ("ALIGN",       (1, 0), (1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [GRIS_CLARO, BLANCO]),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("TOPPADDING",  (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
        ]))
        story.append(tabla_acero)
        story.append(Spacer(1, 12))

    # Sección Metalcon
    if cant_piezas_canal > 0 or total_montantes > 0 or cant_esquinas > 0:
        story.append(HRFlowable(width="100%", thickness=1, color=NARANJA, spaceAfter=6))
        story.append(Paragraph("ACERO NO ESTRUCTURAL - TABIQUES METALCON", estilo_seccion))

        filas_metalcon = [["Elemento", "Tipo", "Cantidad"]]
        if cant_piezas_canal > 0:
            filas_metalcon.append(["Canal / Solera", canal_tipo, f"{cant_piezas_canal:.0f} piezas de {largo_canal}m"])
        if total_montantes > 0:
            filas_metalcon.append(["Montante", montante_tipo, f"{total_montantes} piezas de {largo_montante}m"])
        if cant_esquinas > 0:
            filas_metalcon.append(["Esquinero", esq_tipo, f"{cant_esquinas} piezas de {largo_esq}m"])

        tabla_metalcon = Table(filas_metalcon, colWidths=[4*cm, 9*cm, 4*cm])
        tabla_metalcon.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), GRIS_OSCURO),
            ("TEXTCOLOR",   (0, 0), (-1, 0), BLANCO),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [GRIS_CLARO, BLANCO]),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ]))
        story.append(tabla_metalcon)
        story.append(Spacer(1, 12))

    # Secciones adicionales (acero, moldajes, muros, revestimientos, pisos, terminaciones)
    if pdf_extra:
        # Agrupar partidas por rubro respetando el orden de aparición
        orden_rubros = []
        agrupado = {}
        for reg in pdf_extra:
            r = reg["rubro"]
            if r not in agrupado:
                agrupado[r] = []
                orden_rubros.append(r)
            agrupado[r].append(reg)

        for rubro in orden_rubros:
            story.append(HRFlowable(width="100%", thickness=1, color=NARANJA, spaceAfter=6))
            story.append(Paragraph(rubro.upper(), estilo_seccion))
            for reg in agrupado[rubro]:
                story.append(Paragraph(f"<b>{reg['partida']}</b>", estilo_normal))
                filas = [["Concepto", "Cantidad"]]
                for etiqueta, valor in reg["items"]:
                    filas.append([str(etiqueta), str(valor)])
                tabla_extra = Table(filas, colWidths=[8.5*cm, 8.5*cm])
                tabla_extra.setStyle(TableStyle([
                    ("BACKGROUND",  (0, 0), (-1, 0), NARANJA),
                    ("TEXTCOLOR",   (0, 0), (-1, 0), BLANCO),
                    ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE",    (0, 0), (-1, -1), 9),
                    ("ALIGN",       (1, 0), (1, -1), "CENTER"),
                    ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [GRIS_CLARO, BLANCO]),
                    ("GRID",        (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ("TOPPADDING",  (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
                ]))
                story.append(tabla_extra)
                story.append(Spacer(1, 8))
            story.append(Spacer(1, 6))

    # Pie de página
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"Generado por ObraCubic  |  {fecha_hoy}  |  obracubic.streamlit.app",
        estilo_pie
    ))
    story.append(Paragraph(
        "Este documento es una estimacion de cubicacion. Verifique los valores antes de comprar materiales.",
        estilo_pie
    ))

    def _marca_agua(canvas, doc_):
        """Dibuja la marca de agua diagonal 'ObraCubic' en la página."""
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 60)
        try:
            canvas.setFillColor(NARANJA)
            canvas.setFillAlpha(0.08)  # muy tenue
        except Exception:
            canvas.setFillColor(colors.lightgrey)
        canvas.translate(letter[0] / 2, letter[1] / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "ObraCubic")
        canvas.drawCentredString(0, -120, "ObraCubic")
        canvas.drawCentredString(0, 120, "ObraCubic")
        canvas.restoreState()

    if con_marca_agua:
        doc.build(story, onFirstPage=_marca_agua, onLaterPages=_marca_agua)
    else:
        doc.build(story)
    buffer.seek(0)
    return buffer


def generar_excel_presupuesto(nombre_proyecto, datos_pres, cliente=None, datos_usuario=None):
    """Genera el presupuesto en formato Excel (.xlsx). Solo Plan Pro Élite.
    Devuelve un BytesIO listo para descargar."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NARANJA_HEX = "FF6B00"
    GRIS_HEX = "F2F2F2"
    blanco = Font(color="FFFFFF", bold=True)
    negrita = Font(bold=True)
    fill_naranja = PatternFill("solid", fgColor=NARANJA_HEX)
    fill_gris = PatternFill("solid", fgColor=GRIS_HEX)
    centro = Alignment(horizontal="center", vertical="center")
    derecha = Alignment(horizontal="right")
    borde = Border(*[Side(style="thin", color="DDDDDD")] * 4)

    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    du = datos_usuario or {}
    fila = 1
    # Encabezado profesional
    ws.cell(row=fila, column=1, value=du.get("empresa") or "ObraCubic").font = Font(bold=True, size=16, color=NARANJA_HEX)
    fila += 1
    if du.get("rut"):
        ws.cell(row=fila, column=1, value=f"RUT: {du['rut']}"); fila += 1
    contacto = " · ".join([x for x in [du.get("correo"), du.get("telefono")] if x])
    if contacto:
        ws.cell(row=fila, column=1, value=contacto); fila += 1
    ws.cell(row=fila, column=1, value=f"Proyecto: {nombre_proyecto or '-'}").font = negrita
    fila += 1
    if cliente:
        ws.cell(row=fila, column=1, value=f"Cliente: {cliente}"); fila += 1
    from datetime import datetime as _dt
    ws.cell(row=fila, column=1, value=f"Fecha: {_dt.now().strftime('%d/%m/%Y')}")
    fila += 2

    # Cabecera de la tabla
    encabezados = ["Rubro", "Partida", "Material", "Cantidad", "Unidad", "Precio unit. (neto)", "Subtotal (neto)"]
    for col, txt in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=col, value=txt)
        c.font = blanco; c.fill = fill_naranja; c.alignment = centro; c.border = borde
    fila += 1

    # Filas de materiales
    for it in datos_pres.get("items", []):
        valores = [
            it.get("rubro", ""), it.get("partida", ""), it.get("material", ""),
            round(it.get("cantidad", 0), 2), it.get("unidad", ""),
            it.get("precio", 0), it.get("subtotal", 0),
        ]
        for col, v in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=v)
            c.border = borde
            if col in (6, 7):
                c.number_format = '#,##0'; c.alignment = derecha
            elif col == 4:
                c.alignment = derecha
        fila += 1

    fila += 1
    # Totales
    def fila_total(etiqueta, valor, resaltar=False):
        nonlocal fila
        ws.cell(row=fila, column=6, value=etiqueta).font = negrita
        ws.cell(row=fila, column=6).alignment = derecha
        c = ws.cell(row=fila, column=7, value=round(valor))
        c.number_format = '#,##0'; c.alignment = derecha; c.font = negrita
        if resaltar:
            for col in (6, 7):
                ws.cell(row=fila, column=col).fill = fill_naranja
                ws.cell(row=fila, column=col).font = blanco
        else:
            for col in (6, 7):
                ws.cell(row=fila, column=col).fill = fill_gris
        fila += 1

    fila_total("Subtotal materiales", datos_pres.get("subtotal_materiales", 0))
    fila_total("Mano de obra", datos_pres.get("mano_obra", 0))
    fila_total(f"Margen ({datos_pres.get('margen_pct', 0):.0f}%)", datos_pres.get("margen", 0))
    fila_total("Neto (sin IVA)", datos_pres.get("neto", 0))
    fila_total("IVA (19%)", datos_pres.get("iva", 0))
    fila_total("TOTAL CON IVA", datos_pres.get("total", 0), resaltar=True)

    # Ancho de columnas
    anchos = [22, 24, 34, 11, 9, 18, 18]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Nota al pie
    fila += 1
    ws.cell(row=fila, column=1,
            value="Generado con ObraCubic · Estimación referencial. Verifique los valores antes de comprar.").font = Font(italic=True, size=8, color="888888")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generar_pdf_apu(datos_apu, datos_usuario=None):
    """Genera el PDF del Análisis de Precios Unitarios de una partida. Pro Élite.
    Devuelve un BytesIO."""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, HRFlowable)

    NARANJA = colors.HexColor("#FF6B00")
    CARBON = colors.HexColor("#1E1E1E")
    GRIS = colors.HexColor("#F2F2F2")
    VERDE = colors.HexColor("#1E8E3E")

    estilos = getSampleStyleSheet()
    h_titulo = ParagraphStyle("t", parent=estilos["Title"], fontName="Helvetica-Bold",
                              fontSize=20, textColor=NARANJA, alignment=0, spaceAfter=2)
    h_meta = ParagraphStyle("m", parent=estilos["Normal"], fontName="Helvetica",
                            fontSize=9, textColor=CARBON, leading=13)
    h_sec = ParagraphStyle("s", parent=estilos["Heading2"], fontName="Helvetica-Bold",
                           fontSize=12, textColor=CARBON, spaceBefore=10, spaceAfter=4)
    celda = ParagraphStyle("c", parent=estilos["Normal"], fontName="Helvetica", fontSize=8.5, leading=11)
    celda_b = ParagraphStyle("cb", parent=estilos["Normal"], fontName="Helvetica-Bold",
                             fontSize=8.5, textColor=colors.white, leading=11)

    def clp(v):
        try:
            return "$" + f"{int(round(v)):,}".replace(",", ".")
        except Exception:
            return "$0"

    def sup(texto):
        """Convierte ² y ³ a superíndice válido en reportlab."""
        return str(texto).replace("²", "<super>2</super>").replace("³", "<super>3</super>")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm,
                            topMargin=1.6 * cm, bottomMargin=1.6 * cm,
                            title="ObraCubic - APU", author="ObraCubic")
    S = []
    du = datos_usuario or {}
    S.append(Paragraph(du.get("empresa") or "ObraCubic", h_titulo))
    linea = []
    if du.get("rut"):
        linea.append(f"RUT: {du['rut']}")
    if du.get("telefono"):
        linea.append(du["telefono"])
    if du.get("email"):
        linea.append(du["email"])
    if linea:
        S.append(Paragraph("  ·  ".join(linea), h_meta))
    S.append(Paragraph("Análisis de Precios Unitarios (APU)", ParagraphStyle(
        "x", parent=h_meta, fontName="Helvetica-Bold", fontSize=11, textColor=CARBON)))
    from datetime import datetime as _dt
    S.append(Paragraph(
        f"Partida: <b>{datos_apu.get('partida') or '-'}</b>  ·  "
        f"Unidad: <b>{sup(datos_apu.get('unidad') or '-')}</b>  ·  "
        f"Medida: <b>{datos_apu.get('cantidad', 0):.2f}</b>  ·  "
        f"Fecha: {_dt.now().strftime('%d/%m/%Y')}", h_meta))
    S.append(Spacer(1, 4))
    S.append(HRFlowable(width="100%", thickness=2, color=NARANJA, spaceAfter=8))

    def tabla(titulo, filas, encabezados):
        S.append(Paragraph(titulo, h_sec))
        data = [[Paragraph(h, celda_b) for h in encabezados]]
        for f in filas:
            data.append([Paragraph(sup(c), celda) for c in f])
        anchos = [6.6 * cm, 2.2 * cm, 2.4 * cm, 2.7 * cm, 3.0 * cm]
        t = Table(data, colWidths=anchos)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NARANJA),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        S.append(t)

    # Materiales
    filas_mat = [[m.get("Descripción", ""), m.get("Unidad", ""),
                  f"{m.get('Cantidad', 0):.2f}", clp(m.get("Precio unitario", 0)),
                  clp(m.get("Cantidad", 0) * m.get("Precio unitario", 0))]
                 for m in datos_apu.get("materiales", []) if m.get("Descripción")]
    tabla("1. Materiales", filas_mat, ["Descripción", "Unidad", "Cantidad", "P. unitario", "Subtotal"])
    S.append(Paragraph(f"Subtotal materiales: <b>{clp(datos_apu.get('sub_mat', 0))}</b>", h_meta))

    # Mano de obra
    filas_mo = [[m.get("Descripción", ""), "", f"{m.get('Cantidad', 0):.2f}",
                 clp(m.get("Precio unitario", 0)),
                 clp(m.get("Cantidad", 0) * m.get("Precio unitario", 0))]
                for m in datos_apu.get("mano_obra", []) if m.get("Descripción")]
    tabla("2. Mano de obra", filas_mo, ["Descripción", "", "Cantidad", "P. unitario", "Subtotal"])
    S.append(Paragraph(f"Subtotal mano de obra: <b>{clp(datos_apu.get('sub_mo', 0))}</b>", h_meta))

    # Recargos y totales
    S.append(Paragraph("3. Recargos y totales", h_sec))
    res = [
        ("Leyes sociales", clp(datos_apu.get("leyes", 0))),
        ("Herramientas", clp(datos_apu.get("herr", 0))),
        ("Costo directo", clp(datos_apu.get("costo_directo", 0))),
        ("Gastos grales. + utilidad", clp(datos_apu.get("ggu", 0))),
    ]
    tr = Table([[Paragraph(a, celda), Paragraph(b, ParagraphStyle("r", parent=celda, alignment=2))]
                for a, b in res], colWidths=[12.9 * cm, 4.0 * cm])
    tr.setStyle(TableStyle([("LINEBELOW", (0, 0), (-1, -2), 0.4, colors.HexColor("#DDDDDD")),
                            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    S.append(tr)

    # Total y precio unitario
    S.append(Spacer(1, 6))
    total_box = Table([
        [Paragraph("Total de la partida", ParagraphStyle("tb", parent=celda, textColor=colors.white, fontName="Helvetica-Bold", fontSize=11)),
         Paragraph(clp(datos_apu.get("total", 0)), ParagraphStyle("tv", parent=celda, textColor=colors.white, fontName="Helvetica-Bold", fontSize=13, alignment=2))],
        [Paragraph(sup(f"Precio unitario ({datos_apu.get('unidad') or 'unidad'})"), ParagraphStyle("pb", parent=celda, textColor=colors.white, fontName="Helvetica-Bold", fontSize=11)),
         Paragraph(clp(datos_apu.get("precio_unitario", 0)), ParagraphStyle("pv", parent=celda, textColor=colors.white, fontName="Helvetica-Bold", fontSize=13, alignment=2))],
    ], colWidths=[12.9 * cm, 4.0 * cm])
    total_box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), VERDE),
        ("BACKGROUND", (0, 1), (-1, 1), NARANJA),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    S.append(total_box)

    S.append(Spacer(1, 10))
    S.append(Paragraph("Valores netos (sin IVA). Estimación referencial generada con ObraCubic.",
                       ParagraphStyle("pie", parent=h_meta, fontSize=7.5, textColor=colors.HexColor("#888888"))))
    doc.build(S)
    buf.seek(0)
    return buf


def generar_pdf_presupuesto(nombre_proyecto, datos_pres, cliente=None, datos_usuario=None):
    """Genera un PDF del presupuesto con desglose, IVA y total."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm,
    )
    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("TituloP", parent=styles["Title"],
        fontSize=22, textColor=NARANJA, spaceAfter=4)
    estilo_seccion = ParagraphStyle("SeccionP", parent=styles["Heading2"],
        fontSize=13, textColor=NARANJA, spaceBefore=14, spaceAfter=6)
    estilo_normal = ParagraphStyle("NormalP", parent=styles["Normal"],
        fontSize=10, textColor=GRIS_OSCURO)
    estilo_pie = ParagraphStyle("PieP", parent=styles["Normal"],
        fontSize=8, textColor=colors.grey, alignment=1)

    def _clp(v):
        try:
            return "$" + f"{int(round(v)):,}".replace(",", ".")
        except Exception:
            return "$0"

    story = []
    zona_chile = timezone(timedelta(hours=-4))
    fecha_hoy = datetime.now(zona_chile).strftime("%d/%m/%Y %H:%M")

    # Encabezado con logo (personalizado si es Pro)
    logo = _obtener_logo_pdf(datos_usuario, estilo_normal)

    if datos_usuario and datos_usuario.get("empresa") and "puede_pdf_con_logo" in globals() and puede_pdf_con_logo():
        titulo_enc = f"{datos_usuario['empresa']}<br/><font size=9 color='grey'>Presupuesto de Obra · ObraCubic</font>"
    else:
        titulo_enc = "ObraCubic<br/><font size=9 color='grey'>Presupuesto de Obra</font>"

    encabezado = Table(
        [[
            Paragraph(titulo_enc, estilo_titulo),
            logo
        ]],
        colWidths=[14.5*cm, 2.5*cm]
    )
    encabezado.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    story.append(encabezado)
    story.append(HRFlowable(width="100%", thickness=2, color=NARANJA, spaceAfter=6))

    # Datos del proyecto (incluye datos profesionales del usuario si vienen)
    filas_header = _filas_datos_profesional(datos_usuario) + [
        ["Proyecto:", nombre_proyecto or "Sin nombre"],
        ["Fecha:", fecha_hoy],
    ]
    if cliente:
        filas_header.append(["Cliente:", cliente])
    tabla_header = Table(filas_header, colWidths=[4*cm, 13*cm])
    tabla_header.setStyle(TableStyle([
        ("TEXTCOLOR", (0, 0), (0, -1), NARANJA),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tabla_header)

    # Tabla de materiales
    story.append(Paragraph("Materiales", estilo_seccion))
    filas = [["Material", "Cantidad", "Precio unit.", "Subtotal"]]
    for it in datos_pres.get("items", []):
        unidad = it.get("unidad", "")
        unidad_pdf = unidad.replace("m³", "m3").replace("m²", "m2")
        cant = it.get("cantidad", 0)
        cant_str = f"{cant:.2f} {unidad_pdf}" if unidad in ("m³", "kg") else f"{cant:.0f} {unidad_pdf}"
        filas.append([
            it.get("material", ""),
            cant_str,
            _clp(it.get("precio", 0)),
            _clp(it.get("subtotal", 0)),
        ])
    tabla_mat = Table(filas, colWidths=[7*cm, 4*cm, 3*cm, 3*cm])
    tabla_mat.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NARANJA),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF3E9")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(tabla_mat)

    # Totales
    story.append(Paragraph("Resumen", estilo_seccion))
    filas_tot = [
        ["Subtotal materiales", _clp(datos_pres.get("subtotal_materiales", 0))],
        ["Mano de obra", _clp(datos_pres.get("mano_obra", 0))],
        [f"Margen de ganancia ({datos_pres.get('margen_pct', 0):.0f}%)", _clp(datos_pres.get("margen", 0))],
        ["Neto (sin IVA)", _clp(datos_pres.get("neto", 0))],
        ["IVA (19%)", _clp(datos_pres.get("iva", 0))],
    ]
    tabla_tot = Table(filas_tot, colWidths=[13*cm, 4*cm])
    tabla_tot.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEABOVE", (0, 3), (-1, 3), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tabla_tot)

    # Total con IVA destacado
    tabla_total = Table(
        [["TOTAL CON IVA", _clp(datos_pres.get("total", 0))]],
        colWidths=[13*cm, 4*cm]
    )
    tabla_total.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), NARANJA),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(Spacer(1, 6))
    story.append(tabla_total)

    story.append(Spacer(1, 14))
    story.append(Paragraph(
        "Los precios son referenciales y definidos por el usuario. "
        "Valores netos sin IVA; el IVA se aplica sobre el neto total. "
        "Este presupuesto es una estimación y debe verificarse con proveedores reales.",
        estilo_pie
    ))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"Generado por ObraCubic — {fecha_hoy}",
        estilo_pie
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer


if "ir_a_cubicacion" not in st.session_state:
    st.session_state["ir_a_cubicacion"] = False
if "proyecto_creado" not in st.session_state:
    st.session_state["proyecto_creado"] = False
if "proyecto" not in st.session_state:
    st.session_state["proyecto"] = {}
if "mostrar_proyecto" not in st.session_state:
    st.session_state["mostrar_proyecto"] = False
if "vol_emp" not in st.session_state:
    st.session_state["vol_emp"] = 0.0
if "vol_pilares" not in st.session_state:
    st.session_state["vol_pilares"] = 0.0
if "vol_sc_neto" not in st.session_state:
    st.session_state["vol_sc_neto"] = 0.0
if "vol_radier" not in st.session_state:
    st.session_state["vol_radier"] = 0.0
if "mat_emp" not in st.session_state:
    st.session_state["mat_emp"] = {"cemento_sacos": 0, "gravilla_kg": 0, "arena_kg": 0, "agua_lt": 0}
if "mat_cim" not in st.session_state:
    st.session_state["mat_cim"] = {"cemento_sacos": 0, "gravilla_kg": 0, "arena_kg": 0, "agua_lt": 0}
if "mat_sc" not in st.session_state:
    st.session_state["mat_sc"] = {"cemento_sacos": 0, "gravilla_kg": 0, "arena_kg": 0, "agua_lt": 0}
if "mat_rad" not in st.session_state:
    st.session_state["mat_rad"] = {"cemento_sacos": 0, "gravilla_kg": 0, "arena_kg": 0, "agua_lt": 0}
if "total_hormigon" not in st.session_state:
    st.session_state["total_hormigon"] = st.session_state.get("vol_emp", 0) + st.session_state.get("vol_pilares", 0) + st.session_state.get("vol_sc_neto", 0) + st.session_state.get("vol_radier", 0)
if "emp_perdida" not in st.session_state:
    st.session_state["emp_perdida"] = 5
if "radier_perdida" not in st.session_state:
    st.session_state["radier_perdida"] = 5
if "dos_emp" not in st.session_state:
    st.session_state["dos_emp"] = "G-15"
if "dos_rad" not in st.session_state:
    st.session_state["dos_rad"] = "G-20"
if "dos_cim" not in st.session_state:
    st.session_state["dos_cim"] = "G-20"
if "dos_sc" not in st.session_state:
    st.session_state["dos_sc"] = "G-20"
if "secciones_rad" not in st.session_state:
    st.session_state["secciones_rad"] = [{"largo": 0.0, "ancho": 0.0, "espesor": 0.0}]
if "secciones_emp" not in st.session_state:
    st.session_state["secciones_emp"] = [{"largo": 0.0, "ancho": 0.0, "espesor": 0.0}]
# ============================
# LOGO Y NAVEGACIÓN
# ============================
URL_DEL_LOGO = "https://raw.githubusercontent.com/luissalazarbastias-star/obracubic/refs/heads/main/Foto%201.png"

# Logo centrado en el área principal (se ve en teléfono y computador)
_lc1, _lc2, _lc3 = st.columns([2, 1, 2])
with _lc2:
    st.image(URL_DEL_LOGO, use_container_width=True)

# Selector de navegación SIEMPRE visible (no en el sidebar)
if "nav_option" not in st.session_state:
    st.session_state["nav_option"] = "Cubicacion" if st.session_state.get("ir_a_cubicacion") else "Inicio"

# Cambio de sección solicitado por un botón (se aplica ANTES de crear el radio)
if st.session_state.get("_goto"):
    st.session_state["nav_option"] = st.session_state.pop("_goto")

def _salir_cuenta():
    # Al tocar el menú, salir de la vista de cuenta
    st.session_state["vista_cuenta"] = False

nav_col, cuenta_col = st.columns([3, 1])
with nav_col:
    if st.session_state.get("usuario"):
        # Con cuenta: menú completo, pero Presupuesto solo para planes Pro
        if puede_presupuesto():
            opciones_menu = ["Inicio", "Crear Proyecto", "Cubicacion", "Presupuesto", "Planes"]
        else:
            opciones_menu = ["Inicio", "Crear Proyecto", "Cubicacion", "Planes"]
        # Bitácora de obra: solo Plan Pro Élite
        if puede_bitacora():
            opciones_menu.insert(-1, "Bitácora")
        # APU (presupuesto avanzado): solo Plan Pro Élite
        if puede_apu():
            opciones_menu.insert(-1, "APU")
    else:
        # Usuario sin cuenta: acceso limitado
        opciones_menu = ["Inicio", "Cubicacion", "Planes"]
    # Si la opción guardada ya no está disponible, volver a Inicio
    if st.session_state.get("nav_option") not in opciones_menu:
        st.session_state["nav_option"] = "Inicio"
    option = st.radio(
        "Ir a:",
        opciones_menu,
        horizontal=True,
        key="nav_option",
        on_change=_salir_cuenta,
    )
with cuenta_col:
    st.write("")
    label_cuenta = "👤 Mi cuenta" if st.session_state.get("usuario") else "👤 Iniciar sesión"
    if st.button(label_cuenta, type="primary", use_container_width=True, key="btn_cuenta"):
        st.session_state["vista_cuenta"] = True
        st.rerun()

if option == "Cubicacion":
    st.session_state["ir_a_cubicacion"] = False

st.write("---")

# ============================
# VISTA CUENTA: INICIAR SESIÓN / REGISTRO  (REAL con Supabase)
# ============================
if st.session_state.get("vista_cuenta"):
    if supabase is None:
        st.error("No hay conexión con la base de datos. Intenta más tarde.")
        if st.button("← Volver a la app", use_container_width=True):
            st.session_state["vista_cuenta"] = False
            st.rerun()
        st.stop()

    usuario = st.session_state.get("usuario")

    if usuario:
        # Refrescar plan desde la base (aplica vencimiento automático)
        refrescar_plan_usuario()
        # --- Usuario con sesión iniciada ---
        st.subheader("Mi cuenta")
        nombre_u = st.session_state.get("usuario_nombre") or usuario.get("email", "usuario")
        st.success(f"Sesión iniciada como: {nombre_u}")
        plan_u = st.session_state.get("usuario_plan", "gratis")
        _p = plan_actual()
        _info = PLANES_INFO.get(_p, {"nombre": "Plan Gratis", "emoji": "☕"})
        st.caption(f"Plan actual: **{_info['nombre']} {_info['emoji']}**")
        # Mostrar vencimiento si el plan es de pago y tiene fecha
        _vence = st.session_state.get("usuario_plan_vence")
        if _p in ("pro_basico", "pro_elite") and _vence:
            try:
                from datetime import datetime, timezone
                _fv = datetime.fromisoformat(str(_vence).replace("Z", "+00:00"))
                _dias = (_fv - datetime.now(timezone.utc)).days
                if _dias >= 0:
                    st.caption(f"⏳ Tu plan vence el **{_fv.strftime('%d/%m/%Y')}** (en {_dias} días).")
            except Exception:
                pass
        cc1, cc2 = st.columns(2)
        with cc1:
            if st.button("Cerrar sesión", use_container_width=True):
                try:
                    supabase.auth.sign_out()
                except Exception:
                    pass
                for k in ["usuario", "usuario_nombre", "usuario_plan",
                          "precios_usuario_cargados", "precios_usuario_dict", "_precios_actuales",
                          "usuario_plan_vence", "usuario_empresa", "usuario_rut",
                          "usuario_telefono", "usuario_logo_url"]:
                    st.session_state.pop(k, None)
                st.rerun()
        with cc2:
            if st.button("← Volver a la app", type="primary", use_container_width=True):
                st.session_state["vista_cuenta"] = False
                st.rerun()

        # ---------------------------------------
        # MIS PROYECTOS
        # ---------------------------------------
        st.write("---")
        st.subheader("📁 Mis proyectos")

        proyectos = listar_proyectos(usuario["id"])
        es_premium = puede_cubicaciones_ilimitadas()
        limite = None if es_premium else LIMITE_GRATIS
        n_proy = len(proyectos)

        if limite is None:
            st.caption(f"{n_proy} proyectos guardados · Plan PREMIUM (sin límite)")
        else:
            st.caption(f"{n_proy} / {limite} proyectos guardados")

        # Mensaje de carga exitosa
        if st.session_state.pop("_proyecto_cargado", False):
            st.success("Proyecto cargado. Tus rubros y medidas se restauraron — revísalo en 'Crear Proyecto' o 'Cubicacion'.")

        # --- Guardar la cubicación actual ---
        with st.container(border=True):
            st.markdown("**Guardar proyecto actual**")

            proy_actual = st.session_state.get("proyecto", {})
            rubros_proy = proy_actual.get("partidas", {})
            rubros_activos = [r for r, sub in rubros_proy.items() if sub and any(sub.values())]

            if proy_actual.get("nombre"):
                st.caption(f"Proyecto en curso: **{proy_actual['nombre']}**")
                if rubros_activos:
                    NOMBRES_RUBRO = {
                        "hormigon": "Hormigón", "acero_estructural": "Acero estructural",
                        "metalcon": "Metalcon", "moldajes": "Moldajes", "muros": "Muros",
                        "revestimientos": "Revestimientos", "pisos": "Pisos",
                        "terminaciones": "Terminaciones", "cubierta": "Cubierta",
                    }
                    lista = ", ".join(NOMBRES_RUBRO.get(r, r) for r in rubros_activos)
                    st.caption(f"Rubros: {lista}")
            else:
                st.caption("Sin proyecto creado — se guardará la cubicación general.")

            # Pre-llenar el nombre con el del proyecto creado (solo la primera vez)
            if "nombre_nuevo_proyecto_guardar" not in st.session_state:
                st.session_state["nombre_nuevo_proyecto_guardar"] = proy_actual.get("nombre", "")

            nombre_guardar = st.text_input(
                "Nombre del proyecto",
                placeholder="Ej: Casa Don Pedro - Angol",
                key="nombre_nuevo_proyecto_guardar",
            )
            puede_guardar = (limite is None) or (n_proy < limite)
            if puede_guardar:
                if st.button("💾 Guardar proyecto", type="primary", use_container_width=True):
                    if not nombre_guardar.strip():
                        st.warning("Ponle un nombre al proyecto.")
                    else:
                        try:
                            datos = capturar_cubicacion()
                            guardar_proyecto(usuario["id"], nombre_guardar.strip(), datos)
                            st.success("¡Proyecto guardado!")
                            st.rerun()
                        except Exception:
                            st.error("No se pudo guardar el proyecto. Intenta de nuevo.")
            else:
                st.warning(
                    f"Llegaste al límite de {limite} proyectos del plan gratis. "
                    "Elimina uno o pásate a Premium para guardar más."
                )

        # --- Lista de proyectos guardados ---
        if proyectos:
            st.write("")
            for p in proyectos:
                fecha = (p.get("actualizado_en") or "")[:10]
                col_n, col_c, col_e = st.columns([3, 1, 1])
                with col_n:
                    st.markdown(f"**{p['nombre']}**")
                    if fecha:
                        st.caption(f"Actualizado: {fecha}")
                with col_c:
                    if st.button("Cargar", key=f"cargar_{p['id']}", use_container_width=True):
                        try:
                            datos = cargar_proyecto(p["id"])
                            for k, v in datos.items():
                                st.session_state[k] = v
                            st.session_state["_proyecto_cargado"] = True
                            st.session_state["_goto"] = "Cubicacion"
                            st.session_state["vista_cuenta"] = False
                            st.rerun()
                        except Exception:
                            st.error("No se pudo cargar el proyecto.")
                with col_e:
                    if st.button("Eliminar", key=f"elim_{p['id']}", use_container_width=True):
                        try:
                            eliminar_proyecto(p["id"])
                            st.rerun()
                        except Exception:
                            st.error("No se pudo eliminar.")
        else:
            st.info("Todavía no tienes proyectos guardados.")

        # ---------------------------------------
        # MIS DATOS PROFESIONALES (solo Pro, para el PDF)
        # ---------------------------------------
        if puede_pdf_con_logo():
            st.write("---")
            st.subheader("🏢 Mis datos profesionales")
            st.caption("Estos datos aparecerán en el encabezado de tus PDF (cubicación y presupuesto). "
                       "Déjalos en blanco si no quieres mostrarlos.")

            dp1, dp2 = st.columns(2)
            with dp1:
                in_empresa = st.text_input(
                    "Nombre o empresa",
                    value=st.session_state.get("usuario_empresa") or "",
                    placeholder="Ej: Constructora Salazar Ltda.",
                    key="in_empresa",
                )
                in_rut = st.text_input(
                    "RUT o matrícula (opcional)",
                    value=st.session_state.get("usuario_rut") or "",
                    placeholder="Ej: 76.123.456-7",
                    key="in_rut",
                )
            with dp2:
                in_telefono = st.text_input(
                    "Teléfono de contacto (opcional)",
                    value=st.session_state.get("usuario_telefono") or "",
                    placeholder="Ej: +56 9 1234 5678",
                    key="in_telefono",
                )
                st.text_input(
                    "Correo (de tu cuenta)",
                    value=usuario.get("email", ""),
                    disabled=True,
                    key="in_correo_pro",
                    help="Es el correo de tu cuenta. Aparecerá en el PDF.",
                )

            if st.button("💾 Guardar mis datos profesionales", type="primary", key="btn_guardar_datos_pro"):
                try:
                    guardar_datos_profesional(usuario["id"], in_empresa.strip(), in_rut.strip(), in_telefono.strip())
                    st.session_state["usuario_empresa"] = in_empresa.strip() or None
                    st.session_state["usuario_rut"] = in_rut.strip() or None
                    st.session_state["usuario_telefono"] = in_telefono.strip() or None
                    st.success("¡Datos guardados! Aparecerán en tus próximos PDF.")
                except Exception:
                    st.error("No se pudieron guardar los datos. Intenta de nuevo.")

        # ---------------------------------------
        # MIS PRECIOS (solo planes de pago)
        # ---------------------------------------
        if puede_presupuesto():
            st.write("---")
            st.subheader("💲 Mis precios")
            st.caption("Ajusta el precio de cada material (neto, sin IVA) y guarda tu lista. "
                       "Se aplicará automáticamente en tus presupuestos. Los valores que ves son referenciales.")

            # Cargar precios guardados del usuario
            mis_precios = cargar_precios_usuario(usuario["id"])
            nuevos_precios = {}

            for rubro_cat, materiales_cat in CATALOGO_MATERIALES.items():
                with st.expander(f"📦 {rubro_cat}", expanded=False):
                    for material, unidad in materiales_cat:
                        # Prioridad: precio guardado del usuario > referencial
                        valor_inicial = mis_precios.get(material, precio_referencial(material))
                        col_m, col_p = st.columns([3, 2])
                        with col_m:
                            st.markdown(f"**{material}**")
                            st.caption(f"por {unidad}")
                        with col_p:
                            p = st.number_input(
                                f"Precio {material}",
                                min_value=0, value=int(valor_inicial), step=100,
                                key=f"miprecio_{rubro_cat}_{material}".replace(" ", "_"),
                                label_visibility="collapsed",
                            )
                        nuevos_precios[material] = p

            st.write("")
            if st.button("💾 Guardar mi lista de precios", type="primary", use_container_width=True, key="btn_guardar_mis_precios"):
                try:
                    guardar_precios_usuario(usuario["id"], nuevos_precios)
                    # Refrescar el dict en memoria para que el presupuesto los use
                    st.session_state["precios_usuario_dict"] = cargar_precios_usuario(usuario["id"])
                    st.success("¡Tu lista de precios fue guardada! Se aplicará en tus próximos presupuestos.")
                except Exception:
                    st.error("No se pudieron guardar los precios. Intenta de nuevo.")

        # Términos y condiciones (siempre accesibles con sesión)
        st.write("---")
        with st.expander("📄 Términos y condiciones"):
            mostrar_terminos()

    else:
        # --- Sin sesión: login / registro ---
        st.subheader("Acceso a ObraCubic")

        tab_login, tab_registro = st.tabs(["Iniciar sesión", "Crear cuenta"])

        with tab_login:
            email_l = st.text_input("Correo electrónico", placeholder="tucorreo@ejemplo.com", key="login_email")
            pass_l = st.text_input("Contraseña", type="password", placeholder="••••••••", key="login_pass")
            if st.button("Iniciar sesión", type="primary", use_container_width=True, key="btn_login"):
                if not email_l or not pass_l:
                    st.warning("Ingresa tu correo y contraseña.")
                else:
                    try:
                        res = supabase.auth.sign_in_with_password({"email": email_l, "password": pass_l})
                        st.session_state["usuario"] = {"id": res.user.id, "email": res.user.email}
                        # Cargar perfil (nombre, plan y vencimiento)
                        try:
                            perfil = supabase.table("perfiles").select("nombre, plan, plan_vence, empresa, rut, telefono").eq("id", res.user.id).single().execute()
                            st.session_state["usuario_nombre"] = perfil.data.get("nombre")
                            plan_bd = perfil.data.get("plan", "gratis")
                            vence = perfil.data.get("plan_vence")
                            st.session_state["usuario_plan_vence"] = vence
                            # Si el plan de pago venció, tratarlo como gratis
                            st.session_state["usuario_plan"] = _plan_vigente(plan_bd, vence)
                            # Datos profesionales (para PDF Pro)
                            st.session_state["usuario_empresa"] = perfil.data.get("empresa")
                            st.session_state["usuario_rut"] = perfil.data.get("rut")
                            st.session_state["usuario_telefono"] = perfil.data.get("telefono")
                        except Exception:
                            st.session_state["usuario_plan"] = "gratis"
                            st.session_state["usuario_plan_vence"] = None
                        st.session_state["vista_cuenta"] = False
                        st.session_state["_goto"] = "Inicio"
                        st.rerun()
                    except Exception:
                        st.error("Correo o contraseña incorrectos.")

        with tab_registro:
            nombre_r = st.text_input("Nombre completo", placeholder="Juan Pérez", key="reg_nombre")
            email_r = st.text_input("Correo electrónico", placeholder="tucorreo@ejemplo.com", key="reg_email")
            pass_r = st.text_input("Contraseña", type="password", placeholder="Mínimo 6 caracteres", key="reg_pass")
            pass_r2 = st.text_input("Repetir contraseña", type="password", placeholder="Repite la contraseña", key="reg_pass2")
            st.markdown("Para crear tu cuenta, abre y lee los términos y condiciones. Al final podrás aceptarlos.")
            with st.expander("📄 Leer términos y condiciones"):
                mostrar_terminos()
                st.write("---")
                acepta = st.checkbox("He leído y acepto los términos y condiciones", key="reg_terminos")
            if not st.session_state.get("reg_terminos"):
                st.caption("☝️ Debes abrir los términos y aceptarlos al final para crear tu cuenta.")
            acepta = st.session_state.get("reg_terminos", False)
            if st.button("Crear cuenta", type="primary", use_container_width=True, key="btn_registro"):
                if not nombre_r or not email_r or not pass_r:
                    st.warning("Completa nombre, correo y contraseña.")
                elif len(pass_r) < 6:
                    st.warning("La contraseña debe tener al menos 6 caracteres.")
                elif pass_r != pass_r2:
                    st.warning("Las contraseñas no coinciden.")
                elif not acepta:
                    st.warning("Debes aceptar los términos y condiciones.")
                else:
                    try:
                        supabase.auth.sign_up({
                            "email": email_r,
                            "password": pass_r,
                            "options": {"data": {"nombre": nombre_r}},
                        })
                        st.success("¡Cuenta creada! Revisa tu correo para confirmar la cuenta y luego inicia sesión.")
                    except Exception as e:
                        msg = str(e)
                        if "already" in msg.lower() or "registered" in msg.lower():
                            st.error("Ya existe una cuenta con ese correo.")
                        else:
                            st.error("No se pudo crear la cuenta. Verifica los datos e intenta de nuevo.")

        st.write("")
        if st.button("← Volver a la app", use_container_width=True, key="btn_volver_login"):
            st.session_state["vista_cuenta"] = False
            st.rerun()

    st.stop()  # No mostrar el resto de la app mientras se ve la cuenta

# ============================
# INICIO (pantalla de bienvenida)
# ============================
if option == "Inicio":
    _tiene_sesion = bool(st.session_state.get("usuario"))
    _nombre = st.session_state.get("usuario_nombre", "")

    if _tiene_sesion and _nombre:
        st.title(f"¡Hola, {_nombre}! 👋")
    else:
        st.title("Bienvenido a ObraCubic")
    st.markdown("#### Grandes estructuras se levantan con decisiones precisas")
    st.write("")
    st.markdown(
        "**ObraCubic** es tu herramienta para **cubicar materiales** y **armar presupuestos** "
        "de obras de construcción. Calcula de forma rápida y ordenada lo que necesitas en cada "
        "etapa, ponle precio y genera informes profesionales en PDF, listos para tu cliente."
    )

    st.write("---")
    st.subheader("¿Qué puedes hacer?")
    ini1, ini2 = st.columns(2)
    with ini1:
        st.markdown("##### 📐 Cubicar")
        st.markdown(
            "Estima materiales de muchos rubros:\n"
            "- **Hormigón** — excavación, cimientos, radier\n"
            "- **Acero** estructural y **Metalcon**\n"
            "- **Muros** — hormigón, ladrillo, tabiques\n"
            "- **Revestimientos**, **Pisos** y **Terminaciones**\n"
            "- **Cubierta** — cerchas y planchas"
        )
    with ini2:
        st.markdown("##### 💰 Presupuestar")
        st.markdown(
            "Convierte tu cubicación en un presupuesto:\n"
            "- Precios **referenciales** editables por material\n"
            "- **Mano de obra**, margen e **IVA**\n"
            "- Tu **lista personal de precios** guardada\n"
            "- **PDF profesional** del presupuesto\n"
            "- *(funciones de los Planes Pro)*"
        )

    st.write("---")

    if not _tiene_sesion:
        # Usuario sin cuenta
        st.subheader("Empieza ahora")
        st.markdown(
            "Estás usando ObraCubic **sin cuenta**. Puedes probar la cubicación de **radier, "
            "tabiques y revestimientos**. Crea una cuenta **gratis** para desbloquear todas las "
            "partidas, guardar tus proyectos y más."
        )
        ccta1, ccta2 = st.columns(2)
        with ccta1:
            if st.button("👤 Crear cuenta gratis", type="primary", use_container_width=True):
                st.session_state["vista_cuenta"] = True
                st.rerun()
        with ccta2:
            if st.button("📐 Probar cubicación", use_container_width=True):
                st.session_state["_goto"] = "Cubicacion"
                st.rerun()
        st.write("")
        st.caption("💎 Mira todo lo que incluye cada plan en la sección **Planes**.")
    else:
        # Usuario con cuenta
        st.subheader("¿Cómo empezar?")
        st.markdown(
            "1. Ve a **Crear Proyecto** y selecciona los rubros de tu obra.\n"
            "2. Pasa a **Cubicacion** e ingresa las medidas de cada partida.\n"
            "3. Genera tu **PDF** de cubicación.\n"
            "4. Si tienes Plan Pro, ve a **Presupuesto** para ponerle precio y obtener el total con IVA."
        )
        st.info("También puedes usar la **Cubicacion** directamente, sin crear un proyecto.")

        if puede_presupuesto():
            cini1, cini2, cini3 = st.columns(3)
            with cini1:
                if st.button("🛠️ Crear Proyecto", type="primary", use_container_width=True):
                    st.session_state["_goto"] = "Crear Proyecto"
                    st.rerun()
            with cini2:
                if st.button("📐 Ir a Cubicación", use_container_width=True):
                    st.session_state["_goto"] = "Cubicacion"
                    st.rerun()
            with cini3:
                if st.button("💰 Ir a Presupuesto", use_container_width=True):
                    st.session_state["_goto"] = "Presupuesto"
                    st.rerun()
        else:
            cini1, cini2 = st.columns(2)
            with cini1:
                if st.button("🛠️ Crear Proyecto", type="primary", use_container_width=True):
                    st.session_state["_goto"] = "Crear Proyecto"
                    st.rerun()
            with cini2:
                if st.button("📐 Ir a Cubicación", use_container_width=True):
                    st.session_state["_goto"] = "Cubicacion"
                    st.rerun()
            st.write("")
            st.caption("💎 ¿Quieres armar presupuestos con precios e IVA? Mira los **Planes Pro**.")

# ============================
# CREAR PROYECTO
# ============================
if option == "Crear Proyecto":
    st.subheader("Crear Nuevo Proyecto")
    st.caption("Configurá tu obra y seleccioná solo los rubros que vas a trabajar.")

    with st.container(border=True):
        nombre_proy = st.text_input("Nombre del proyecto *",
            placeholder="Ej: Casa Don Pedro - Angol",
            key="input_nombre_proy")

        col1, col2 = st.columns(2)
        with col1:
            tipo_obra = st.text_input("Tipo de obra",
                placeholder="Ej: Vivienda Unifamiliar",
                key="input_tipo_obra")
        with col2:
            profesional = st.text_input("Profesional / Empresa (opcional)",
                placeholder="Ej: Juan Pérez - Constructor",
                key="input_profesional")

        st.write("**Seleccioná los rubros de tu obra:**")
        col1, col2, col3 = st.columns(3)
        with col1:
            usar_hormigon    = st.checkbox("Hormigón y Mov. de tierra", key="usar_hormigon")
            usar_moldajes    = st.checkbox("Moldajes", key="usar_moldajes")
            usar_pisos       = st.checkbox("Pisos y Pavimentos", key="usar_pisos")
        with col2:
            usar_acero_est   = st.checkbox("Acero estructural", key="usar_acero_est")
            usar_muros       = st.checkbox("Muros", key="usar_muros")
            usar_terminaciones = st.checkbox("Terminaciones", key="usar_terminaciones")
        with col3:
            usar_metalcon    = st.checkbox("Acero No Estructural (Metalcon)", key="usar_metalcon")
            usar_revestimientos = st.checkbox("Revestimientos", key="usar_revestimientos")
            usar_cubierta    = st.checkbox("Cubierta / Techumbre", key="usar_cubierta")

        partidas_seleccionadas = {}

        if usar_hormigon:
            st.write("---")
            st.markdown("**Hormigón — selecciona las partidas:**")
            hc1, hc2, hc3 = st.columns(3)
            with hc1:
                p_exc = st.checkbox("Excavación", key="p_exc")
                p_emp = st.checkbox("Emplantillado", key="p_emp")
            with hc2:
                p_cim = st.checkbox("Cimiento", key="p_cim")
                p_sc  = st.checkbox("Sobrecimiento", key="p_sc")
            with hc3:
                p_rad = st.checkbox("Radier", key="p_rad")
            partidas_seleccionadas["hormigon"] = {
                "excavacion": p_exc, "emplantillado": p_emp,
                "cimiento": p_cim, "sobrecimiento": p_sc, "radier": p_rad,
            }

        if usar_acero_est:
            st.write("---")
            st.markdown("**Acero estructural — selecciona las partidas:**")
            ac1, ac2, ac3 = st.columns(3)
            with ac1:
                p_losa   = st.checkbox("Losa", key="p_losa")
                p_pilar  = st.checkbox("Pilar", key="p_pilar")
            with ac2:
                p_viga   = st.checkbox("Viga", key="p_viga")
                p_rad_ac = st.checkbox("Radier", key="p_rad_ac")
            with ac3:
                p_cim_ac = st.checkbox("Cimiento", key="p_cim_ac")
            partidas_seleccionadas["acero_estructural"] = {
                "losa": p_losa, "pilar": p_pilar, "viga": p_viga,
                "radier": p_rad_ac, "cimiento": p_cim_ac,
            }

        if usar_metalcon:
            st.write("---")
            st.markdown("**Metalcon — selecciona las partidas:**")
            mc1, mc2, mc3 = st.columns(3)
            with mc1:
                p_canal = st.checkbox("Canal / Solera", key="p_canal")
            with mc2:
                p_mont  = st.checkbox("Montante", key="p_mont")
            with mc3:
                p_esq   = st.checkbox("Esquinero", key="p_esq")
            partidas_seleccionadas["metalcon"] = {
                "canal": p_canal, "montante": p_mont, "esquinero": p_esq,
            }

        if usar_moldajes:
            st.write("---")
            st.markdown("**Moldajes — selecciona las partidas:**")
            mold1, mold2, mold3 = st.columns(3)
            with mold1:
                p_mold_cim  = st.checkbox("Cimiento", key="p_mold_cim")
                p_mold_muro = st.checkbox("Muro", key="p_mold_muro")
            with mold2:
                p_mold_losa = st.checkbox("Losa", key="p_mold_losa")
                p_mold_viga = st.checkbox("Viga", key="p_mold_viga")
            with mold3:
                p_mold_pilar = st.checkbox("Pilar", key="p_mold_pilar")
            partidas_seleccionadas["moldajes"] = {
                "cimiento": p_mold_cim, "muro": p_mold_muro,
                "losa": p_mold_losa, "viga": p_mold_viga, "pilar": p_mold_pilar,
            }

        if usar_muros:
            st.write("---")
            st.markdown("**Muros — selecciona las partidas:**")
            mu1, mu2, mu3 = st.columns(3)
            with mu1:
                p_muro_h = st.checkbox("Muro Hormigón", key="p_muro_h")
            with mu2:
                p_muro_l = st.checkbox("Muro Ladrillo", key="p_muro_l")
            with mu3:
                p_tab_met = st.checkbox("Tabique Metalcon", key="p_tab_met")
                p_tab_mad = st.checkbox("Tabique Madera", key="p_tab_mad")
            partidas_seleccionadas["muros"] = {
                "hormigon": p_muro_h, "ladrillo": p_muro_l,
                "tabique_metalcon": p_tab_met, "tabique_madera": p_tab_mad,
            }

        if usar_revestimientos:
            st.write("---")
            st.markdown("**Revestimientos — selecciona las partidas:**")
            re1, re2 = st.columns(2)
            with re1:
                p_rev_int = st.checkbox("Interior", key="p_rev_int")
            with re2:
                p_rev_ext = st.checkbox("Exterior", key="p_rev_ext")
            partidas_seleccionadas["revestimientos"] = {
                "interior": p_rev_int, "exterior": p_rev_ext,
            }

        if usar_pisos:
            st.write("---")
            st.markdown("**Pisos y Pavimentos — selecciona las partidas:**")
            pi1, pi2, pi3 = st.columns(3)
            with pi1:
                p_ceramico = st.checkbox("Cerámico / Porcelanato", key="p_ceramico")
                p_flotante = st.checkbox("Piso Flotante", key="p_flotante")
            with pi2:
                p_baldosa  = st.checkbox("Baldosa", key="p_baldosa")
            with pi3:
                p_deck     = st.checkbox("Deck Madera", key="p_deck")
            partidas_seleccionadas["pisos"] = {
                "ceramico": p_ceramico, "flotante": p_flotante,
                "baldosa": p_baldosa, "deck": p_deck,
            }

        if usar_terminaciones:
            st.write("---")
            st.markdown("**Terminaciones — selecciona las partidas:**")
            te1, te2, te3 = st.columns(3)
            with te1:
                p_pintura = st.checkbox("Pintura", key="p_pintura")
                p_estuco  = st.checkbox("Estuco / Revoque", key="p_estuco")
            with te2:
                p_cielos  = st.checkbox("Cielos", key="p_cielos")
            with te3:
                p_zocalos = st.checkbox("Zócalos", key="p_zocalos")
            partidas_seleccionadas["terminaciones"] = {
                "pintura": p_pintura, "estuco": p_estuco,
                "cielos": p_cielos, "zocalos": p_zocalos,
            }

        if usar_cubierta:
            st.write("---")
            st.markdown("**Cubierta / Techumbre — selecciona las partidas:**")
            cu1, cu2 = st.columns(2)
            with cu1:
                p_est_mad = st.checkbox("Estructura de Madera", key="p_est_mad")
                p_est_met = st.checkbox("Estructura Metálica", key="p_est_met")
            with cu2:
                p_planchas = st.checkbox("Cubierta (planchas)", key="p_planchas")
                p_aislacion = st.checkbox("Aislación y Fieltro", key="p_aislacion")
            partidas_seleccionadas["cubierta"] = {
                "estructura_madera": p_est_mad, "estructura_metalica": p_est_met,
                "planchas": p_planchas, "aislacion": p_aislacion,
            }

        total_partidas = sum(
            v for rubro in partidas_seleccionadas.values()
            for v in rubro.values()
        )

        st.write("---")
        col_cancel, col_count, col_crear = st.columns([2, 3, 2])
        with col_cancel:
            if st.button("Cancelar", use_container_width=True):
                st.session_state["proyecto_creado"] = False
                st.session_state["proyecto"] = {}
        with col_count:
            st.caption(f"{total_partidas} partidas seleccionadas")
        with col_crear:
            crear = st.button("Crear Proyecto", type="primary",
                use_container_width=True, disabled=not nombre_proy)

        if crear and nombre_proy:
            st.session_state["proyecto_creado"] = True
            st.session_state["mostrar_proyecto"] = True
            st.session_state["proyecto"] = {
                "nombre": nombre_proy,
                "tipo_obra": tipo_obra,
                "profesional": profesional,
                "partidas": partidas_seleccionadas,
            }
            # Limpiar materiales acumulados del proyecto anterior
            st.session_state["materiales_persistente"] = {}
            st.rerun()

    # --- Expander del proyecto creado ---
    if st.session_state.get("proyecto_creado"):
        proy = st.session_state["proyecto"]
        partidas = proy.get("partidas", {})

        with st.expander(f"Tu proyecto: {proy['nombre']}", expanded=True):
            col1, col2 = st.columns(2)
            with col1:
                if proy.get("tipo_obra"):
                    st.caption(f"Tipo de obra: {proy['tipo_obra']}")
            with col2:
                if proy.get("profesional"):
                    st.caption(f"Profesional: {proy['profesional']}")

            st.write("---")
            st.markdown("**Partidas seleccionadas:**")

            NOMBRES = {
                "hormigon": "Hormigón y Mov. de tierra",
                "acero_estructural": "Acero estructural",
                "metalcon": "Acero No Estructural (Metalcon)",
                "moldajes": "Moldajes",
                "muros": "Muros",
                "revestimientos": "Revestimientos",
                "pisos": "Pisos y Pavimentos",
                "terminaciones": "Terminaciones",
                "cubierta": "Cubierta / Techumbre",
                "cierres": "Cierres Perimetrales y Faena",
            }
            NOMBRES_PARTIDAS = {
                "excavacion": "Excavación", "emplantillado": "Emplantillado",
                "cimiento": "Cimiento", "sobrecimiento": "Sobrecimiento",
                "radier": "Radier", "losa": "Losa", "pilar": "Pilar",
                "viga": "Viga", "canal": "Canal / Solera",
                "montante": "Montante", "esquinero": "Esquinero",
                "muro": "Muro", "hormigon": "Muro Hormigón",
                "ladrillo": "Muro Ladrillo", "tabique_metalcon": "Tabique Metalcon",
                "tabique_madera": "Tabique Madera", "interior": "Interior",
                "exterior": "Exterior", "ceramico": "Cerámico / Porcelanato",
                "flotante": "Piso Flotante", "baldosa": "Baldosa",
                "deck": "Deck Madera", "pintura": "Pintura",
                "estuco": "Estuco / Revoque", "cielos": "Cielos",
                "zocalos": "Zócalos",
                "estructura_madera": "Estructura de Madera",
                "estructura_metalica": "Estructura Metálica",
                "planchas": "Cubierta (planchas)", "aislacion": "Aislación y Fieltro",
            }

            for rubro, sub in partidas.items():
                activas = [NOMBRES_PARTIDAS.get(k, k) for k, v in sub.items() if v]
                if activas:
                    st.markdown(f"**{NOMBRES.get(rubro, rubro)}**")
                    for p in activas:
                        st.markdown(f"- {p}")

            st.write("---")
            col_ir, col_cerrar = st.columns(2)
            with col_ir:
                if st.button("Ir a cubicación", type="primary", use_container_width=True):
                    st.session_state["ir_a_cubicacion"] = True
                    st.session_state["_goto"] = "Cubicacion"
                    st.rerun()
            with col_cerrar:
                if st.button("Cerrar proyecto", use_container_width=True):
                    st.session_state["proyecto_creado"] = False
                    st.session_state["proyecto"] = {}
                    st.rerun()

    st.info("También podés usar la Cubicación General sin crear un proyecto — está siempre disponible.")

# ============================
# CUBICACIÓN
# ============================
if option == "Cubicacion":
    proy_creado = st.session_state.get("proyecto_creado", False)
    partidas_proy = st.session_state.get("proyecto", {}).get("partidas", {})
    horm  = partidas_proy.get("hormigon", {})
    acero = partidas_proy.get("acero_estructural", {})
    metal = partidas_proy.get("metalcon", {})
    mold  = partidas_proy.get("moldajes", {})
    muros = partidas_proy.get("muros", {})
    revest = partidas_proy.get("revestimientos", {})
    pisos  = partidas_proy.get("pisos", {})
    term   = partidas_proy.get("terminaciones", {})
    cubierta = partidas_proy.get("cubierta", {})
    cierres = partidas_proy.get("cierres", {})

    # Usuario sin cuenta: acceso limitado a algunas partidas (según plan)
    sin_cuenta = not st.session_state.get("usuario")

    # Rubros/partidas permitidas sin cuenta: radier, tabiques (madera/metalcon), revestimiento
    RUBROS_SIN_CUENTA = {"hormigon", "muros", "revestimientos"}
    PARTIDAS_SIN_CUENTA = {
        "hormigon": {"radier"},
        "muros": {"tabique_metalcon", "tabique_madera"},
        "revestimientos": None,  # None = todas las de ese rubro
    }

    def rubro_permitido(nombre_rubro):
        if not sin_cuenta:
            return True
        return nombre_rubro in RUBROS_SIN_CUENTA

    def partida_permitida(nombre_rubro, nombre_partida):
        if not sin_cuenta:
            return True
        if nombre_rubro not in RUBROS_SIN_CUENTA:
            return False
        permitidas = PARTIDAS_SIN_CUENTA.get(nombre_rubro)
        if permitidas is None:
            return True
        return nombre_partida in permitidas

    # Mapeo de dict de rubro a su nombre interno (para chequear permisos sin cuenta)
    _MAPA_RUBRO = {
        id(horm): "hormigon", id(acero): "acero_estructural", id(metal): "metalcon",
        id(mold): "moldajes", id(muros): "muros", id(revest): "revestimientos",
        id(pisos): "pisos", id(term): "terminaciones", id(cubierta): "cubierta",
        id(cierres): "cierres",
    }

    def ver(rubro_dict, partida):
        # Sin cuenta: respetar partidas permitidas
        if sin_cuenta:
            nombre_rubro = _MAPA_RUBRO.get(id(rubro_dict))
            if nombre_rubro and not partida_permitida(nombre_rubro, partida):
                return False
        return not proy_creado or rubro_dict.get(partida, False)

    def ver_rubro(rubro_dict):
        return not proy_creado or (rubro_dict and any(rubro_dict.values()))
    
    st.subheader("CUBICACIONES")

    if sin_cuenta:
        st.info("👋 Estás usando ObraCubic **sin cuenta**. Puedes probar la cubicación de "
                "**radier, tabiques (Metalcon y madera) y revestimientos**. "
                "**Crea una cuenta gratis** para acceder a todas las partidas, guardar tus proyectos y más.")

    # pdf_extra se reinicia para reflejar lo que se renderiza en este ciclo,
    # pero el acumulador persistente conserva todo lo cubicado entre secciones.
    st.session_state["pdf_extra"] = []
    st.session_state.setdefault("materiales_persistente", {})

# ============================
# CIERRES PERIMETRALES Y FAENA
# ============================
    if ver_rubro(cierres) and rubro_permitido("cierres"):
        with st.expander("Cierres Perimetrales y Faena", expanded=False):
            st.caption("Cierres provisionales de obra: delimitan la faena, impiden el acceso "
                       "de terceros y protegen a peatones y trabajadores.")

            # Nota técnica OGUC
            st.warning(
                "⚖️ **Normativa (OGUC):** El cierre provisional **no debe invadir la vía "
                "pública** más allá de lo autorizado y debe mantener condiciones de seguridad "
                "e higiene durante toda la ejecución de la obra (Ordenanza General de "
                "Urbanismo y Construcciones)."
            )

            # --- Tramos del cierre: cada uno con su largo y altura ---
            st.markdown("**Tramos del cierre** (agrega cada tramo con su medida)")
            _tramos_cierre = secciones_input(
                "cierre", [("largo", "Largo (mL)"), ("alto", "Alto (m)")], "tramo")
            largo_cierre = sum(s["largo"] for s in _tramos_cierre)
            area_cierre = sum(s["largo"] * s["alto"] for s in _tramos_cierre)
            st.info(f"Superficie total de cierre: {area_cierre:.2f} m²  ·  Largo total: {largo_cierre:.2f} mL")

            tipo_cierre = st.selectbox(
                "Tipo de cierre",
                [
                    "Planchas de Zinc Alum (metálico opaco)",
                    "Paneles tipo Acmafor (provisional móvil)",
                    "Malla eslabonada / simple torsión",
                    "Planchas de OSB / Aglomerada (madera)",
                    "Tablas de Pino (empalizada)",
                ],
                key="cierre_tipo",
            )

            desp_cierre = st.slider("% Desperdicio", 0, 20, 10, key="cierre_desp")
            factor_desp = 1 + desp_cierre / 100

            # Separación estándar de los apoyos verticales (postes/polines)
            sep_apoyo = st.selectbox(
                "Separación entre apoyos verticales (m)", ["2,00", "2,50", "3,00"],
                index=1, key="cierre_sep_apoyo",
                help="Distancia estándar entre polines, postes o perfiles de soporte.")
            sep_apoyo_v = float(sep_apoyo.replace(",", "."))

            # Resumen informativo (estos items NO se cobran: el valor parte con un símbolo)
            items_cierre = [
                ("Tipo de cierre", tipo_cierre),
                ("Largo total", f"≈ {largo_cierre:.2f} mL"),
                ("Tramos", f"≈ {len(_tramos_cierre)} tramo(s)"),
                ("Superficie", f"≈ {area_cierre:.2f} m²"),
            ]

            st.write("---")
            st.subheader("📦 Materiales calculados")

            # =====================================================
            # 1.A — PLANCHAS DE ZINC ALUM (metálico opaco)
            # =====================================================
            if tipo_cierre.startswith("Planchas de Zinc"):
                largo_plancha_z = st.selectbox(
                    "Largo plancha zinc", ["2,00m", "2,50m", "3,00m", "3,66m"],
                    key="cierre_z_largo")
                sop_z = st.selectbox(
                    "Estructura de soporte", ["Polines de madera", "Perfiles metálicos"],
                    key="cierre_z_sop")

                ancho_util_z = 0.80  # ancho útil plancha zinc acanalada
                n_planchas_z = math.ceil(largo_cierre / ancho_util_z) if largo_cierre > 0 else 0
                n_planchas_z_d = math.ceil(n_planchas_z * factor_desp)
                n_apoyos_z = (math.ceil(largo_cierre / sep_apoyo_v) + 1) if largo_cierre > 0 else 0
                ml_rieles_z = 2 * largo_cierre  # solera superior + inferior
                tornillos_z = n_planchas_z_d * 10  # ~10 por plancha

                st.success(f"Planchas de zinc ({largo_plancha_z}): {n_planchas_z_d} unidades "
                           f"(c/{desp_cierre}% desp.)")
                st.info(f"Apoyos verticales: {n_apoyos_z} · Rieles horizontales: {ml_rieles_z:.1f} mL · "
                        f"Tornillos: {tornillos_z}")

                items_cierre.append(("Planchas de zinc acanaladas", f"{n_planchas_z_d} unidades de {largo_plancha_z}"))
                if sop_z == "Polines de madera":
                    items_cierre.append(("Polines impregnados (soporte)", f"{n_apoyos_z} piezas"))
                    items_cierre.append(("Rieles horizontales (madera)", f"{ml_rieles_z:.1f} mL"))
                else:
                    items_cierre.append(("Perfiles metálicos (soporte)", f"{n_apoyos_z} piezas"))
                    items_cierre.append(("Rieles horizontales (perfil)", f"{ml_rieles_z:.1f} mL"))
                items_cierre.append(("Tornillos de fijación", f"{tornillos_z} unidades"))

            # =====================================================
            # 1.B — PANELES TIPO ACMAFOR (provisional móvil)
            # =====================================================
            elif tipo_cierre.startswith("Paneles tipo Acmafor"):
                largo_panel_ac = 3.50  # panel móvil estándar
                n_paneles_ac = math.ceil(largo_cierre / largo_panel_ac) if largo_cierre > 0 else 0
                n_bases_ac = (n_paneles_ac + 1) if n_paneles_ac > 0 else 0
                n_abraz_ac = n_paneles_ac * 2  # abrazaderas de unión entre paneles

                st.success(f"Paneles Acmafor (3,50×2,00m): {n_paneles_ac} unidades")
                st.info(f"Bases de hormigón móviles: {n_bases_ac} · Abrazaderas: {n_abraz_ac}")
                st.caption("Cierre rápido de montar/desmontar. No entrega privacidad; "
                           "puedes complementarlo con malla Rachel (abajo).")

                items_cierre.append(("Paneles Acmafor provisionales", f"{n_paneles_ac} unidades"))
                items_cierre.append(("Base hormigón móvil", f"{n_bases_ac} unidades"))
                items_cierre.append(("Abrazaderas de unión", f"{n_abraz_ac} unidades"))

            # =====================================================
            # 1.C — MALLA ESLABONADA / SIMPLE TORSIÓN
            # =====================================================
            elif tipo_cierre.startswith("Malla eslabonada"):
                largo_rollo_me = st.selectbox(
                    "Largo del rollo de malla", ["10m", "20m", "25m"], index=2,
                    key="cierre_me_rollo")
                sop_me = st.selectbox(
                    "Postes de soporte", ["Metálicos (tubo/PHS)", "Madera (polines)"],
                    key="cierre_me_sop")
                largo_rollo_me_v = float(largo_rollo_me.replace("m", ""))

                n_rollos_me = math.ceil(largo_cierre / largo_rollo_me_v) if largo_cierre > 0 else 0
                n_postes_me = (math.ceil(largo_cierre / sep_apoyo_v) + 1) if largo_cierre > 0 else 0
                ml_alambre_me = 3 * largo_cierre  # 3 hebras tensoras (sup/medio/inf)

                st.success(f"Malla eslabonada: {n_rollos_me} rollos de {largo_rollo_me}")
                st.info(f"Postes: {n_postes_me} · Alambre tensor: {ml_alambre_me:.1f} mL (3 hebras)")

                items_cierre.append(("Malla eslabonada (rollos)", f"{n_rollos_me} rollos de {largo_rollo_me}"))
                if sop_me.startswith("Metálicos"):
                    items_cierre.append(("Postes metálicos / tubo PHS", f"{n_postes_me} piezas"))
                else:
                    items_cierre.append(("Polines de soporte", f"{n_postes_me} piezas"))
                items_cierre.append(("Alambre tensor", f"{ml_alambre_me:.1f} mL"))

            # =====================================================
            # 2.A — PLANCHAS DE OSB / AGLOMERADA (madera)
            # =====================================================
            elif tipo_cierre.startswith("Planchas de OSB"):
                area_plancha_osb = 1.22 * 2.44  # plancha estándar (≈2,98 m²)
                n_planchas_osb = math.ceil((area_cierre / area_plancha_osb) * factor_desp) if area_cierre > 0 else 0
                n_pies_osb = (math.ceil(largo_cierre / 0.60) + 1) if largo_cierre > 0 else 0  # pies derechos c/0,60m
                ml_soleras_osb = 2 * largo_cierre  # solera sup + inf
                n_diag_osb = math.ceil(largo_cierre / sep_apoyo_v) if largo_cierre > 0 else 0  # 1 diagonal por vano
                tornillos_osb = n_planchas_osb * 20  # ~20 fijaciones por plancha

                st.success(f"Planchas de OSB (1,22×2,44m): {n_planchas_osb} unidades "
                           f"(c/{desp_cierre}% desp.)")
                st.info(f"Pies derechos: {n_pies_osb} · Soleras: {ml_soleras_osb:.1f} mL · "
                        f"Diagonales: {n_diag_osb} · Tornillos: {tornillos_osb}")
                st.caption("Panel ciego: excelente barrera visual y de seguridad. "
                           "Puede pintarse para mejorar la estética urbana.")

                items_cierre.append(("Planchas de OSB", f"{n_planchas_osb} unidades"))
                items_cierre.append(("Pie derecho (entramado)", f"{n_pies_osb} piezas"))
                items_cierre.append(("Soleras (entramado)", f"{ml_soleras_osb:.1f} mL"))
                items_cierre.append(("Diagonales", f"{n_diag_osb} piezas"))
                items_cierre.append(("Tornillos de fijación", f"{tornillos_osb} unidades"))

            # =====================================================
            # 2.B — TABLAS DE PINO (empalizada)
            # =====================================================
            elif tipo_cierre.startswith("Tablas de Pino"):
                orient_t = st.selectbox(
                    "Orientación de las tablas", ["Horizontal", "Vertical"],
                    key="cierre_t_orient")
                largo_tabla_t = st.selectbox(
                    "Largo tabla pino", ["3,20m", "4,00m"], key="cierre_t_largo")
                ancho_tabla_t = st.selectbox(
                    "Ancho tabla", ['1x4" (≈0,10m)', '1x6" (≈0,15m)'], key="cierre_t_ancho")

                largo_tabla_t_v = float(largo_tabla_t.replace("m", "").replace(",", "."))
                ancho_tabla_t_v = 0.10 if "0,10" in ancho_tabla_t else 0.15

                # Metros lineales de tabla = área total / ancho de la tabla
                # (válido en horizontal o vertical, y con tramos de medidas distintas)
                ml_tablas_t = area_cierre / ancho_tabla_t_v if ancho_tabla_t_v > 0 else 0
                n_tablas_t = math.ceil((ml_tablas_t / largo_tabla_t_v) * factor_desp) if largo_tabla_t_v > 0 else 0
                n_postes_t = (math.ceil(largo_cierre / sep_apoyo_v) + 1) if largo_cierre > 0 else 0
                clavos_t = n_tablas_t * 4  # ~4 clavos por tabla

                st.success(f"Tablas de pino ({largo_tabla_t}): {n_tablas_t} unidades "
                           f"(c/{desp_cierre}% desp.)")
                st.info(f"Polines de soporte: {n_postes_t} · Clavos: {clavos_t}")

                items_cierre.append(("Tablas de pino", f"{n_tablas_t} unidades de {largo_tabla_t}"))
                items_cierre.append(("Polines de soporte", f"{n_postes_t} piezas"))
                items_cierre.append(("Clavos", f"{clavos_t} unidades"))

            # =====================================================
            # 3 — COMPLEMENTOS DE PLÁSTICO Y SEÑALIZACIÓN (checkbox)
            # =====================================================
            st.write("---")
            st.subheader("➕ Complementos (opcional)")

            usar_rachel = st.checkbox(
                "Agregar malla Rachel (sombra / control de polvo / privacidad)",
                key="cierre_rachel")
            if usar_rachel:
                color_rachel = st.selectbox("Color malla Rachel", ["Verde", "Negra"],
                                            key="cierre_rachel_color")
                largo_rollo_rachel = 50.0  # rollo estándar 50 m
                m2_rachel = area_cierre * factor_desp
                n_rollos_rachel = math.ceil(largo_cierre / largo_rollo_rachel) if largo_cierre > 0 else 0
                st.info(f"Malla Rachel {color_rachel}: {n_rollos_rachel} rollos (≈{m2_rachel:.1f} m²)")
                items_cierre.append((f"Malla Rachel {color_rachel}",
                                     f"{n_rollos_rachel} rollos (≈{m2_rachel:.1f} m²)"))

            usar_senal = st.checkbox(
                "Agregar malla plástica de señalización (naranja, faena / excavaciones)",
                key="cierre_senal")
            if usar_senal:
                ml_senal = st.number_input(
                    "Metros a señalizar (mL)", value=float(largo_cierre), min_value=0.0, step=1.0,
                    key="cierre_senal_ml")
                n_rollos_senal = math.ceil(ml_senal / 50.0) if ml_senal > 0 else 0
                st.info(f"Malla señalización naranja: {n_rollos_senal} rollos de 50m")
                items_cierre.append(("Malla señalización naranja", f"{n_rollos_senal} rollos de 50m"))

            # Registrar para PDF y presupuesto solo si hay un cierre real cubicado.
            # Como el tipo se elige con un selectbox, al cambiar de tipo hay que
            # borrar el registro del tipo anterior (si no, aparecen los dos).
            _persist_c = st.session_state.setdefault("materiales_persistente", {})
            if largo_cierre > 0:
                nombre_partida_cierre = tipo_cierre.split(" (")[0]
                _clave_actual = f"Cierres Perimetrales y Faena||{nombre_partida_cierre}"
                # Eliminar cualquier OTRO tipo de cierre registrado antes
                for _k in list(_persist_c.keys()):
                    if _k.startswith("Cierres Perimetrales y Faena||") and _k != _clave_actual:
                        _persist_c.pop(_k, None)
                registrar_pdf("Cierres Perimetrales y Faena", nombre_partida_cierre, items_cierre)
            else:
                # Sin largo: no hay cierre cubicado, limpiar todos sus registros
                for _k in list(_persist_c.keys()):
                    if _k.startswith("Cierres Perimetrales y Faena||"):
                        _persist_c.pop(_k, None)

    if ver_rubro(horm):
            with st.expander("Hormigón y Movimiento de tierra", expanded=False):
                if ver(horm, "excavacion"):
                    with st.expander("1. Excavación", expanded=False):

                        if "secciones_exc" not in st.session_state:
                            st.session_state.secciones_exc = [{"largo": 0.0, "ancho": 0.0, "prof": 0.0}]

                        col_add, col_del = st.columns(2)
                        with col_add:
                            if st.button("➕ Agregar sección", key="add_exc"):
                                st.session_state.secciones_exc.append({"largo": 0.0, "ancho": 0.0, "prof": 0.0})
                        with col_del:
                            if st.button("🗑️ Eliminar última sección", key="del_exc"):
                                if len(st.session_state.secciones_exc) > 1:
                                    st.session_state.secciones_exc.pop()

                        vol_exc_total = 0.0
                        for i, sec in enumerate(st.session_state.secciones_exc):
                            st.markdown(f"**Sección {i+1}**")
                            ex1, ex2, ex3 = st.columns(3)
                            with ex1:
                                sec["largo"] = st.number_input(f"Largo (m)", value=sec["largo"], key=f"exc_largo_{i}")
                            with ex2:
                                sec["ancho"] = st.number_input(f"Ancho (m)", value=sec["ancho"], key=f"exc_ancho_{i}")
                            with ex3:
                                sec["prof"] = st.number_input(f"Profundidad (m)", value=sec["prof"], key=f"exc_prof_{i}")

                            vol_sec_exc = sec["largo"] * sec["ancho"] * sec["prof"]
                            st.caption(f"Volumen sección {i+1}: {vol_sec_exc:.2f} m³")
                            vol_exc_total += vol_sec_exc

                        exc_perdida = st.slider("% Esponjamiento", 0, 30, 20, key="exc_perdida")
                        vol_exc_final = vol_exc_total * (1 + exc_perdida / 100)

                        st.write("---")
                        st.info(f"Volumen neto excavación: {vol_exc_total:.2f} m³")
                        st.success(f"Volumen con {exc_perdida}% esponjamiento: {vol_exc_final:.2f} m³")

        # --- 2. Emplantillado ---
                
                if ver(horm, "emplantillado"):
                    with st.expander("2. Emplantillado", expanded=False):
                        col_add, col_del = st.columns(2)
                        with col_add:
                            if st.button("➕ Agregar sección", key="add_emp"):
                                st.session_state.secciones_emp.append({"largo": 0.0, "ancho": 0.0, "espesor": 0.0})
                        with col_del:
                            if st.button("🗑️ Eliminar última sección", key="del_emp"):
                                if len(st.session_state.secciones_emp) > 1:
                                    st.session_state.secciones_emp.pop()

                        vol_emp_total = 0.0
                        for i, sec in enumerate(st.session_state.secciones_emp):
                            st.markdown(f"**Sección {i+1}**")
                            em1, em2, em3 = st.columns(3)
                            with em1:
                                sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"emp_largo_{i}")
                            with em2:
                                sec["ancho"] = st.number_input("Ancho (m)", value=sec["ancho"], key=f"emp_ancho_{i}")
                            with em3:
                                sec["espesor"] = st.number_input("Espesor (m)", value=sec["espesor"], key=f"emp_espesor_{i}")
                            vol_sec_emp = sec["largo"] * sec["ancho"] * sec["espesor"]
                            st.caption(f"Volumen sección {i+1}: {vol_sec_emp:.2f} m³")
                            vol_emp_total += vol_sec_emp

                        emp_perdida = st.slider("% Pérdida", 0, 15, 5, key="emp_perdida")
                        dos_emp = st.selectbox("Dosificación", list(DOSIFICACIONES.keys()),
                                                key="dos_emp", help=DOSIFICACIONES["G-15"]["descripcion"])

                        st.session_state["_emp_vol"] = vol_emp_total
                        st.session_state["_emp_perdida"] = emp_perdida
                        st.session_state["_emp_dos"] = dos_emp

                        vol_emp_final = vol_emp_total * (1 + emp_perdida / 100)
                        mat_emp = calcular_materiales(vol_emp_final, dos_emp)
                        st.session_state["mat_emp"] = mat_emp
                        st.session_state["vol_emp"] = vol_emp_final

                        st.write("---")
                        st.info(f"Volumen neto emplantillado: {vol_emp_total:.2f} m³")
                        st.success(f"Volumen con {emp_perdida}% pérdida: {vol_emp_final:.2f} m³")
                        mostrar_materiales(mat_emp, "Hormigón y Movimiento de tierra", "Emplantillado")

                # Recalculo FUERA del expander
                _vol_emp = st.session_state.get("_emp_vol", 0)
                _perd_emp = st.session_state.get("_emp_perdida", 5)
                _dos_emp = st.session_state.get("_emp_dos", "G-15")
                vol_emp_final = _vol_emp * (1 + _perd_emp / 100)
                mat_emp = calcular_materiales(vol_emp_final, _dos_emp)
                st.session_state["mat_emp"] = mat_emp
                st.session_state["vol_emp"] = vol_emp_final

            # --- 3. Cimiento ---
                if ver(horm, "cimiento"):
                    with st.expander("3. Cimiento", expanded=False):        

                        st.write("**Selecciona los tipos de cimiento que tiene la obra:**")

                        usar_zapata_aislada   = st.checkbox("Zapata Aislada", key="usar_zapata_aislada")
                        usar_zapata_corrida   = st.checkbox("Zapata Corrida", key="usar_zapata_corrida")
                        usar_zapata_combinada = st.checkbox("Zapata Combinada", key="usar_zapata_combinada")
                        usar_losa_cim         = st.checkbox("Losa de Cimentación", key="usar_losa_cim")

                        vol_pilares = 0.0

                        # --- Zapata Aislada ---
                        if usar_zapata_aislada:
                            st.write("---")
                            st.markdown("### 📐 Zapata Aislada")

                            if "secciones_zapata" not in st.session_state:
                                st.session_state.secciones_zapata = [{"cant": 0, "seccion": 0.0, "alto": 0.0}]

                            col_add, col_del = st.columns(2)
                            with col_add:
                                if st.button("➕ Agregar grupo", key="add_zapata"):
                                    st.session_state.secciones_zapata.append({"cant": 0, "seccion": 0.0, "alto": 0.0})
                            with col_del:
                                if st.button("🗑️ Eliminar último grupo", key="del_zapata"):
                                    if len(st.session_state.secciones_zapata) > 1:
                                        st.session_state.secciones_zapata.pop()

                            vol_zapata_aislada = 0.0
                            for i, sec in enumerate(st.session_state.secciones_zapata):
                                st.markdown(f"**Grupo {i+1}**")
                                c1, c2, c3 = st.columns(3)
                                with c1:
                                    sec["cant"] = st.number_input("Cantidad zapatas", value=sec["cant"], step=1, key=f"zapata_cant_{i}")
                                with c2:
                                    sec["seccion"] = st.number_input("Sección (m)", value=sec["seccion"], key=f"zapata_sec_{i}")
                                with c3:
                                    sec["alto"] = st.number_input("Profundidad (m)", value=sec["alto"], key=f"zapata_alto_{i}")

                                vol_sec = (sec["seccion"] * sec["seccion"] * sec["alto"]) * sec["cant"]
                                st.caption(f"Volumen grupo {i+1}: {vol_sec:.2f} m³")
                                vol_zapata_aislada += vol_sec

                            st.info(f"Volumen Zapata Aislada: {vol_zapata_aislada:.2f} m³")
                            vol_pilares += vol_zapata_aislada

                        # --- Zapata Corrida ---
                        if usar_zapata_corrida:
                            st.write("---")
                            st.markdown("### 📐 Zapata Corrida")

                            if "secciones_corrida" not in st.session_state:
                                st.session_state.secciones_corrida = [{"largo": 0.0, "ancho": 0.0, "prof": 0.0}]

                            col_add, col_del = st.columns(2)
                            with col_add:
                                if st.button("➕ Agregar sección", key="add_corrida"):
                                    st.session_state.secciones_corrida.append({"largo": 0.0, "ancho": 0.0, "prof": 0.0})
                            with col_del:
                                if st.button("🗑️ Eliminar última sección", key="del_corrida"):
                                    if len(st.session_state.secciones_corrida) > 1:
                                        st.session_state.secciones_corrida.pop()

                            vol_zapata_corrida = 0.0
                            for i, sec in enumerate(st.session_state.secciones_corrida):
                                st.markdown(f"**Sección {i+1}**")
                                c1, c2, c3 = st.columns(3)
                                with c1:
                                    sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"corrida_largo_{i}")
                                with c2:
                                    sec["ancho"] = st.number_input("Ancho/Espesor (m)", value=sec["ancho"], key=f"corrida_ancho_{i}")
                                with c3:
                                    sec["prof"] = st.number_input("Profundidad (m)", value=sec["prof"], key=f"corrida_prof_{i}")

                                vol_sec = sec["largo"] * sec["ancho"] * sec["prof"]
                                st.caption(f"Volumen sección {i+1}: {vol_sec:.2f} m³")
                                vol_zapata_corrida += vol_sec

                            st.info(f"Volumen Zapata Corrida: {vol_zapata_corrida:.2f} m³")
                            vol_pilares += vol_zapata_corrida

                        # --- Zapata Combinada ---
                        if usar_zapata_combinada:
                            st.write("---")
                            st.markdown("### 📐 Zapata Combinada")
                            st.caption("Une dos o más pilares cercanos en una sola zapata.")

                            if "secciones_combinada" not in st.session_state:
                                st.session_state.secciones_combinada = [{"cant": 0, "largo": 0.0, "ancho": 0.0, "prof": 0.0}]

                            col_add, col_del = st.columns(2)
                            with col_add:
                                if st.button("➕ Agregar sección", key="add_combinada"):
                                    st.session_state.secciones_combinada.append({"cant": 0, "largo": 0.0, "ancho": 0.0, "prof": 0.0})
                            with col_del:
                                if st.button("🗑️ Eliminar última sección", key="del_combinada"):
                                    if len(st.session_state.secciones_combinada) > 1:
                                        st.session_state.secciones_combinada.pop()

                            vol_zapata_combinada = 0.0
                            for i, sec in enumerate(st.session_state.secciones_combinada):
                                st.markdown(f"**Sección {i+1}**")
                                c1, c2, c3, c4 = st.columns(4)
                                with c1:
                                    sec["cant"] = st.number_input("Cantidad zapatas", value=sec["cant"], step=1, key=f"comb_cant_{i}")
                                with c2:
                                    sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"comb_largo_{i}")
                                with c3:
                                    sec["ancho"] = st.number_input("Ancho (m)", value=sec["ancho"], key=f"comb_ancho_{i}")
                                with c4:
                                    sec["prof"] = st.number_input("Profundidad (m)", value=sec["prof"], key=f"comb_prof_{i}")

                                vol_sec = sec["largo"] * sec["ancho"] * sec["prof"] * sec["cant"]
                                st.caption(f"Volumen sección {i+1}: {vol_sec:.2f} m³")
                                vol_zapata_combinada += vol_sec

                            st.info(f"Volumen Zapata Combinada: {vol_zapata_combinada:.2f} m³")
                            vol_pilares += vol_zapata_combinada

                        # --- Losa de Cimentación ---
                        if usar_losa_cim:
                            st.write("---")
                            st.markdown("### 📐 Losa de Cimentación")
                            st.caption("Placa continua bajo toda la estructura. Se usa en terrenos de baja capacidad portante.")

                            if "secciones_losa_cim" not in st.session_state:
                                st.session_state.secciones_losa_cim = [{"largo": 0.0, "ancho": 0.0, "esp": 0.0}]

                            col_add, col_del = st.columns(2)
                            with col_add:
                                if st.button("➕ Agregar sección", key="add_losa_cim"):
                                    st.session_state.secciones_losa_cim.append({"largo": 0.0, "ancho": 0.0, "esp": 0.0})
                            with col_del:
                                if st.button("🗑️ Eliminar última sección", key="del_losa_cim"):
                                    if len(st.session_state.secciones_losa_cim) > 1:
                                        st.session_state.secciones_losa_cim.pop()

                            vol_losa_cim = 0.0
                            for i, sec in enumerate(st.session_state.secciones_losa_cim):
                                st.markdown(f"**Sección {i+1}**")
                                c1, c2, c3 = st.columns(3)
                                with c1:
                                    sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"losacim_largo_{i}")
                                with c2:
                                    sec["ancho"] = st.number_input("Ancho (m)", value=sec["ancho"], key=f"losacim_ancho_{i}")
                                with c3:
                                    sec["esp"] = st.number_input("Espesor (m)", value=sec["esp"], key=f"losacim_esp_{i}")

                                vol_sec = sec["largo"] * sec["ancho"] * sec["esp"]
                                st.caption(f"Volumen sección {i+1}: {vol_sec:.2f} m³")
                                vol_losa_cim += vol_sec

                            st.info(f"Volumen Losa de Cimentación: {vol_losa_cim:.2f} m³")
                            vol_pilares += vol_losa_cim

                        # --- Resumen y Dosificación ---
                        if vol_pilares > 0:
                            st.write("---")
                            st.success(f"### Volumen Total Cimiento: {vol_pilares:.2f} m³")
                            st.caption("Para pilotes o micropilotes consulte con un ingeniero especialista.")

                            dos_cim = st.selectbox("Dosificación", list(DOSIFICACIONES.keys()),
                                                index=1, key="dos_cim",
                                                help=DOSIFICACIONES["G-20"]["descripcion"])
                            mat_cim = calcular_materiales(vol_pilares, dos_cim)
                            st.session_state["mat_cim"] = mat_cim
                            st.session_state["vol_pilares"] = vol_pilares
                            mostrar_materiales(mat_cim, "Hormigón y Movimiento de tierra", "Cimiento")
                        else:
                            st.caption("Selecciona al menos un tipo de cimiento para calcular.")

                        st.session_state["_cim_vol"] = vol_pilares
                        st.session_state["_cim_dos"] = st.session_state.get("dos_cim", "G-20")

                        mat_cim = calcular_materiales(vol_pilares, st.session_state.get("dos_cim", "G-20"))
                        st.session_state["mat_cim"] = mat_cim
                        st.session_state["vol_pilares"] = vol_pilares
                        mostrar_materiales(mat_cim, "Hormigón y Movimiento de tierra", "Cimiento")

                _vol_cim = st.session_state.get("_cim_vol", 0)
                _dos_cim = st.session_state.get("_cim_dos", "G-20")
                mat_cim = calcular_materiales(_vol_cim, _dos_cim)
                st.session_state["mat_cim"] = mat_cim
                st.session_state["vol_pilares"] = _vol_cim

            # --- 4. Sobrecimiento ---
                if ver(horm, "sobrecimiento"):
                    with st.expander("4. Sobrecimiento", expanded=False):
                        if "secciones_sc" not in st.session_state:
                            st.session_state.secciones_sc = [{"largo": 0.0, "ancho": 0.0, "alto": 0.0}]

                        col_add, col_del = st.columns(2)
                        with col_add:
                            if st.button("➕ Agregar sección", key="add_sc"):
                                st.session_state.secciones_sc.append({"largo": 0.0, "ancho": 0.0, "alto": 0.0})
                        with col_del:
                            if st.button("🗑️ Eliminar última sección", key="del_sc"):
                                if len(st.session_state.secciones_sc) > 1:
                                    st.session_state.secciones_sc.pop()

                        vol_sc_total = 0.0
                        for i, sec in enumerate(st.session_state.secciones_sc):
                            st.markdown(f"**Sección {i+1}**")
                            sc1, sc2, sc3 = st.columns(3)
                            with sc1:
                                sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"sc_largo_{i}")
                            with sc2:
                                sec["ancho"] = st.number_input("Ancho/Espesor (m)", value=sec["ancho"], key=f"sc_ancho_{i}")
                            with sc3:
                                sec["alto"] = st.number_input("Alto (m)", value=sec["alto"], key=f"sc_alto_{i}")

                            vol_sec_sc = sec["largo"] * sec["ancho"] * sec["alto"]
                            st.caption(f"Volumen bruto sección {i+1}: {vol_sec_sc:.2f} m³")
                            vol_sc_total += vol_sec_sc

                        st.write("**Descuento de Vanos**")
                        v1, v2 = st.columns(2)
                        with v1:
                            n_vanos_sc = st.number_input("Cantidad de vanos/puertas", value=0, step=1, key="vanos_cant_sc")
                        with v2:
                            ancho_vano_sc = st.number_input("Ancho del vano (m)", value=0.0, key="vanos_ancho_sc")

                        espesor_sc = st.session_state.secciones_sc[0]["ancho"] if st.session_state.secciones_sc else 0.15
                        alto_sc = st.session_state.secciones_sc[0]["alto"] if st.session_state.secciones_sc else 0.30
                        vol_vanos_sc = n_vanos_sc * ancho_vano_sc * espesor_sc * alto_sc
                        vol_sc_neto = vol_sc_total - vol_vanos_sc

                        dos_sc = st.selectbox("Dosificación", list(DOSIFICACIONES.keys()),
                                                index=1, key="dos_sc",
                                                help=DOSIFICACIONES["G-20"]["descripcion"])
                        mat_sc = calcular_materiales(vol_sc_neto, dos_sc)
                        st.session_state["mat_sc"] = mat_sc
                        st.session_state["vol_sc_neto"] = vol_sc_neto

                        st.write("---")
                        st.text(f"Volumen bruto total: {vol_sc_total:.2f} m³")
                        st.text(f"Descuento vanos: -{vol_vanos_sc:.2f} m³")
                        st.info(f"Volumen neto sobrecimiento: {vol_sc_neto:.2f} m³")
                        mostrar_materiales(mat_sc, "Hormigón y Movimiento de tierra", "Sobrecimiento")

                        st.session_state["_sc_vol"] = vol_sc_neto
                        st.session_state["_sc_dos"] = dos_sc

                        mat_sc = calcular_materiales(vol_sc_neto, dos_sc)
                        st.session_state["mat_sc"] = mat_sc
                        st.session_state["vol_sc_neto"] = vol_sc_neto
                        mostrar_materiales(mat_sc, "Hormigón y Movimiento de tierra", "Sobrecimiento")

                _vol_sc = st.session_state.get("_sc_vol", 0)
                _dos_sc = st.session_state.get("_sc_dos", "G-20")
                mat_sc = calcular_materiales(_vol_sc, _dos_sc)
                st.session_state["mat_sc"] = mat_sc
                st.session_state["vol_sc_neto"] = _vol_sc

            # --- 5. Radier ---
                if ver(horm, "radier"):
                        with st.expander("5. Radier", expanded=False):

                            if "secciones_rad" not in st.session_state:
                                st.session_state.secciones_rad = [{"largo": 0.0, "ancho": 0.0, "espesor": 0.0}]

                            col_add, col_del = st.columns(2)
                            with col_add:
                                if st.button("➕ Agregar sección", key="add_rad"):
                                    st.session_state.secciones_rad.append({"largo": 0.0, "ancho": 0.0, "espesor": 0.0})
                            with col_del:
                                if st.button("🗑️ Eliminar última sección", key="del_rad"):
                                    if len(st.session_state.secciones_rad) > 1:
                                        st.session_state.secciones_rad.pop()

                            area_radier_total = 0.0
                            vol_radier = 0.0
                            for i, sec in enumerate(st.session_state.secciones_rad):
                                st.markdown(f"**Sección {i+1}**")
                                ra1, ra2, ra3 = st.columns(3)
                                with ra1:
                                    sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"rad_largo_{i}")
                                with ra2:
                                    sec["ancho"] = st.number_input("Ancho (m)", value=sec["ancho"], key=f"rad_ancho_{i}")
                                with ra3:
                                    sec["espesor"] = st.number_input("Espesor (m)", value=sec["espesor"], key=f"rad_espesor_{i}")

                                vol_sec_rad = sec["largo"] * sec["ancho"] * sec["espesor"]
                                area_sec_rad = sec["largo"] * sec["ancho"]
                                st.caption(f"Volumen sección {i+1}: {vol_sec_rad:.2f} m³ | Área: {area_sec_rad:.2f} m²")
                                vol_radier += vol_sec_rad
                                area_radier_total += area_sec_rad

                            rad_perdida = st.slider("% Pérdida Radier", 0, 15, 5, key="radier_perdida")
                            vol_radier_final = vol_radier * (1 + rad_perdida / 100)

                            dos_rad = st.selectbox("Dosificación", list(DOSIFICACIONES.keys()),
                                                    index=1, key="dos_rad",
                                                    help=DOSIFICACIONES["G-20"]["descripcion"])
                            mat_rad = calcular_materiales(vol_radier_final, dos_rad)
                            st.session_state["mat_rad"] = mat_rad
                            st.session_state["vol_radier"] = vol_radier_final

                            # --- Malla ACMA ---
                            usar_malla = st.checkbox("¿Agregar malla ACMA al radier?", key="usar_malla")
                            if usar_malla:
                                malla_tipo = st.selectbox("Tipo de malla ACMA", list(MALLA_ACMA.keys()), key="malla_tipo")
                                desp_malla = st.slider("% Desperdicio malla", 0, 20, 10, key="desp_malla")

                                kg_malla = area_radier_total * MALLA_ACMA[malla_tipo]["kg_m2"]
                                kg_malla_desp = kg_malla * (1 + desp_malla / 100)
                                area_plancha_malla = 14.10  # plancha 2,35x6,00m
                                planchas_malla = area_radier_total / area_plancha_malla
                                planchas_malla_desp = planchas_malla * (1 + desp_malla / 100)

                            st.write("---")
                            st.info(f"Volumen neto radier: {vol_radier:.2f} m³")
                            st.info(f"Área total radier: {area_radier_total:.2f} m²")
                            st.success(f"Volumen con {rad_perdida}% pérdida: {vol_radier_final:.2f} m³")
                            mostrar_materiales(mat_rad, "Hormigón y Movimiento de tierra", "Radier")

                            if usar_malla:
                                st.write("---")
                                st.subheader("🔩 Malla ACMA")
                                st.info(f"Tipo: {malla_tipo}")
                                st.info(f"Peso malla: {kg_malla:.1f} kg")
                                st.success(f"Con {desp_malla}% desperdicio: {kg_malla_desp:.1f} kg")
                                st.success(f"Planchas 2,35x6,00m: {planchas_malla_desp:.0f} planchas")
                                st.caption(f"Área total: {area_radier_total:.2f} m² | Plancha cubre: 14,1 m²")

                            st.session_state["_rad_vol"] = vol_radier
                            st.session_state["_rad_perdida"] = rad_perdida
                            st.session_state["_rad_dos"] = dos_rad

                _vol_rad = st.session_state.get("_rad_vol", 0)
                _perd_rad = st.session_state.get("_rad_perdida", 5)
                _dos_rad = st.session_state.get("_rad_dos", "G-20")
                vol_radier_final = _vol_rad * (1 + _perd_rad / 100)
                mat_rad = calcular_materiales(vol_radier_final, _dos_rad)
                st.session_state["mat_rad"] = mat_rad
                st.session_state["vol_radier"] = vol_radier_final

                st.write("---")
                total_hormigon = (
                    st.session_state.get("vol_emp", 0) +
                    st.session_state.get("vol_pilares", 0) +
                    st.session_state.get("vol_sc_neto", 0) +
                    st.session_state.get("vol_radier", 0)
)
                # Recalculo siempre desde las secciones guardadas
                # Emplantillado
                _vol_emp = sum(s["largo"] * s["ancho"] * s["espesor"] for s in st.session_state.get("secciones_emp", [{"largo":0,"ancho":0,"espesor":0}]))
                _perd_emp = st.session_state.get("emp_perdida", 5)
                _dos_emp = st.session_state.get("dos_emp", "G-15")
                vol_emp_final = _vol_emp * (1 + _perd_emp / 100)
                mat_emp = calcular_materiales(vol_emp_final, _dos_emp)
                st.session_state["mat_emp"] = mat_emp
                st.session_state["vol_emp"] = vol_emp_final

                # Cimiento
                _vol_cim = st.session_state.get("_cim_vol", 0)
                _dos_cim = st.session_state.get("dos_cim", "G-20")
                mat_cim = calcular_materiales(_vol_cim, _dos_cim)
                st.session_state["mat_cim"] = mat_cim
                st.session_state["vol_pilares"] = _vol_cim

                # Sobrecimiento
                _vol_sc = st.session_state.get("_sc_vol", 0)
                _dos_sc = st.session_state.get("dos_sc", "G-20")
                mat_sc = calcular_materiales(_vol_sc, _dos_sc)
                st.session_state["mat_sc"] = mat_sc
                st.session_state["vol_sc_neto"] = _vol_sc

                # Radier
                _vol_rad = sum(s["largo"] * s["ancho"] * s["espesor"] for s in st.session_state.get("secciones_rad", [{"largo":0,"ancho":0,"espesor":0}]))
                _perd_rad = st.session_state.get("radier_perdida", 5)
                _dos_rad = st.session_state.get("dos_rad", "G-20")
                vol_radier_final = _vol_rad * (1 + _perd_rad / 100)
                mat_rad = calcular_materiales(vol_radier_final, _dos_rad)
                st.session_state["mat_rad"] = mat_rad
                st.session_state["vol_radier"] = vol_radier_final
                
                st.subheader("Resumen Total de Materiales")
                st.caption("Suma de todas las partidas con sus respectivas dosificaciones y desperdicios")

                total_sacos    = (st.session_state.get("mat_emp", {}).get("cemento_sacos", 0) +
                                st.session_state.get("mat_cim", {}).get("cemento_sacos", 0) +
                                st.session_state.get("mat_sc",  {}).get("cemento_sacos", 0) +
                                st.session_state.get("mat_rad", {}).get("cemento_sacos", 0))

                total_gravilla = (st.session_state.get("mat_emp", {}).get("gravilla_kg", 0) +
                                st.session_state.get("mat_cim", {}).get("gravilla_kg", 0) +
                                st.session_state.get("mat_sc",  {}).get("gravilla_kg", 0) +
                                st.session_state.get("mat_rad", {}).get("gravilla_kg", 0))

                total_arena    = (st.session_state.get("mat_emp", {}).get("arena_kg", 0) +
                                st.session_state.get("mat_cim", {}).get("arena_kg", 0) +
                                st.session_state.get("mat_sc",  {}).get("arena_kg", 0) +
                                st.session_state.get("mat_rad", {}).get("arena_kg", 0))

                total_agua     = (st.session_state.get("mat_emp", {}).get("agua_lt", 0) +
                                st.session_state.get("mat_cim", {}).get("agua_lt", 0) +
                                st.session_state.get("mat_sc",  {}).get("agua_lt", 0) +
                                st.session_state.get("mat_rad", {}).get("agua_lt", 0))

                total_hormigon = (st.session_state.get("vol_emp", 0) +
                                st.session_state.get("vol_pilares", 0) +
                                st.session_state.get("vol_sc_neto", 0) +
                                st.session_state.get("vol_radier", 0))

                st.success(f"### Volumen Total Neto de la Obra: {total_hormigon:.2f} m³")

                r1, r2, r3, r4 = st.columns(4)
                r1.metric("Cemento Total",  f"{total_sacos} sacos")
                r2.metric("Gravilla Total", f"{total_gravilla} kg")
                r3.metric("Arena Total",    f"{total_arena} kg")
                r4.metric("Agua Total",     f"{total_agua} lt")

    # --- Acero estructural ---
    if ver_rubro(acero) and rubro_permitido("acero_estructural"):
        with st.expander("Acero estructural", expanded=False):
            if ver(acero, "losa"):
                with st.expander("1. Losa", expanded=False):
                    modo_losa = st.radio(
                        "Modo de cálculo",
                        ["🔨 Modo Simple", "📐 Modo Detallado"],
                        horizontal=True, key="modo_losa"
                    )
                    if modo_losa == "🔨 Modo Simple":
                        st.caption("Estimación por ratio kg/m²")
                        ls1, ls2 = st.columns(2)
                        with ls1:
                            largo_losa = st.number_input("Largo losa (m)", value=0.0, key="ls_largo")
                        with ls2:
                            ancho_losa = st.number_input("Ancho losa (m)", value=0.0, key="ls_ancho")
                        area_losa = largo_losa * ancho_losa
                        kg_acero_losa = area_losa * 8
                        diametro_losa = st.selectbox("Diámetro de barra", list(PESO_BARRAS.keys()), key="diam_losa")
                        largo_barra_losa = st.selectbox("Largo de barra", ["6m", "12m"], key="largo_losa")
                        largo_metros_losa = float(largo_barra_losa.replace("m", ""))
                        kg_por_barra_losa = PESO_BARRAS[diametro_losa] * largo_metros_losa
                        cant_barras_losa = kg_acero_losa / kg_por_barra_losa if kg_por_barra_losa > 0 else 0
                        st.info(f"Área losa: {area_losa:.2f} m²")
                        st.text(f"Acero estimado: {kg_acero_losa:.1f} kg")
                        st.text(f"Cantidad de barras {diametro_losa}: {cant_barras_losa:.0f} barras de {largo_barra_losa}")
                        if kg_acero_losa > 0:
                            registrar_pdf("Acero estructural", "Losa", [
                                ("Acero total", f"{kg_acero_losa:.1f} kg"),
                                ("Barras", f"{cant_barras_losa:.0f} de {diametro_losa} ({largo_barra_losa})"),
                            ])
                    elif modo_losa == "📐 Modo Detallado":
                        st.caption("Cálculo por barras en ambas direcciones")
                        ld1, ld2 = st.columns(2)
                        with ld1:
                            largo_losa_d = st.number_input("Largo losa (m)", value=0.0, key="ld_largo")
                        with ld2:
                            ancho_losa_d = st.number_input("Ancho losa (m)", value=0.0, key="ld_ancho")
                        st.write("**Barras dirección X (largo)**")
                        dx1, dx2, dx3 = st.columns(3)
                        with dx1:
                            diam_x = st.selectbox("Diámetro", list(PESO_BARRAS.keys()), key="diam_x")
                        with dx2:
                            sep_x = st.selectbox("Separación (m)", [0.10, 0.15, 0.20, 0.25], key="sep_x")
                        with dx3:
                            largo_barra_x = st.selectbox("Largo barra", ["6m", "12m"], key="largo_x")
                        st.write("**Barras dirección Y (ancho)**")
                        dy1, dy2, dy3 = st.columns(3)
                        with dy1:
                            diam_y = st.selectbox("Diámetro", list(PESO_BARRAS.keys()), key="diam_y")
                        with dy2:
                            sep_y = st.selectbox("Separación (m)", [0.10, 0.15, 0.20, 0.25], key="sep_y")
                        with dy3:
                            largo_barra_y = st.selectbox("Largo barra", ["6m", "12m"], key="largo_y")
                        cant_barras_x = ancho_losa_d / sep_x if sep_x > 0 else 0
                        largo_m_x = float(largo_barra_x.replace("m", ""))
                        kg_x = cant_barras_x * largo_losa_d * PESO_BARRAS[diam_x]
                        barras_x = (cant_barras_x * largo_losa_d) / largo_m_x
                        cant_barras_y = largo_losa_d / sep_y if sep_y > 0 else 0
                        largo_m_y = float(largo_barra_y.replace("m", ""))
                        kg_y = cant_barras_y * ancho_losa_d * PESO_BARRAS[diam_y]
                        barras_y = (cant_barras_y * ancho_losa_d) / largo_m_y
                        kg_total_losa = kg_x + kg_y
                        st.write("---")
                        st.info(f"Acero total losa: {kg_total_losa:.1f} kg")
                        st.text(f"Dirección X: {barras_x:.0f} barras {diam_x} de {largo_barra_x}")
                        st.text(f"Dirección Y: {barras_y:.0f} barras {diam_y} de {largo_barra_y}")
                        st.text(f"Kg dirección X: {kg_x:.1f} kg")
                        st.text(f"Kg dirección Y: {kg_y:.1f} kg")
                        if kg_total_losa > 0:
                            registrar_pdf("Acero estructural", "Losa", [
                                ("Acero total", f"{kg_total_losa:.1f} kg"),
                                ("Dirección X", f"{barras_x:.0f} barras {diam_x} de {largo_barra_x}"),
                                ("Dirección Y", f"{barras_y:.0f} barras {diam_y} de {largo_barra_y}"),
                            ])

            if ver(acero, "pilar"):
                with st.expander("2. Pilar", expanded=False):
                    modo_pilar = st.radio(
                        "Modo de cálculo",
                        ["🔨 Modo Simple", "📐 Modo Detallado"],
                        horizontal=True, key="modo_pilar"
                    )
                    if modo_pilar == "🔨 Modo Simple":
                        st.caption("Cálculo por barras longitudinales y estribos")
                        p1, p2, p3 = st.columns(3)
                        with p1:
                            cant_pilares = st.number_input("Cantidad de pilares", value=1, step=1, key="cant_pil")
                        with p2:
                            alto_pilar_a = st.number_input("Alto pilar (m)", value=0.0, key="alto_pil")
                        with p3:
                            barras_long = st.selectbox("Barras longitudinales", [4, 6, 8], key="barras_long")
                        p4, p5, p6 = st.columns(3)
                        with p4:
                            diam_long = st.selectbox("Diámetro barra long.", list(PESO_BARRAS.keys()), index=2, key="diam_long")
                        with p5:
                            sep_estribo = st.selectbox("Separación estribos (m)", ["0.10", "0.15", "0.20"], key="sep_estribo")
                            sep_estribo = float(sep_estribo)
                        with p6:
                            diam_estribo = st.selectbox("Diámetro estribo", list(PESO_BARRAS.keys()), index=0, key="diam_estribo")
                        p7, p8 = st.columns(2)
                        with p7:
                            ancho_pilar_a = st.number_input("Ancho pilar (m)", value=0.0, key="ancho_pil")
                        with p8:
                            largo_pilar_a = st.number_input("Largo pilar (m)", value=0.0, key="largo_pil")
                        ml_long = barras_long * alto_pilar_a * cant_pilares
                        kg_long = ml_long * PESO_BARRAS[diam_long]
                        n_estribos = (alto_pilar_a / sep_estribo) * cant_pilares
                        perimetro_estribo = ((ancho_pilar_a + largo_pilar_a) * 2) + 0.20
                        ml_estribos = n_estribos * perimetro_estribo
                        kg_estribos = ml_estribos * PESO_BARRAS[diam_estribo]
                        kg_total_pilar = kg_long + kg_estribos
                        st.write("---")
                        st.info(f"Acero total pilares: {kg_total_pilar:.1f} kg")
                        st.text(f"Barras long.: {barras_long * cant_pilares} barras {diam_long} x {alto_pilar_a}m")
                        st.text(f"Estribos: {n_estribos:.0f} estribos {diam_estribo} c/{sep_estribo}m")
                        st.text(f"Kg longitudinal: {kg_long:.1f} kg")
                        st.text(f"Kg estribos: {kg_estribos:.1f} kg")
                        if kg_total_pilar > 0:
                            registrar_pdf("Acero estructural", "Pilar", [
                                ("Acero total", f"{kg_total_pilar:.1f} kg"),
                                ("Barras long.", f"{barras_long * cant_pilares} {diam_long} x {alto_pilar_a}m"),
                                ("Estribos", f"{n_estribos:.0f} {diam_estribo} c/{sep_estribo}m"),
                            ])
                    elif modo_pilar == "📐 Modo Detallado":
                        st.caption("Cálculo exacto por barra")
                        pd1, pd2 = st.columns(2)
                        with pd1:
                            cant_pil_d = st.number_input("Cantidad de pilares", value=1, step=1, key="cant_pil_d")
                        with pd2:
                            alto_pil_d = st.number_input("Alto pilar (m)", value=0.0, key="alto_pil_d")
                        st.write("**Barras longitudinales**")
                        bl1, bl2 = st.columns(2)
                        with bl1:
                            cant_bl = st.number_input("Cantidad barras", value=4, step=1, key="cant_bl")
                        with bl2:
                            diam_bl = st.selectbox("Diámetro", list(PESO_BARRAS.keys()), index=2, key="diam_bl")
                        st.write("**Estribos**")
                        be1, be2, be3 = st.columns(3)
                        with be1:
                            ancho_p = st.number_input("Ancho pilar (m)", value=0.0, key="ancho_p_d")
                        with be2:
                            largo_p = st.number_input("Largo pilar (m)", value=0.0, key="largo_p_d")
                        with be3:
                            sep_be = st.selectbox("Separación (m)", ["0.10", "0.15", "0.20"], key="sep_be")
                        sep_be = float(sep_be)
                        diam_be = st.selectbox("Diámetro estribo", list(PESO_BARRAS.keys()), index=0, key="diam_be")
                        ml_bl = cant_bl * alto_pil_d * cant_pil_d
                        kg_bl = ml_bl * PESO_BARRAS[diam_bl]
                        n_estribos_d = (alto_pil_d / sep_be) * cant_pil_d
                        perimetro_d = ((ancho_p + largo_p) * 2) + 0.20
                        kg_be = n_estribos_d * perimetro_d * PESO_BARRAS[diam_be]
                        kg_total_d = kg_bl + kg_be
                        st.write("---")
                        st.info(f"Acero total pilares: {kg_total_d:.1f} kg")
                        st.text(f"Barras long.: {cant_bl * cant_pil_d} barras {diam_bl} x {alto_pil_d}m → {kg_bl:.1f} kg")
                        st.text(f"Estribos: {n_estribos_d:.0f} estribos {diam_be} c/{sep_be}m → {kg_be:.1f} kg")
                        if kg_total_d > 0:
                            registrar_pdf("Acero estructural", "Pilar", [
                                ("Acero total", f"{kg_total_d:.1f} kg"),
                                ("Barras long.", f"{cant_bl * cant_pil_d} {diam_bl} x {alto_pil_d}m"),
                                ("Estribos", f"{n_estribos_d:.0f} {diam_be} c/{sep_be}m"),
                            ])

            if ver(acero, "viga"):
                    with st.expander("3. Viga", expanded=False):
                        modo_viga = st.radio(
                            "Modo de cálculo",
                            ["🔨 Modo Simple", "📐 Modo Detallado"],
                            horizontal=True, key="modo_viga"
                        )
                        if modo_viga == "🔨 Modo Simple":
                            st.caption("Estimación por ratio kg/m³")
                            v1, v2, v3 = st.columns(3)
                            with v1:
                                cant_vigas = st.number_input("Cantidad de vigas", value=0, step=1, key="cant_vigas")
                            with v2:
                                largo_viga = st.number_input("Largo viga (m)", value=0.0, key="largo_viga")
                            with v3:
                                alto_viga = st.number_input("Alto viga (m)", value=0.0, key="alto_viga")
                            ancho_viga = st.number_input("Ancho viga (m)", value=0.0, key="ancho_viga")
                            vol_viga = cant_vigas * largo_viga * alto_viga * ancho_viga
                            kg_acero_viga = vol_viga * 120
                            diam_viga_s = st.selectbox("Diámetro de barra", list(PESO_BARRAS.keys()), key="diam_viga_s")
                            largo_barra_viga_s = st.selectbox("Largo de barra", ["6m", "12m"], key="largo_viga_s")
                            largo_metros_viga_s = float(largo_barra_viga_s.replace("m", ""))
                            kg_por_barra_viga = PESO_BARRAS[diam_viga_s] * largo_metros_viga_s
                            cant_barras_viga = kg_acero_viga / kg_por_barra_viga if kg_por_barra_viga > 0 else 0
                            st.write("---")
                            st.info(f"Acero estimado: {kg_acero_viga:.1f} kg")
                            st.text(f"Cantidad de barras {diam_viga_s}: {cant_barras_viga:.0f} barras de {largo_barra_viga_s}")
                            st.caption(f"Volumen vigas: {vol_viga:.2f} m³ | Ratio: 120 kg/m³")
                            if kg_acero_viga > 0:
                                registrar_pdf("Acero estructural", "Viga", [
                                    ("Acero total", f"{kg_acero_viga:.1f} kg"),
                                    ("Barras", f"{cant_barras_viga:.0f} de {diam_viga_s} ({largo_barra_viga_s})"),
                                ])
                        elif modo_viga == "📐 Modo Detallado":
                            st.caption("Cálculo por barras superiores, inferiores y estribos")
                            vd1, vd2, vd3 = st.columns(3)
                            with vd1:
                                cant_vigas_d = st.number_input("Cantidad de vigas", value=0, step=1, key="cant_vigas_d")
                            with vd2:
                                largo_viga_d = st.number_input("Largo viga (m)", value=0.0, key="largo_viga_d")
                            with vd3:
                                ancho_viga_d = st.number_input("Ancho viga (m)", value=0.0, key="ancho_viga_d")
                            st.write("**Barras superiores**")
                            vs1, vs2 = st.columns(2)
                            with vs1:
                                cant_sup = st.number_input("Cantidad barras sup.", value=0, step=1, key="cant_sup")
                            with vs2:
                                diam_sup = st.selectbox("Diámetro", list(PESO_BARRAS.keys()), key="diam_sup")
                            st.write("**Barras inferiores**")
                            vi1, vi2 = st.columns(2)
                            with vi1:
                                cant_inf = st.number_input("Cantidad barras inf.", value=0, step=1, key="cant_inf")
                            with vi2:
                                diam_inf = st.selectbox("Diámetro", list(PESO_BARRAS.keys()), key="diam_inf")
                            st.write("**Estribos**")
                            ve1, ve2 = st.columns(2)
                            with ve1:
                                sep_estribo_v = st.selectbox("Separación (m)", ["0.10", "0.15", "0.20"], key="sep_estribo_v")
                            with ve2:
                                diam_estribo_v = st.selectbox("Diámetro estribo", list(PESO_BARRAS.keys()), index=0, key="diam_estribo_v")
                            sep_estribo_v = float(sep_estribo_v)
                            ml_sup = cant_sup * largo_viga_d * cant_vigas_d
                            kg_sup = ml_sup * PESO_BARRAS[diam_sup]
                            ml_inf = cant_inf * largo_viga_d * cant_vigas_d
                            kg_inf = ml_inf * PESO_BARRAS[diam_inf]
                            n_estribos_v = (largo_viga_d / sep_estribo_v) * cant_vigas_d
                            perimetro_estribo_v = ((ancho_viga_d + 0.30) * 2) + 0.20
                            kg_estribo_v = n_estribos_v * perimetro_estribo_v * PESO_BARRAS[diam_estribo_v]
                            kg_total_viga = kg_sup + kg_inf + kg_estribo_v
                            st.write("---")
                            st.info(f"Acero total vigas: {kg_total_viga:.1f} kg")
                            st.text(f"Barras sup.: {cant_sup * cant_vigas_d} barras {diam_sup} → {kg_sup:.1f} kg")
                            st.text(f"Barras inf.: {cant_inf * cant_vigas_d} barras {diam_inf} → {kg_inf:.1f} kg")
                            st.text(f"Estribos: {n_estribos_v:.0f} estribos {diam_estribo_v} c/{sep_estribo_v}m → {kg_estribo_v:.1f} kg")
                            if kg_total_viga > 0:
                                registrar_pdf("Acero estructural", "Viga", [
                                    ("Acero total", f"{kg_total_viga:.1f} kg"),
                                    ("Barras sup.", f"{cant_sup * cant_vigas_d} {diam_sup}"),
                                    ("Barras inf.", f"{cant_inf * cant_vigas_d} {diam_inf}"),
                                    ("Estribos", f"{n_estribos_v:.0f} {diam_estribo_v} c/{sep_estribo_v}m"),
                                ])

            if ver(acero, "radier"):
                    with st.expander("4. Radier", expanded=False):
                        modo_radier = st.radio(
                            "Modo de cálculo",
                            ["🔨 Modo Simple", "📐 Modo Detallado"],
                            horizontal=True, key="modo_radier"
                        )
                        if modo_radier == "🔨 Modo Simple":
                            st.caption("Estimación por ratio kg/m²")
                            rr1, rr2 = st.columns(2)
                            with rr1:
                                largo_rad_a = st.number_input("Largo radier (m)", value=0.0, key="largo_rad_a")
                            with rr2:
                                ancho_rad_a = st.number_input("Ancho radier (m)", value=0.0, key="ancho_rad_a")
                            area_rad = largo_rad_a * ancho_rad_a
                            kg_acero_rad = area_rad * 5
                            diam_rad_s = st.selectbox("Diámetro de barra", list(PESO_BARRAS.keys()), key="diam_rad_s")
                            largo_barra_rad = st.selectbox("Largo de barra", ["6m", "12m"], key="largo_rad_s")
                            largo_metros_rad = float(largo_barra_rad.replace("m", ""))
                            kg_por_barra_rad = PESO_BARRAS[diam_rad_s] * largo_metros_rad
                            cant_barras_rad = kg_acero_rad / kg_por_barra_rad if kg_por_barra_rad > 0 else 0
                            st.write("---")
                            st.info(f"Acero estimado: {kg_acero_rad:.1f} kg")
                            st.text(f"Cantidad de barras {diam_rad_s}: {cant_barras_rad:.0f} barras de {largo_barra_rad}")
                            st.caption(f"Área radier: {area_rad:.2f} m² | Ratio: 5 kg/m²")
                            if kg_acero_rad > 0:
                                registrar_pdf("Acero estructural", "Radier", [
                                    ("Acero total", f"{kg_acero_rad:.1f} kg"),
                                    ("Barras", f"{cant_barras_rad:.0f} de {diam_rad_s} ({largo_barra_rad})"),
                                ])
                        elif modo_radier == "📐 Modo Detallado":
                            st.caption("Cálculo por barras en ambas direcciones")
                            rd1, rd2 = st.columns(2)
                            with rd1:
                                largo_rad_d = st.number_input("Largo radier (m)", value=0.0, key="largo_rad_d")
                            with rd2:
                                ancho_rad_d = st.number_input("Ancho radier (m)", value=0.0, key="ancho_rad_d")
                            st.write("**Barras dirección X (largo)**")
                            rx1, rx2, rx3 = st.columns(3)
                            with rx1:
                                diam_rx = st.selectbox("Diámetro", list(PESO_BARRAS.keys()), key="diam_rx")
                            with rx2:
                                sep_rx = st.selectbox("Separación (m)", ["0.10", "0.15", "0.20", "0.25"], key="sep_rx")
                            with rx3:
                                largo_barra_rx = st.selectbox("Largo barra", ["6m", "12m"], key="largo_rx")
                            sep_rx = float(sep_rx)
                            st.write("**Barras dirección Y (ancho)**")
                            ry1, ry2, ry3 = st.columns(3)
                            with ry1:
                                diam_ry = st.selectbox("Diámetro", list(PESO_BARRAS.keys()), key="diam_ry")
                            with ry2:
                                sep_ry = st.selectbox("Separación (m)", ["0.10", "0.15", "0.20", "0.25"], key="sep_ry")
                            with ry3:
                                largo_barra_ry = st.selectbox("Largo barra", ["6m", "12m"], key="largo_ry")
                            sep_ry = float(sep_ry)
                            cant_barras_rx = ancho_rad_d / sep_rx if sep_rx > 0 else 0
                            largo_m_rx = float(largo_barra_rx.replace("m", ""))
                            kg_rx = cant_barras_rx * largo_rad_d * PESO_BARRAS[diam_rx]
                            barras_rx = (cant_barras_rx * largo_rad_d) / largo_m_rx
                            cant_barras_ry = largo_rad_d / sep_ry if sep_ry > 0 else 0
                            largo_m_ry = float(largo_barra_ry.replace("m", ""))
                            kg_ry = cant_barras_ry * ancho_rad_d * PESO_BARRAS[diam_ry]
                            barras_ry = (cant_barras_ry * ancho_rad_d) / largo_m_ry
                            kg_total_rad = kg_rx + kg_ry
                            st.write("---")
                            st.info(f"Acero total radier: {kg_total_rad:.1f} kg")
                            st.text(f"Dirección X: {barras_rx:.0f} barras {diam_rx} de {largo_barra_rx} → {kg_rx:.1f} kg")
                            st.text(f"Dirección Y: {barras_ry:.0f} barras {diam_ry} de {largo_barra_ry} → {kg_ry:.1f} kg")
                            if kg_total_rad > 0:
                                registrar_pdf("Acero estructural", "Radier", [
                                    ("Acero total", f"{kg_total_rad:.1f} kg"),
                                    ("Dirección X", f"{barras_rx:.0f} barras {diam_rx} de {largo_barra_rx}"),
                                    ("Dirección Y", f"{barras_ry:.0f} barras {diam_ry} de {largo_barra_ry}"),
                                ])

            if ver(acero, "cimiento"):
                    with st.expander("5. Cimiento", expanded=False):
                        
                        modo_cimiento = st.radio(
                            "Modo de cálculo",
                            ["🔨 Modo Simple", "📐 Modo Detallado"],
                            horizontal=True,
                            key="modo_cimiento"
                        )
                                
                        if modo_cimiento == "🔨 Modo Simple":
                            st.caption("Estimación por ratio kg/m³")
                                    
                            ci1, ci2, ci3 = st.columns(3)
                            with ci1:
                                cant_cim = st.number_input("Cantidad de cimientos", value=0, step=1, key="cant_cim")
                            with ci2:
                                largo_cim = st.number_input("Largo cimiento (m)", value=0.0, key="largo_cim")
                            with ci3:
                                ancho_cim = st.number_input("Ancho cimiento (m)", value=0.0, key="ancho_cim")
                                    
                            alto_cim = st.number_input("Alto cimiento (m)", value=0.0, key="alto_cim")
                                    
                            vol_cim = cant_cim * largo_cim * ancho_cim * alto_cim
                            kg_acero_cim = vol_cim * 80  # ratio 80 kg/m³
                                    
                            diam_cim_s = st.selectbox("Diámetro de barra", list(PESO_BARRAS.keys()), key="diam_cim_s")
                            largo_barra_cim = st.selectbox("Largo de barra", ["6m", "12m"], key="largo_cim_s")
                            largo_metros_cim = float(largo_barra_cim.replace("m", ""))
                                    
                            kg_por_barra_cim = PESO_BARRAS[diam_cim_s] * largo_metros_cim
                            cant_barras_cim = kg_acero_cim / kg_por_barra_cim if kg_por_barra_cim > 0 else 0
                                    
                            st.write("---")
                            st.info(f"Acero estimado: {kg_acero_cim:.1f} kg")
                            st.text(f"Cantidad de barras {diam_cim_s}: {cant_barras_cim:.0f} barras de {largo_barra_cim}")
                            st.caption(f"Volumen cimiento: {vol_cim:.2f} m³ | Ratio: 80 kg/m³")
                            if kg_acero_cim > 0:
                                registrar_pdf("Acero estructural", "Cimiento", [
                                    ("Acero total", f"{kg_acero_cim:.1f} kg"),
                                    ("Barras", f"{cant_barras_cim:.0f} de {diam_cim_s} ({largo_barra_cim})"),
                                ])

                        elif modo_cimiento == "📐 Modo Detallado":
                            st.caption("Cálculo por barras longitudinales y estribos")
                                    
                            cd1, cd2, cd3 = st.columns(3)
                            with cd1:
                                cant_cim_d = st.number_input("Cantidad de cimientos", value=0, step=1, key="cant_cim_d")
                            with cd2:
                                largo_cim_d = st.number_input("Largo cimiento (m)", value=0.0, key="largo_cim_d")
                            with cd3:
                                ancho_cim_d = st.number_input("Ancho cimiento (m)", value=0.0, key="ancho_cim_d")

                            st.write("**Barras longitudinales**")
                            cl1, cl2 = st.columns(2)
                            with cl1:
                                cant_bl_cim = st.number_input("Cantidad barras", value=0, step=1, key="cant_bl_cim")
                            with cl2:
                                diam_bl_cim = st.selectbox("Diámetro", list(PESO_BARRAS.keys()), key="diam_bl_cim")

                            st.write("**Estribos**")
                            ce1, ce2 = st.columns(2)
                            with ce1:
                                sep_estribo_cim = st.selectbox("Separación (m)", ["0.10", "0.15", "0.20", "0.25"], key="sep_estribo_cim")
                            with ce2:
                                diam_estribo_cim = st.selectbox("Diámetro estribo", list(PESO_BARRAS.keys()), index=0, key="diam_estribo_cim")
                            sep_estribo_cim = float(sep_estribo_cim)

                            # Cálculo barras longitudinales
                            ml_bl_cim = cant_bl_cim * largo_cim_d * cant_cim_d
                            kg_bl_cim = ml_bl_cim * PESO_BARRAS[diam_bl_cim]

                            # Cálculo estribos
                            n_estribos_cim = (largo_cim_d / sep_estribo_cim) * cant_cim_d
                            perimetro_estribo_cim = ((ancho_cim_d + 0.30) * 2) + 0.20
                            kg_estribo_cim = n_estribos_cim * perimetro_estribo_cim * PESO_BARRAS[diam_estribo_cim]

                            kg_total_cim = kg_bl_cim + kg_estribo_cim

                            st.write("---")
                            st.info(f"Acero total cimientos: {kg_total_cim:.1f} kg")
                            st.text(f"Barras long.: {cant_bl_cim * cant_cim_d} barras {diam_bl_cim} → {kg_bl_cim:.1f} kg")
                            st.text(f"Estribos: {n_estribos_cim:.0f} estribos {diam_estribo_cim} c/{sep_estribo_cim}m → {kg_estribo_cim:.1f} kg") 
                            if kg_total_cim > 0:
                                registrar_pdf("Acero estructural", "Cimiento", [
                                    ("Acero total", f"{kg_total_cim:.1f} kg"),
                                    ("Barras long.", f"{cant_bl_cim * cant_cim_d} {diam_bl_cim}"),
                                    ("Estribos", f"{n_estribos_cim:.0f} {diam_estribo_cim} c/{sep_estribo_cim}m"),
                                ])
                

    # --- Acero No estructural ---
    if ver_rubro(metal) and rubro_permitido("metalcon"):
        with st.expander("Acero No Estructural (Tabiques Metalcon)", expanded=False):

            # ============================
            # DATOS SEGÚN MANUAL METALCON
            # ============================
            MONTANTES = {
                # Perforados
                "Montante Normal Perf. 60x38x0,5 - 2,40m":  {"largo": 2.40, "peso": 0.56},
                "Montante Normal Perf. 60x38x0,5 - 3,00m":  {"largo": 3.00, "peso": 0.56},
                # Económico (sin perforar)
                "Montante Económico 38x38x0,5 - 2,40m": {"largo": 2.40, "peso": 0.48},
                "Montante Económico 38x38x0,5 - 3,00m": {"largo": 3.00, "peso": 0.48},
            }

            CANALES = {
                "Canal Normal 61x20x0,5 - 2,40m":    {"largo": 2.40, "peso": 0.39},
                "Canal Normal 61x20x0,5 - 3,00m":    {"largo": 3.00, "peso": 0.39},
                "Canal Económico 39x20x0,5 - 2,40m": {"largo": 2.40, "peso": 0.31},
                "Canal Económico 39x20x0,5 - 3,00m": {"largo": 3.00, "peso": 0.31},
            }

            ESQUINEROS = {
                "Esquinero Perf. 30x30 - 2,40m":     {"largo": 2.40, "peso": 0.18},
                "Esquinero Perf. Eco. 25x25 - 3,00m": {"largo": 3.00, "peso": 0.15},
            }

# --- 1. Canal / Solera ---
            if ver(metal, "canal"):
                    with st.expander("1. Canal / Solera", expanded=False):
                        st.caption("Se usa como solera en piso y cielo del tabique")

                        canal_tipo = st.selectbox("Tipo de canal", list(CANALES.keys()), key="canal_tipo")

                        ca1, ca2 = st.columns(2)
                        with ca1:
                            largo_tabique_c = st.number_input("Largo del tabique (m)", value=0.0, key="largo_tab_c")
                        with ca2:
                            cant_tabiques_c = st.number_input("Cantidad de tabiques", value=0, step=1, key="cant_tab_c")

                        largo_canal = CANALES[canal_tipo]["largo"]

                        # Solera inf + sup = largo tabique * 2 * cantidad tabiques
                        ml_canal = largo_tabique_c * 2 * cant_tabiques_c
                        cant_piezas_canal = ml_canal / largo_canal if largo_canal > 0 else 0

                        # Desperdicio solera
                        ultimo_tramo = ml_canal % largo_canal
                        desperdicio_canal = (largo_canal - ultimo_tramo) if ultimo_tramo > 0 else 0
                                
                        # Sugerencia canal óptimo
                        desperdicio_canal_240 = None
                        desperdicio_canal_300 = None
                                
                        if largo_tabique_c > 0:
                            ml_total_240 = largo_tabique_c * 2 * (cant_tabiques_c if cant_tabiques_c > 0 else 1)
                            ml_total_300 = ml_total_240
                            sobra_240 = ml_total_240 % 2.40
                            sobra_300 = ml_total_300 % 3.00
                            desperdicio_canal_240 = (2.40 - sobra_240) if sobra_240 > 0 else 0
                            desperdicio_canal_300 = (3.00 - sobra_300) if sobra_300 > 0 else 0

                        st.write("---")
                        st.info(f"Metros lineales necesarios: {ml_canal:.2f} ml")
                        st.text(f"Cantidad de canales {largo_canal}m: {cant_piezas_canal:.0f} piezas")

                        if largo_tabique_c > 0:
                            st.write("---")
                            st.subheader("📐 Análisis de desperdicio")
                            st.text(f"Desperdicio último tramo: {desperdicio_canal:.2f}m")

                            st.write("**💡 Sugerencia de canal más conveniente:**")
                            if desperdicio_canal_240 is not None and desperdicio_canal_300 is not None:
                                if desperdicio_canal_240 <= desperdicio_canal_300:
                                    st.success(f"✅ Usa canal de 2,40m → desperdicio {desperdicio_canal_240:.2f}m")
                                    st.warning(f"Con 3,00m → desperdicio {desperdicio_canal_300:.2f}m")
                                else:
                                    st.success(f"✅ Usa canal de 3,00m → desperdicio {desperdicio_canal_300:.2f}m")
                                    st.warning(f"Con 2,40m → desperdicio {desperdicio_canal_240:.2f}m")

                        st.caption("Considera solera inferior + solera superior")

# --- 2. Montante / Pie Derecho ---
            if ver(metal, "montante"):
                    with st.expander("2. Montante / Pie Derecho", expanded=False):
                        st.caption("Se instala cada 40 o 60 cm según revestimiento")

                        montante_tipo = st.selectbox("Tipo de montante", list(MONTANTES.keys()), key="montante_tipo")

                        if "Perf" in montante_tipo:
                            st.caption("✅ Perforado: permite pasar instalaciones eléctricas y sanitarias por dentro")
                        else:
                            st.caption("⚠️ Sin perforaciones: usar cuando no pasan instalaciones por el tabique")

                        mo1, mo2, mo3 = st.columns(3)
                        with mo1:
                            largo_tabique_m = st.number_input("Largo tabique (m)", value=0.0, key="largo_tab_m")
                        with mo2:
                            alto_tabique_m = st.number_input("Alto tabique (m)", value=0.0, key="alto_tab_m")
                        with mo3:
                            cant_tabiques_m = st.number_input("Cantidad tabiques", value=0, step=1, key="cant_tab_m")

                        separacion_m = st.selectbox("Separación entre montantes",
                                                    ["0,40m (recomendado)", "0,60m (máximo)"],
                                                    key="sep_mont")
                        sep_valor = 0.40 if "0,40" in separacion_m else 0.60

                        largo_montante = MONTANTES[montante_tipo]["largo"]

                        # Cálculo montantes
                        montantes_por_tabique = int(largo_tabique_m / sep_valor) + 1
                        total_montantes = montantes_por_tabique * cant_tabiques_m

                        # Desperdicio
                        corte_por_perfil = largo_montante - alto_tabique_m
                        desperdicio_total = corte_por_perfil * total_montantes
                        perfiles_desperdiciados = desperdicio_total / largo_montante if largo_montante > 0 else 0

                        # Sugerencia de perfil óptimo
                        desperdicio_240 = 2.40 - alto_tabique_m if alto_tabique_m <= 2.40 else None
                        desperdicio_300 = 3.00 - alto_tabique_m if alto_tabique_m <= 3.00 else None

                        st.write("---")
                        st.info(f"Montantes por tabique: {montantes_por_tabique} piezas")
                        st.info(f"Total montantes: {total_montantes} piezas de {largo_montante}m")

                        if alto_tabique_m > 0 and largo_montante > 0:
                            if corte_por_perfil < 0:
                                st.error(f"⚠️ El tabique ({alto_tabique_m}m) supera el montante ({largo_montante}m). Necesita empalme.")
                            else:
                                st.write("---")
                                st.subheader("📐 Análisis de desperdicio")
                                st.text(f"Corte por perfil: {corte_por_perfil:.2f}m")
                                st.text(f"Desperdicio total: {desperdicio_total:.2f}m lineales")
                                st.text(f"Equivale a: {perfiles_desperdiciados:.1f} perfiles desperdiciados")

                                # Sugerencia perfil óptimo
                                st.write("**💡 Sugerencia de perfil más conveniente:**")
                                if desperdicio_240 is not None and desperdicio_300 is not None:
                                    if desperdicio_240 <= desperdicio_300:
                                        st.success(f"✅ Usa montante de 2,40m → sobran {desperdicio_240:.2f}m por perfil")
                                        st.warning(f"Con 3,00m → sobrarían {desperdicio_300:.2f}m por perfil")
                                    else:
                                        st.success(f"✅ Usa montante de 3,00m → sobran {desperdicio_300:.2f}m por perfil")
                                        st.warning(f"Con 2,40m → sobrarían {desperdicio_240:.2f}m por perfil")
                                elif desperdicio_240 is None:
                                    st.success(f"✅ Usa montante de 3,00m → sobran {desperdicio_300:.2f}m por perfil")
                                    st.error("❌ Montante de 2,40m no alcanza para este tabique")
                                        
                        st.caption(f"Separación: {sep_valor}m según manual Metalcon | +1 montante en extremo")

# --- 3. Esquinero ---
            if ver(metal, "esquinero"):
                    with st.expander("3. Esquinero", expanded=False):
                        st.caption("Se usa en encuentros de muros y esquinas")

                        esq_tipo = st.selectbox("Tipo de esquinero", list(ESQUINEROS.keys()), key="esq_tipo")
                        cant_esquinas = st.number_input("Cantidad de esquinas/encuentros", value=0, step=1, key="cant_esq")
                                
                        largo_esq = ESQUINEROS[esq_tipo]["largo"]
                                
                        st.write("---")
                        st.info(f"Cantidad de esquineros: {cant_esquinas} piezas de {largo_esq}m") 

                        # Guardar en session_state para el PDF
                        st.session_state["pdf_canal_tipo"] = canal_tipo
                        st.session_state["pdf_cant_piezas_canal"] = cant_piezas_canal
                        st.session_state["pdf_ml_canal"] = ml_canal
                        st.session_state["pdf_largo_canal"] = largo_canal

                        st.session_state["pdf_montante_tipo"] = montante_tipo
                        st.session_state["pdf_total_montantes"] = total_montantes
                        st.session_state["pdf_largo_montante"] = largo_montante

                        st.session_state["pdf_esq_tipo"] = esq_tipo
                        st.session_state["pdf_cant_esquinas"] = cant_esquinas
                        st.session_state["pdf_largo_esq"] = largo_esq

# ============================
# MOLDAJES
# ============================
    if ver_rubro(mold) and rubro_permitido("moldajes"):
        with st.expander("Moldajes", expanded=False):

            # Datos materiales
            MATERIALES_MOLDAJE = {
                "Tabla 1\"x8\" (ancho 19cm)": {"ancho": 0.19, "largo": 3.20},
                "Tabla 1\"x10\" (ancho 24cm)": {"ancho": 0.24, "largo": 3.20},
                "Terciado Film 18mm (1,22x2,44m)": {"area_plancha": 2.98},
                "Moldaje Metálico": {"solo_m2": True},
            }

            def resultado_moldaje(partida, material, m2):
                """Devuelve el item de material principal para el PDF según el tipo."""
                mat_d = MATERIALES_MOLDAJE[material]
                if "solo_m2" in mat_d:
                    return ("Moldaje metálico", f"{m2:.2f} m²")
                elif "area_plancha" in mat_d:
                    planchas = (m2 / mat_d["area_plancha"]) * 1.10
                    return ("Planchas terciado (10% desp.)", f"{planchas:.0f} unidades")
                else:
                    tablas = (m2 / mat_d["ancho"] / mat_d["largo"]) * 1.10
                    return (f"Tablas {mat_d['largo']}m (10% desp.)", f"{tablas:.0f} unidades")

# --- 1. Cimiento ---
            if ver(mold, "cimiento"):
                    with st.expander("1. Moldaje de Cimiento", expanded=False):
                        st.caption("Moldaje para caras laterales del cimiento")

                        material_cim = st.selectbox(
                            "Material de moldaje",
                            list(MATERIALES_MOLDAJE.keys()),
                            key="mat_mold_cim"
                        )

                        st.markdown("**Cimientos** (agrega cada uno con su medida)")
                        _sec_cim = secciones_input(
                            "mold_cim", [("largo", "Largo (m)"), ("alto", "Alto (m)")], "cimiento")
                        # m² = suma de (largo * alto * 2 caras) de cada cimiento
                        m2_cimiento = sum(s["largo"] * s["alto"] * 2 for s in _sec_cim)

                        st.write("---")
                        st.info(f"Superficie de moldaje: {m2_cimiento:.2f} m²")

                        mat = MATERIALES_MOLDAJE[material_cim]

                        if "solo_m2" in mat:
                            st.text(f"Moldaje metálico requerido: {m2_cimiento:.2f} m²")

                        elif "area_plancha" in mat:
                            cant_planchas = m2_cimiento / mat["area_plancha"]
                            st.text(f"Planchas terciado 1,22x2,44m: {cant_planchas:.0f} unidades")
                            st.caption("Considera un 10% de desperdicio por cortes")
                            cant_planchas_real = cant_planchas * 1.10
                            st.text(f"Con 10% desperdicio: {cant_planchas_real:.0f} planchas")

                        else:
                            # Tabla
                            ml_tabla = m2_cimiento / mat["ancho"]
                            cant_tablas = ml_tabla / mat["largo"]
                            st.text(f"Metros lineales de tabla: {ml_tabla:.2f} ml")
                            st.text(f"Cantidad de tablas de {mat['largo']}m: {cant_tablas:.0f} unidades")
                            st.caption("Considera un 10% de desperdicio por cortes")
                            cant_tablas_real = cant_tablas * 1.10
                            st.text(f"Con 10% desperdicio: {cant_tablas_real:.0f} tablas")   
                        if m2_cimiento > 0:
                            registrar_pdf("Moldajes", "Cimiento", [
                                ("Material", material_cim),
                                ("Superficie", f"{m2_cimiento:.2f} m²"),
                                resultado_moldaje("Cimiento", material_cim, m2_cimiento),
                            ])

# --- 2. Moldaje de Muro ---
            if ver(mold, "muro"):
                    with st.expander("2. Moldaje de Muro", expanded=False):
                        st.caption("Moldaje para ambas caras del muro")

                        material_muro = st.selectbox(
                            "Material de moldaje",
                            list(MATERIALES_MOLDAJE.keys()),
                            key="mat_mold_muro"
                        )

                        st.markdown("**Muros** (agrega cada uno con su medida)")
                        _sec_muro_m = secciones_input(
                            "mold_muro", [("largo", "Largo (m)"), ("alto", "Alto (m)")], "muro")
                        # m² = suma de (largo * alto * 2 caras) de cada muro
                        m2_muro = sum(s["largo"] * s["alto"] * 2 for s in _sec_muro_m)

                        st.write("---")
                        st.info(f"Superficie de moldaje: {m2_muro:.2f} m²")

                        mat_m = MATERIALES_MOLDAJE[material_muro]

                        if "solo_m2" in mat_m:
                            st.text(f"Moldaje metálico requerido: {m2_muro:.2f} m²")
                        elif "area_plancha" in mat_m:
                            cant_planchas = m2_muro / mat_m["area_plancha"]
                            cant_planchas_real = cant_planchas * 1.10
                            st.text(f"Planchas terciado 1,22x2,44m: {cant_planchas:.0f} unidades")
                            st.text(f"Con 10% desperdicio: {cant_planchas_real:.0f} planchas")
                        else:
                            ml_tabla = m2_muro / mat_m["ancho"]
                            cant_tablas = ml_tabla / mat_m["largo"]
                            cant_tablas_real = cant_tablas * 1.10
                            st.text(f"Metros lineales de tabla: {ml_tabla:.2f} ml")
                            st.text(f"Cantidad de tablas de {mat_m['largo']}m: {cant_tablas:.0f} unidades")
                            st.text(f"Con 10% desperdicio: {cant_tablas_real:.0f} tablas")
                        if m2_muro > 0:
                            registrar_pdf("Moldajes", "Muro", [
                                ("Material", material_muro),
                                ("Superficie", f"{m2_muro:.2f} m²"),
                                resultado_moldaje("Muro", material_muro, m2_muro),
                            ])

# --- 3. Moldaje de Losa ---
            if ver(mold, "losa"):
                    with st.expander("3. Moldaje de Losa", expanded=False):
                        st.caption("Moldaje para cara inferior de la losa")

                        material_losa = st.selectbox(
                            "Material de moldaje",
                            list(MATERIALES_MOLDAJE.keys()),
                            key="mat_mold_losa"
                        )

                        st.markdown("**Losas** (agrega cada una con su medida)")
                        _sec_losa = secciones_input(
                            "mold_losa", [("largo", "Largo (m)"), ("ancho", "Ancho (m)")], "losa")
                        # m² = suma de (largo * ancho) de cada losa (solo cara inferior)
                        m2_losa = sum(s["largo"] * s["ancho"] for s in _sec_losa)

                        st.write("---")
                        st.info(f"Superficie de moldaje: {m2_losa:.2f} m²")

                        mat_l = MATERIALES_MOLDAJE[material_losa]

                        if "solo_m2" in mat_l:
                            st.text(f"Moldaje metálico requerido: {m2_losa:.2f} m²")
                        elif "area_plancha" in mat_l:
                            cant_planchas = m2_losa / mat_l["area_plancha"]
                            cant_planchas_real = cant_planchas * 1.10
                            st.text(f"Planchas terciado 1,22x2,44m: {cant_planchas:.0f} unidades")
                            st.text(f"Con 10% desperdicio: {cant_planchas_real:.0f} planchas")
                        else:
                            ml_tabla = m2_losa / mat_l["ancho"]
                            cant_tablas = ml_tabla / mat_l["largo"]
                            cant_tablas_real = cant_tablas * 1.10
                            st.text(f"Metros lineales de tabla: {ml_tabla:.2f} ml")
                            st.text(f"Cantidad de tablas de {mat_l['largo']}m: {cant_tablas:.0f} unidades")
                            st.text(f"Con 10% desperdicio: {cant_tablas_real:.0f} tablas")
                        if m2_losa > 0:
                            registrar_pdf("Moldajes", "Losa", [
                                ("Material", material_losa),
                                ("Superficie", f"{m2_losa:.2f} m²"),
                                resultado_moldaje("Losa", material_losa, m2_losa),
                            ])

# --- 4. Moldaje de Viga ---
            if ver(mold, "viga"):
                    with st.expander("4. Moldaje de Viga", expanded=False):
                        st.caption("Moldaje para fondo y caras laterales de la viga")

                        material_viga = st.selectbox(
                            "Material de moldaje",
                            list(MATERIALES_MOLDAJE.keys()),
                            key="mat_mold_viga"
                        )

                        st.markdown("**Vigas** (agrega cada una con su medida)")
                        _sec_viga = secciones_input(
                            "mold_viga",
                            [("largo", "Largo (m)"), ("ancho", "Ancho (m)"), ("alto", "Alto (m)")], "viga")
                        # m² = suma de (fondo + 2 caras laterales) * largo de cada viga
                        m2_viga = sum((s["ancho"] + s["alto"] * 2) * s["largo"] for s in _sec_viga)

                        st.write("---")
                        st.info(f"Superficie de moldaje: {m2_viga:.2f} m²")

                        mat_v = MATERIALES_MOLDAJE[material_viga]

                        if "solo_m2" in mat_v:
                            st.text(f"Moldaje metálico requerido: {m2_viga:.2f} m²")
                        elif "area_plancha" in mat_v:
                            cant_planchas = m2_viga / mat_v["area_plancha"]
                            cant_planchas_real = cant_planchas * 1.10
                            st.text(f"Planchas terciado 1,22x2,44m: {cant_planchas:.0f} unidades")
                            st.text(f"Con 10% desperdicio: {cant_planchas_real:.0f} planchas")
                        else:
                            ml_tabla = m2_viga / mat_v["ancho"]
                            cant_tablas = ml_tabla / mat_v["largo"]
                            cant_tablas_real = cant_tablas * 1.10
                            st.text(f"Metros lineales de tabla: {ml_tabla:.2f} ml")
                            st.text(f"Cantidad de tablas de {mat_v['largo']}m: {cant_tablas:.0f} unidades")
                            st.text(f"Con 10% desperdicio: {cant_tablas_real:.0f} tablas")
                        if m2_viga > 0:
                            registrar_pdf("Moldajes", "Viga", [
                                ("Material", material_viga),
                                ("Superficie", f"{m2_viga:.2f} m²"),
                                resultado_moldaje("Viga", material_viga, m2_viga),
                            ])

# --- 5. Moldaje de Pilar ---
            if ver(mold, "pilar"):
                    with st.expander("5. Moldaje de Pilar", expanded=False):
                        st.caption("Moldaje para las 4 caras del pilar")

                        material_pilar = st.selectbox(
                            "Material de moldaje",
                            list(MATERIALES_MOLDAJE.keys()),
                            key="mat_mold_pilar"
                        )

                        mp1, mp2, mp3, mp4 = st.columns(4)
                        with mp1:
                            ancho_pilar_m = st.number_input("Ancho pilar (m)", value=0.0, key="ancho_pilar_mold")
                        with mp2:
                            largo_pilar_m = st.number_input("Largo pilar (m)", value=0.0, key="largo_pilar_mold")
                        with mp3:
                            alto_pilar_m = st.number_input("Alto pilar (m)", value=0.0, key="alto_pilar_mold")
                        with mp4:
                            cant_pilar_m = st.number_input("Cantidad pilares", value=0, step=1, key="cant_pilar_mold")

                        # m² = perimetro * alto * cantidad
                        perimetro_pilar = (ancho_pilar_m + largo_pilar_m) * 2
                        m2_pilar = perimetro_pilar * alto_pilar_m * cant_pilar_m

                        st.write("---")
                        st.info(f"Superficie de moldaje: {m2_pilar:.2f} m²")

                        mat_p = MATERIALES_MOLDAJE[material_pilar]

                        if "solo_m2" in mat_p:
                            st.text(f"Moldaje metálico requerido: {m2_pilar:.2f} m²")
                        elif "area_plancha" in mat_p:
                            cant_planchas = m2_pilar / mat_p["area_plancha"]
                            cant_planchas_real = cant_planchas * 1.10
                            st.text(f"Planchas terciado 1,22x2,44m: {cant_planchas:.0f} unidades")
                            st.text(f"Con 10% desperdicio: {cant_planchas_real:.0f} planchas")
                        else:
                            ml_tabla = m2_pilar / mat_p["ancho"]
                            cant_tablas = ml_tabla / mat_p["largo"]
                            cant_tablas_real = cant_tablas * 1.10
                            st.text(f"Metros lineales de tabla: {ml_tabla:.2f} ml")
                            st.text(f"Cantidad de tablas de {mat_p['largo']}m: {cant_tablas:.0f} unidades")
                            st.text(f"Con 10% desperdicio: {cant_tablas_real:.0f} tablas")     
                        if m2_pilar > 0:
                            registrar_pdf("Moldajes", "Pilar", [
                                ("Material", material_pilar),
                                ("Superficie", f"{m2_pilar:.2f} m²"),
                                resultado_moldaje("Pilar", material_pilar, m2_pilar),
                            ])

# ============================
# Muros
# ============================   

    if ver_rubro(muros):
        with st.expander("Muros", expanded=False):
# ============================
# MURO DE HORMIGÓN
# ============================
            if ver(muros, "hormigon"):
                    with st.expander("1. Muro de Hormigón", expanded=False):

                        if "secciones_muro_h" not in st.session_state:
                            st.session_state.secciones_muro_h = [{"largo": 0.0, "alto": 0.0, "espesor": 0.0, "cant": 0}]

                        col_add, col_del = st.columns(2)
                        with col_add:
                            if st.button("➕ Agregar sección", key="add_muro_h"):
                                st.session_state.secciones_muro_h.append({"largo": 0.0, "alto": 0.0, "espesor": 0.0, "cant": 0})
                        with col_del:
                            if st.button("🗑️ Eliminar última sección", key="del_muro_h"):
                                if len(st.session_state.secciones_muro_h) > 1:
                                    st.session_state.secciones_muro_h.pop()

                        vol_bruto_h = 0.0
                        largo_muro_h = 0.0
                        alto_muro_h = 0.0
                        espesor_muro_h = 0.0
                        cant_muros_h = 0

                        for i, sec in enumerate(st.session_state.secciones_muro_h):
                            st.markdown(f"**Sección {i+1}**")
                            mh1, mh2, mh3, mh4 = st.columns(4)
                            with mh1:
                                sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"muro_h_largo_{i}")
                            with mh2:
                                sec["alto"] = st.number_input("Alto (m)", value=sec["alto"], key=f"muro_h_alto_{i}")
                            with mh3:
                                sec["espesor"] = st.number_input("Espesor (m)", value=sec["espesor"], key=f"muro_h_esp_{i}")
                            with mh4:
                                sec["cant"] = st.number_input("Cantidad", value=sec["cant"], step=1, key=f"muro_h_cant_{i}")

                            vol_sec = sec["largo"] * sec["alto"] * sec["espesor"] * sec["cant"]
                            st.caption(f"Volumen sección {i+1}: {vol_sec:.2f} m³")
                            vol_bruto_h += vol_sec

                            # Usamos última sección para cálculos de enfierradura
                            largo_muro_h = sec["largo"]
                            alto_muro_h = sec["alto"]
                            espesor_muro_h = sec["espesor"]
                            cant_muros_h = sec["cant"]

                        st.write("---")
                        st.info(f"Volumen bruto total muros hormigón: {vol_bruto_h:.2f} m³")

                        # Vanos
                        st.subheader("🚪 Vanos")
                        vh1, vh2 = st.columns(2)
                        with vh1:
                            cant_puertas_h = st.number_input("Cantidad puertas", value=0, step=1, key="cant_puertas_h")
                            ancho_puerta_h = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_h")
                            alto_puerta_h = st.number_input("Alto puerta (m)", value=0.0, key="alto_puerta_h")
                        with vh2:
                            cant_ventanas_h = st.number_input("Cantidad ventanas", value=0, step=1, key="cant_ventanas_h")
                            ancho_ventana_h = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_h")
                            alto_ventana_h = st.number_input("Alto ventana (m)", value=0.0, key="alto_ventana_h")

                        vol_vanos_h = ((cant_puertas_h * ancho_puerta_h * alto_puerta_h) +
                                    (cant_ventanas_h * ancho_ventana_h * alto_ventana_h)) * espesor_muro_h
                        vol_neto_h = vol_bruto_h - vol_vanos_h

                        dos_muro_h = st.selectbox("Dosificación hormigón", list(DOSIFICACIONES.keys()),
                                                index=2, key="dos_muro_h",
                                                help=DOSIFICACIONES["G-25"]["descripcion"])
                        mat_muro_h = calcular_materiales(vol_neto_h, dos_muro_h)

                        st.write("---")
                        st.subheader("📊 Modo de cálculo enfierradura")
                        modo_muro_h = st.radio(
                            "Selecciona el modo",
                            ["🔨 Modo Simple (estimación)", "📐 Modo Detallado (con planos)"],
                            horizontal=True, key="modo_muro_h"
                        )

                        if modo_muro_h == "🔨 Modo Simple (estimación)":
                            st.caption("Valores típicos según NCh430 para muros estructurales en Chile")
                            if espesor_muro_h >= 0.20:
                                st.warning("⚠️ Espesor ≥ 20cm: NCh430 exige doble malla obligatoriamente")
                                tipo_malla = "Doble"
                            else:
                                st.info("ℹ️ Espesor < 20cm: Se puede usar malla simple o doble según cálculo")
                                tipo_malla = st.selectbox("Tipo de malla", ["Simple", "Doble"], key="malla_simple_h")
                            ms1, ms2 = st.columns(2)
                            with ms1:
                                diam_vert_s = st.selectbox("Diámetro barra vertical", ["Ø8mm", "Ø10mm", "Ø12mm"], index=1, key="diam_vert_s")
                                sep_vert_s = st.selectbox("Separación vertical (m)", ["0.15", "0.20", "0.25"], index=1, key="sep_vert_s")
                            with ms2:
                                diam_horiz_s = st.selectbox("Diámetro barra horizontal", ["Ø8mm", "Ø10mm", "Ø12mm"], index=1, key="diam_horiz_s")
                                sep_horiz_s = st.selectbox("Separación horizontal (m)", ["0.15", "0.20", "0.25"], index=1, key="sep_horiz_s")
                            sep_v_s = float(sep_vert_s)
                            sep_h_s = float(sep_horiz_s)
                            n_mallas = 2 if tipo_malla == "Doble" else 1
                            cant_barras_v_s = (largo_muro_h / sep_v_s + 1) * cant_muros_h * n_mallas
                            ml_v_s = cant_barras_v_s * alto_muro_h
                            kg_v_s = ml_v_s * PESO_BARRAS[diam_vert_s]
                            cant_barras_h_s = (alto_muro_h / sep_h_s + 1) * cant_muros_h * n_mallas
                            ml_h_s = cant_barras_h_s * largo_muro_h
                            kg_h_s = ml_h_s * PESO_BARRAS[diam_horiz_s]
                            st.info("💡 NCh430: Se recomiendan barras de borde en extremos del muro")
                            diam_borde = st.selectbox("Diámetro barra de borde", ["Ø12mm", "Ø16mm", "Ø20mm"], index=0, key="diam_borde_s")
                            ml_borde = alto_muro_h * 4 * cant_muros_h
                            kg_borde = ml_borde * PESO_BARRAS[diam_borde]
                            kg_diag = 0
                            if cant_puertas_h > 0 or cant_ventanas_h > 0:
                                st.info("💡 NCh430: Se requieren barras diagonales Ø12mm en esquinas de vanos")
                                ml_diag = (cant_puertas_h + cant_ventanas_h) * 4 * 0.60 * cant_muros_h
                                kg_diag = ml_diag * PESO_BARRAS["Ø12mm"]
                            kg_total_h = kg_v_s + kg_h_s + kg_borde + kg_diag
                            st.write("---")
                            st.subheader("📦 Resultados Hormigón")
                            # mostrar_materiales se auto-limpia si el volumen es 0
                            mostrar_materiales(mat_muro_h, "Muros", "Muro de hormigón")
                            st.subheader("📦 Resultados Enfierradura")
                            re1, re2 = st.columns(2)
                            with re1:
                                st.info(f"Volumen neto muro: {vol_neto_h:.2f} m³")
                                st.info(f"Malla: {tipo_malla} ({n_mallas} capa/s)")
                                st.info(f"Barras verticales {diam_vert_s}: {cant_barras_v_s:.0f} barras → {kg_v_s:.1f} kg")
                                st.info(f"Barras horizontales {diam_horiz_s}: {cant_barras_h_s:.0f} barras → {kg_h_s:.1f} kg")
                            with re2:
                                st.info(f"Barras de borde {diam_borde}: {ml_borde:.1f} ml → {kg_borde:.1f} kg")
                                if kg_diag > 0:
                                    st.info(f"Refuerzo diagonal vanos: {kg_diag:.1f} kg")
                                st.success(f"TOTAL ACERO: {kg_total_h:.1f} kg")
                            if vol_neto_h > 0:
                                registrar_pdf("Muros", "Muro Hormigón", [
                                    ("Volumen neto", f"{vol_neto_h:.2f} m³"),
                                    ("Dosificación", dos_muro_h),
                                    ("Cemento", f"{mat_muro_h['cemento_sacos']} sacos"),
                                    ("Gravilla", f"{mat_muro_h['gravilla_kg']} kg"),
                                    ("Arena", f"{mat_muro_h['arena_kg']} kg"),
                                    ("Acero total", f"{kg_total_h:.1f} kg"),
                                ])
                            else:
                                quitar_pdf("Muros", "Muro Hormigón")

                        elif modo_muro_h == "📐 Modo Detallado (con planos)":
                            st.caption("Ingresa los datos exactos según planos estructurales")
                            if espesor_muro_h >= 0.20:
                                st.warning("⚠️ Espesor ≥ 20cm: NCh430 exige doble malla obligatoriamente")
                                tipo_malla_d = "Doble"
                            else:
                                tipo_malla_d = st.selectbox("Tipo de malla", ["Simple", "Doble"], key="malla_det_h")
                            n_mallas_d = 2 if tipo_malla_d == "Doble" else 1
                            st.write("**Armadura vertical**")
                            dv1, dv2 = st.columns(2)
                            with dv1:
                                diam_vert_d = st.selectbox("Diámetro", list(PESO_BARRAS.keys()), index=1, key="diam_vert_d")
                            with dv2:
                                sep_vert_d = st.selectbox("Separación (m)", ["0.10", "0.15", "0.20", "0.25", "0.30"], index=2, key="sep_vert_d")
                            st.write("**Armadura horizontal**")
                            dh1, dh2 = st.columns(2)
                            with dh1:
                                diam_horiz_d = st.selectbox("Diámetro", list(PESO_BARRAS.keys()), index=1, key="diam_horiz_d")
                            with dh2:
                                sep_horiz_d = st.selectbox("Separación (m)", ["0.10", "0.15", "0.20", "0.25", "0.30"], index=2, key="sep_horiz_d")
                            st.write("**Barras de borde**")
                            bb1, bb2, bb3 = st.columns(3)
                            with bb1:
                                cant_barras_borde_d = st.number_input("Cantidad barras borde", value=4, step=1, key="cant_bb_d")
                            with bb2:
                                diam_borde_d = st.selectbox("Diámetro borde", list(PESO_BARRAS.keys()), index=3, key="diam_borde_d")
                            with bb3:
                                sep_estribos_d = st.selectbox("Sep. estribos borde (m)", ["0.08", "0.10", "0.15"], key="sep_est_d")
                            st.write("**Refuerzo diagonal vanos**")
                            rd1, rd2 = st.columns(2)
                            with rd1:
                                diam_diag_d = st.selectbox("Diámetro diagonal", list(PESO_BARRAS.keys()), index=2, key="diam_diag_d")
                            with rd2:
                                largo_diag_d = st.number_input("Largo diagonal (m)", value=0.60, key="largo_diag_d")
                            sep_v_d = float(sep_vert_d)
                            sep_h_d = float(sep_horiz_d)
                            cant_v_d = (largo_muro_h / sep_v_d + 1) * cant_muros_h * n_mallas_d
                            ml_v_d = cant_v_d * alto_muro_h
                            kg_v_d = ml_v_d * PESO_BARRAS[diam_vert_d]
                            cant_h_d = (alto_muro_h / sep_h_d + 1) * cant_muros_h * n_mallas_d
                            ml_h_d = cant_h_d * largo_muro_h
                            kg_h_d = ml_h_d * PESO_BARRAS[diam_horiz_d]
                            ml_borde_d = alto_muro_h * cant_barras_borde_d * cant_muros_h
                            kg_borde_d = ml_borde_d * PESO_BARRAS[diam_borde_d]
                            n_estribos_borde = (alto_muro_h / float(sep_estribos_d)) * cant_muros_h
                            kg_estribos_borde = n_estribos_borde * (espesor_muro_h * 2 + 0.20) * PESO_BARRAS["Ø8mm"]
                            ml_diag_d = (cant_puertas_h + cant_ventanas_h) * 4 * largo_diag_d * cant_muros_h
                            kg_diag_d = ml_diag_d * PESO_BARRAS[diam_diag_d]
                            kg_total_d = kg_v_d + kg_h_d + kg_borde_d + kg_estribos_borde + kg_diag_d
                            st.write("---")
                            st.subheader("📦 Resultados Hormigón")
                            # mostrar_materiales se auto-limpia si el volumen es 0
                            mostrar_materiales(mat_muro_h, "Muros", "Muro de hormigón")
                            st.subheader("📦 Resultados Enfierradura")
                            rd1, rd2 = st.columns(2)
                            with rd1:
                                st.info(f"Volumen neto: {vol_neto_h:.2f} m³")
                                st.info(f"Malla: {tipo_malla_d} ({n_mallas_d} capa/s)")
                                st.info(f"Barras vert. {diam_vert_d}: {cant_v_d:.0f} barras → {kg_v_d:.1f} kg")
                                st.info(f"Barras horiz. {diam_horiz_d}: {cant_h_d:.0f} barras → {kg_h_d:.1f} kg")
                            with rd2:
                                st.info(f"Barras borde {diam_borde_d}: {ml_borde_d:.1f} ml → {kg_borde_d:.1f} kg")
                                st.info(f"Estribos borde Ø8mm: {n_estribos_borde:.0f} unid → {kg_estribos_borde:.1f} kg")
                                if kg_diag_d > 0:
                                    st.info(f"Diagonal vanos {diam_diag_d}: {kg_diag_d:.1f} kg")
                                st.success(f"TOTAL ACERO: {kg_total_d:.1f} kg")
                            if vol_neto_h > 0:
                                registrar_pdf("Muros", "Muro Hormigón", [
                                    ("Volumen neto", f"{vol_neto_h:.2f} m³"),
                                    ("Dosificación", dos_muro_h),
                                    ("Cemento", f"{mat_muro_h['cemento_sacos']} sacos"),
                                    ("Gravilla", f"{mat_muro_h['gravilla_kg']} kg"),
                                    ("Arena", f"{mat_muro_h['arena_kg']} kg"),
                                    ("Acero total", f"{kg_total_d:.1f} kg"),
                                ])
                            else:
                                quitar_pdf("Muros", "Muro Hormigón")
# ============================
# MURO DE LADRILLO
# ============================
            if ver(muros, "ladrillo"):
                    with st.expander("2. Muro de Ladrillo", expanded=False):

                        LADRILLOS = {
                            "Fiscal (29x14x5cm)": {"largo": 0.29, "ancho": 0.14, "alto": 0.05, "junta": 0.013, "descripcion": "Muros estructurales y de carga"},
                            "Princesa (29x14x7,1cm)": {"largo": 0.29, "ancho": 0.14, "alto": 0.071, "junta": 0.010, "descripcion": "Tabiques y muros no soportantes"},
                            "Mechón/Hueco Titán (29x14x9,4cm)": {"largo": 0.29, "ancho": 0.14, "alto": 0.094, "junta": 0.010, "descripcion": "Tabiquería interior y muros divisorios"},
                            "Caravista (29x14x7,1cm)": {"largo": 0.29, "ancho": 0.14, "alto": 0.071, "junta": 0.010, "descripcion": "Terminación estética, exteriores sin estuco"},
                        }

                        ladrillo_tipo = st.selectbox("Tipo de ladrillo", list(LADRILLOS.keys()), key="ladrillo_tipo")
                        st.caption(LADRILLOS[ladrillo_tipo]["descripcion"])

                        if "secciones_muro_l" not in st.session_state:
                            st.session_state.secciones_muro_l = [{"largo": 0.0, "alto": 0.0, "cant": 0}]

                        col_add, col_del = st.columns(2)
                        with col_add:
                            if st.button("➕ Agregar sección", key="add_muro_l"):
                                st.session_state.secciones_muro_l.append({"largo": 0.0, "alto": 0.0, "cant": 0})
                        with col_del:
                            if st.button("🗑️ Eliminar última sección", key="del_muro_l"):
                                if len(st.session_state.secciones_muro_l) > 1:
                                    st.session_state.secciones_muro_l.pop()

                        area_bruta_l = 0.0
                        for i, sec in enumerate(st.session_state.secciones_muro_l):
                            st.markdown(f"**Sección {i+1}**")
                            lb1, lb2, lb3 = st.columns(3)
                            with lb1:
                                sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"muro_l_largo_{i}")
                            with lb2:
                                sec["alto"] = st.number_input("Alto (m)", value=sec["alto"], key=f"muro_l_alto_{i}")
                            with lb3:
                                sec["cant"] = st.number_input("Cantidad", value=sec["cant"], step=1, key=f"muro_l_cant_{i}")
                            area_sec = sec["largo"] * sec["alto"] * sec["cant"]
                            st.caption(f"Área sección {i+1}: {area_sec:.2f} m²")
                            area_bruta_l += area_sec

                        st.write("---")
                        st.info(f"Área bruta total: {area_bruta_l:.2f} m²")

                        st.subheader("🚪 Vanos")
                        vl1, vl2 = st.columns(2)
                        with vl1:
                            cant_puertas_l = st.number_input("Cantidad puertas", value=0, step=1, key="cant_puertas_l")
                            ancho_puerta_l = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_l")
                            alto_puerta_l = st.number_input("Alto puerta (m)", value=0.0, key="alto_puerta_l")
                        with vl2:
                            cant_ventanas_l = st.number_input("Cantidad ventanas", value=0, step=1, key="cant_ventanas_l")
                            ancho_ventana_l = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_l")
                            alto_ventana_l = st.number_input("Alto ventana (m)", value=0.0, key="alto_ventana_l")

                        desp_ladrillo = st.slider("% Desperdicio ladrillos", 0, 15, 5, key="desp_ladrillo")

                        lad = LADRILLOS[ladrillo_tipo]
                        area_vanos_l = ((cant_puertas_l * ancho_puerta_l * alto_puerta_l) +
                                    (cant_ventanas_l * ancho_ventana_l * alto_ventana_l))
                        area_neta_l = area_bruta_l - area_vanos_l

                        largo_con_junta = lad["largo"] + lad["junta"]
                        alto_con_junta = lad["alto"] + lad["junta"]
                        ladrillos_por_m2 = 1 / (largo_con_junta * alto_con_junta)
                        total_ladrillos = area_neta_l * ladrillos_por_m2
                        total_ladrillos_desp = total_ladrillos * (1 + desp_ladrillo / 100)

                        vol_mortero = area_neta_l * lad["ancho"] * 0.20
                        cemento_mortero_kg = vol_mortero * 400
                        cemento_mortero_sacos = cemento_mortero_kg / 25
                        arena_mortero_m3 = vol_mortero * 1.10

                        st.write("---")
                        st.subheader("📦 Resultados")
                        rl1, rl2 = st.columns(2)
                        with rl1:
                            st.info(f"Área neta muro: {area_neta_l:.2f} m²")
                            st.info(f"Ladrillos por m²: {ladrillos_por_m2:.1f} unidades")
                            st.info(f"Total ladrillos: {total_ladrillos:.0f} unidades")
                            st.success(f"Con {desp_ladrillo}% desperdicio: {total_ladrillos_desp:.0f} ladrillos")
                        with rl2:
                            st.info(f"Volumen mortero: {vol_mortero:.3f} m³")
                            st.info(f"Cemento mortero: {cemento_mortero_sacos:.0f} sacos de 25kg")
                            st.info(f"Arena mortero: {arena_mortero_m3:.2f} m³")

                        if "Caravista" in ladrillo_tipo:
                            st.warning("⚠️ Ladrillo Caravista: Requiere mortero especial de color y junta vista.")

                        st.write("---")
                        st.subheader("💡 Datos técnicos")
                        st.text(f"Dimensiones ladrillo: {lad['largo']*100:.0f}x{lad['ancho']*100:.0f}x{lad['alto']*100:.1f}cm")
                        st.text(f"Junta de mortero: {lad['junta']*10:.0f}mm")
                        st.text(f"Dosificación mortero: 1:4 (cemento:arena)")
                        if area_neta_l > 0:
                            registrar_pdf("Muros", "Muro Ladrillo", [
                                ("Tipo", ladrillo_tipo),
                                ("Área neta", f"{area_neta_l:.2f} m²"),
                                ("Ladrillos (c/desp.)", f"{total_ladrillos_desp:.0f} unidades"),
                                ("Cemento mortero", f"{cemento_mortero_sacos:.0f} sacos"),
                                ("Arena mortero", f"{arena_mortero_m3:.2f} m³"),
                            ])
                        # ============================
# TABIQUE METALCON
# ============================
            if ver(muros, "tabique_metalcon"):
                with st.expander("3. Tabique Metalcon", expanded=False):

                    MONTANTES_MU = {
                        "Montante Normal Perf. 60x38x0,5 - 2,40m": {"largo": 2.40, "peso": 0.56},
                        "Montante Normal Perf. 60x38x0,5 - 3,00m": {"largo": 3.00, "peso": 0.56},
                        "Montante Económico Perf. 38x38x0,5 - 2,40m": {"largo": 2.40, "peso": 0.48},
                        "Montante Económico Perf. 38x38x0,5 - 3,00m": {"largo": 3.00, "peso": 0.48},
                        "Montante Normal 60x38x0,5 - 2,40m": {"largo": 2.40, "peso": 0.56},
                        "Montante Normal 60x38x0,5 - 3,00m": {"largo": 3.00, "peso": 0.56},
                        "Montante Económico 38x38x0,5 - 2,40m": {"largo": 2.40, "peso": 0.48},
                        "Montante Económico 38x38x0,5 - 3,00m": {"largo": 3.00, "peso": 0.48},
                    }
                    CANALES_MU = {
                        "Canal Normal 61x20x0,5 - 3,00m": {"largo": 3.00, "peso": 0.39},
                        "Canal Económico 39x20x0,5 - 3,00m": {"largo": 3.00, "peso": 0.31},
                    }

                    if "secciones_tab_mu" not in st.session_state:
                        st.session_state.secciones_tab_mu = [{"largo": 0.0, "alto": 0.0, "cant": 0}]

                    col_add, col_del = st.columns(2)
                    with col_add:
                        if st.button("➕ Agregar sección", key="add_tab_mu"):
                            st.session_state.secciones_tab_mu.append({"largo": 0.0, "alto": 0.0, "cant": 0})
                    with col_del:
                        if st.button("🗑️ Eliminar última sección", key="del_tab_mu"):
                            if len(st.session_state.secciones_tab_mu) > 1:
                                st.session_state.secciones_tab_mu.pop()

                    largo_tab_mu = 0.0
                    alto_tab_mu = 0.0
                    cant_tab_mu = 0

                    for i, sec in enumerate(st.session_state.secciones_tab_mu):
                        st.markdown(f"**Sección {i+1}**")
                        tm1, tm2, tm3 = st.columns(3)
                        with tm1:
                            sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"tab_mu_largo_{i}")
                        with tm2:
                            sec["alto"] = st.number_input("Alto (m)", value=sec["alto"], key=f"tab_mu_alto_{i}")
                        with tm3:
                            sec["cant"] = st.number_input("Cantidad", value=sec["cant"], step=1, key=f"tab_mu_cant_{i}")
                        st.caption(f"Área sección {i+1}: {sec['largo'] * sec['alto'] * sec['cant']:.2f} m²")
                        largo_tab_mu += sec["largo"] * sec["cant"]
                        alto_tab_mu = max(alto_tab_mu, sec["alto"])
                        cant_tab_mu += sec["cant"]

                    st.write("---")

                    # Resto del código igual que tenías desde "sep_mu = st.selectbox..."
                    sep_mu = st.selectbox("Separación montantes",
                                        ["0,40m (recomendado)", "0,60m (máximo)"], key="sep_mu")
                    sep_valor_mu = 0.40 if "0,40" in sep_mu else 0.60
                    canal_tipo_mu = st.selectbox("Tipo de canal", list(CANALES_MU.keys()), key="canal_mu")
                    st.write("**Tipo de montantes**")
                    mt1, mt2 = st.columns(2)
                    with mt1:
                        montante_medio = st.selectbox("Montante central (perforado)",
                            [k for k in MONTANTES_MU.keys() if "Perf" in k], key="mont_medio_mu")
                        st.caption("✅ Va en el medio - permite paso de instalaciones")
                    with mt2:
                        montante_esquina = st.selectbox("Montante esquinas y encuentros (normal)",
                            [k for k in MONTANTES_MU.keys() if "Perf" not in k], key="mont_esq_mu")
                        st.caption("⚠️ Va en esquinas y encuentros")

                    st.subheader("🚪 Vanos")
                    v1, v2 = st.columns(2)
                    with v1:
                        cant_puertas_mu = st.number_input("Cantidad de puertas", value=0, step=1, key="cant_puertas_mu")
                        ancho_puerta_mu = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_mu")
                    with v2:
                        cant_ventanas_mu = st.number_input("Cantidad de ventanas", value=0, step=1, key="cant_ventanas_mu")
                        ancho_ventana_mu = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_mu")

                    st.subheader("🔲 Esquinas y Encuentros")
                    ec1, ec2 = st.columns(2)
                    with ec1:
                        cant_esquinas_mu = st.number_input("Cantidad de esquinas", value=0, step=1, key="cant_esq_mu")
                    with ec2:
                        cant_encuentros_mu = st.number_input("Cantidad de encuentros", value=0, step=1, key="cant_enc_mu")

                    incluir_diagonales_mu = st.checkbox(
                        "Incluir diagonales (riostras)", value=True, key="incluir_diag_mu",
                        help="Las diagonales rigidizan el tabique. Inclúyelas si tu diseño las requiere.",
                    )

                    # Cálculos — igual que tenías
                    largo_canal_mu = CANALES_MU[canal_tipo_mu]["largo"]
                    largo_mont_medio = MONTANTES_MU[montante_medio]["largo"]
                    largo_mont_esq = MONTANTES_MU[montante_esquina]["largo"]
                    ml_canal_mu = largo_tab_mu * 2
                    cant_canales_mu = ml_canal_mu / largo_canal_mu if largo_canal_mu > 0 else 0
                    montantes_medio_por_tab = max(0, int((largo_tab_mu / cant_tab_mu if cant_tab_mu > 0 else 0) / sep_valor_mu) - 1)
                    total_mont_medio = montantes_medio_por_tab * cant_tab_mu
                    mont_extremos = 2 * cant_tab_mu
                    mont_esquinas = cant_esquinas_mu * 3
                    mont_encuentros = cant_encuentros_mu * 4
                    mont_vanos = (cant_puertas_mu + cant_ventanas_mu) * 2
                    total_mont_normal = mont_extremos + mont_esquinas + mont_encuentros + mont_vanos
                    canales_extra_vanos = (cant_puertas_mu + cant_ventanas_mu) * 2
                    total_canales_final = cant_canales_mu + canales_extra_vanos
                    total_diagonales = (cant_tab_mu * 2) if incluir_diagonales_mu else 0
                    total_pletinas = cant_esquinas_mu
                    tornillos = ((total_mont_medio + total_mont_normal) * 4) + (total_canales_final * 2)
                    m2_aislacion = largo_tab_mu * alto_tab_mu
                    m2_vanos_mu = ((cant_puertas_mu * ancho_puerta_mu * alto_tab_mu) +
                                (cant_ventanas_mu * ancho_ventana_mu * alto_tab_mu))
                    m2_aislacion_neta = m2_aislacion - m2_vanos_mu

                    st.write("---")
                    st.subheader("📦 Resultados")
                    r1, r2 = st.columns(2)
                    with r1:
                        st.info(f"Canales totales: {total_canales_final:.0f} piezas de {largo_canal_mu}m")
                        st.info(f"Montantes perforados (medio): {total_mont_medio:.0f} piezas de {largo_mont_medio}m")
                        st.info(f"Montantes normales (esquinas/extremos): {total_mont_normal:.0f} piezas de {largo_mont_esq}m")
                        st.info(f"Total montantes: {total_mont_medio + total_mont_normal:.0f} piezas")
                    with r2:
                        if incluir_diagonales_mu:
                            st.info(f"Diagonales: {total_diagonales} piezas")
                        st.info(f"Pletinas esquinas: {total_pletinas} unidades")
                        st.info(f"Tornillos autoperf.: {tornillos:.0f} unidades")
                        st.info(f"Lana de vidrio: {m2_aislacion_neta:.2f} m²")

                    st.session_state["pdf_canal_tipo"] = canal_tipo_mu
                    st.session_state["pdf_cant_piezas_canal"] = total_canales_final
                    st.session_state["pdf_ml_canal"] = ml_canal_mu
                    st.session_state["pdf_largo_canal"] = largo_canal_mu
                    st.session_state["pdf_montante_tipo"] = montante_medio
                    st.session_state["pdf_total_montantes"] = total_mont_medio + total_mont_normal
                    st.session_state["pdf_largo_montante"] = largo_mont_medio
                    st.session_state["pdf_esq_tipo"] = ""
                    st.session_state["pdf_cant_esquinas"] = cant_esquinas_mu
                    st.session_state["pdf_largo_esq"] = 0
                    if largo_tab_mu > 0 and total_canales_final > 0:
                        items_metalcon = [
                            ("Canales", f"{total_canales_final:.0f} de {largo_canal_mu}m"),
                            ("Montantes perf. (medio)", f"{total_mont_medio:.0f} de {largo_mont_medio}m"),
                            ("Montantes normales", f"{total_mont_normal:.0f} de {largo_mont_esq}m"),
                        ]
                        if incluir_diagonales_mu:
                            items_metalcon.append(("Diagonales", f"{total_diagonales} piezas"))
                        items_metalcon.extend([
                            ("Tornillos autoperf.", f"{tornillos:.0f} unidades"),
                            ("Lana de vidrio", f"{m2_aislacion_neta:.2f} m²"),
                        ])
                        registrar_pdf("Muros", "Tabique Metalcon", items_metalcon)
# ============================
# TABIQUE DE MADERA
# ============================
            if ver(muros, "tabique_madera"):
                    with st.expander("4. Tabique de Madera", expanded=False):

                        MADERA_TIPOS = {
                            "2x2 Pino Dimensionado Verde 3,2m": {"largo": 3.20, "ancho": 0.038, "alto": 0.038},
                            "2x3 Pino Dimensionado Verde 3,2m": {"largo": 3.20, "ancho": 0.038, "alto": 0.063},
                            "1x2 Pino Certificado Seco Cepillado 3,2m": {"largo": 3.20, "ancho": 0.025, "alto": 0.038},
                        }

                        if "secciones_tab_ma" not in st.session_state:
                            st.session_state.secciones_tab_ma = [{"largo": 0.0, "alto": 0.0, "cant": 0}]

                        col_add, col_del = st.columns(2)
                        with col_add:
                            if st.button("➕ Agregar sección", key="add_tab_ma"):
                                st.session_state.secciones_tab_ma.append({"largo": 0.0, "alto": 0.0, "cant": 0})
                        with col_del:
                            if st.button("🗑️ Eliminar última sección", key="del_tab_ma"):
                                if len(st.session_state.secciones_tab_ma) > 1:
                                    st.session_state.secciones_tab_ma.pop()

                        largo_tab_ma = 0.0
                        alto_tab_ma = 0.0
                        cant_tab_ma = 0

                        for i, sec in enumerate(st.session_state.secciones_tab_ma):
                            st.markdown(f"**Sección {i+1}**")
                            tw1, tw2, tw3 = st.columns(3)
                            with tw1:
                                sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"tab_ma_largo_{i}")
                            with tw2:
                                sec["alto"] = st.number_input("Alto (m)", value=sec["alto"], key=f"tab_ma_alto_{i}")
                            with tw3:
                                sec["cant"] = st.number_input("Cantidad", value=sec["cant"], step=1, key=f"tab_ma_cant_{i}")
                            st.caption(f"Área sección {i+1}: {sec['largo'] * sec['alto'] * sec['cant']:.2f} m²")
                            largo_tab_ma += sec["largo"] * sec["cant"]
                            alto_tab_ma = max(alto_tab_ma, sec["alto"])
                            cant_tab_ma += sec["cant"]

                        st.write("---")

                        madera_tipo = st.selectbox("Tipo de madera", list(MADERA_TIPOS.keys()), key="madera_tipo")
                        sep_ma = st.selectbox("Separación entre montantes", ["0,40m", "0,60m"], key="sep_ma")
                        sep_valor_ma = 0.40 if "0,40" in sep_ma else 0.60
                        cant_cad = st.selectbox("Filas de cadenetas", ["1 fila (cada 80cm)", "2 filas (cada 60cm)", "3 filas (cada 50cm)"], key="cant_cad")
                        n_cadenetas = int(cant_cad[0])

                        st.subheader("🚪 Vanos")
                        vw1, vw2 = st.columns(2)
                        with vw1:
                            cant_puertas_ma = st.number_input("Cantidad de puertas", value=0, step=1, key="cant_puertas_ma")
                            ancho_puerta_ma = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_ma")
                        with vw2:
                            cant_ventanas_ma = st.number_input("Cantidad de ventanas", value=0, step=1, key="cant_ventanas_ma")
                            ancho_ventana_ma = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_ma")

                        cant_esq_ma = st.number_input("Cantidad de esquinas", value=0, step=1, key="cant_esq_ma")

                        largo_mad = MADERA_TIPOS[madera_tipo]["largo"]
                        ml_solera_ma = largo_tab_ma * 2
                        cant_soleras_ma = ml_solera_ma / largo_mad if largo_mad > 0 else 0
                        largo_por_tab = largo_tab_ma / cant_tab_ma if cant_tab_ma > 0 else 0
                        montantes_por_tab_ma = int(largo_por_tab / sep_valor_ma) + 1
                        total_mont_ma = montantes_por_tab_ma * cant_tab_ma
                        total_diag_ma = cant_tab_ma * 2
                        ml_cadenetas = largo_tab_ma * n_cadenetas
                        cant_cadenetas = ml_cadenetas / largo_mad if largo_mad > 0 else 0
                        piezas_puertas = cant_puertas_ma * 1
                        piezas_ventanas = cant_ventanas_ma * 2
                        piezas_esquinas = cant_esq_ma
                        total_listones = cant_soleras_ma + total_mont_ma + total_diag_ma + cant_cadenetas + piezas_puertas + piezas_ventanas + piezas_esquinas
                        clavos = total_listones * 4
                        m2_papel = largo_tab_ma * alto_tab_ma
                        m2_vanos_ma = (cant_puertas_ma * ancho_puerta_ma * alto_tab_ma) + (cant_ventanas_ma * ancho_ventana_ma * alto_tab_ma)
                        m2_papel_neto = m2_papel - m2_vanos_ma
                        m2_aislante_ma = m2_papel_neto

                        st.write("---")
                        st.subheader("📦 Resultados")
                        rm1, rm2 = st.columns(2)
                        with rm1:
                            st.info(f"Soleras y carreras: {cant_soleras_ma:.0f} piezas de {largo_mad}m")
                            st.info(f"Montantes: {total_mont_ma:.0f} piezas de {largo_mad}m")
                            st.info(f"Diagonales: {total_diag_ma} piezas de {largo_mad}m")
                            st.info(f"Cadenetas: {cant_cadenetas:.0f} piezas de {largo_mad}m")
                        with rm2:
                            st.info(f"Vanos puertas: {piezas_puertas} piezas")
                            st.info(f"Vanos ventanas: {piezas_ventanas} piezas")
                            st.info(f"Esquinas: {piezas_esquinas} piezas")
                            st.info(f"Total listones: {total_listones:.0f} piezas")
                        st.write("---")
                        rm3, rm4 = st.columns(2)
                        with rm3:
                            st.info(f"Papel kraft: {m2_papel_neto:.2f} m²")
                            st.info(f"Aislante: {m2_aislante_ma:.2f} m²")
                        with rm4:
                            st.info(f"Clavos aprox.: {clavos:.0f} unidades")
                        if largo_tab_ma > 0:
                            registrar_pdf("Muros", "Tabique Madera", [
                                ("Tipo madera", madera_tipo),
                                ("Soleras y carreras", f"{cant_soleras_ma:.0f} de {largo_mad}m"),
                                ("Montantes", f"{total_mont_ma:.0f} de {largo_mad}m"),
                                ("Diagonales", f"{total_diag_ma} de {largo_mad}m"),
                                ("Cadenetas", f"{cant_cadenetas:.0f} de {largo_mad}m"),
                                ("Total listones", f"{total_listones:.0f} piezas"),
                                ("Aislante", f"{m2_aislante_ma:.2f} m²"),
                                ("Clavos", f"{clavos:.0f} unidades"),
                            ])

# ============================
# REVESTIMIENTOS
# ============================
    if ver_rubro(revest):
            with st.expander("Revestimientos", expanded=False):

                YESO_CARTON = {
                    "ST Estándar 10mm (Zonas secas)":        {"espesor": 10,   "area": 2.88, "tipo": "ST"},
                    "ST Estándar 12,5mm (Zonas secas)":      {"espesor": 12.5, "area": 2.88, "tipo": "ST"},
                    "RH Humedad 12,5mm (Baños/Cocinas)":     {"espesor": 12.5, "area": 2.88, "tipo": "RH"},
                    "RH Humedad 15mm (Baños/Cocinas)":       {"espesor": 15,   "area": 2.88, "tipo": "RH"},
                    "RF Fuego 12,5mm (Shaft/Bodegas)":       {"espesor": 12.5, "area": 2.88, "tipo": "RF"},
                    "RF Fuego 15mm (Shaft/Bodegas)":         {"espesor": 15,   "area": 2.88, "tipo": "RF"},
                }
                TERCIADO_RANURADO = {
                    "Terciado Ranurado 5,5-6mm (Colonial liviano)": {"area": 2.98},
                    "Terciado Ranurado 9mm (Uso general)":          {"area": 2.98},
                    "Terciado Ranurado 12mm (Alta resistencia)":    {"area": 2.98},
                }
                OSB_TIPOS = {
                    "OSB 9,5mm":  {"area": 2.98},
                    "OSB 11,1mm": {"area": 2.98},
                    "OSB 15,1mm": {"area": 2.98},
                }
                FIBROCEMENTO = {
                    "Lisa 4mm (Sobre tablero madera)":      {"area": 2.88},
                    "Lisa 5mm (Cielos y aleros húmedos)":   {"area": 2.88},
                    "Simplísima 6mm (Fachada)":             {"area": 2.88},
                    "Textura Madera 6mm (Fachada)":         {"area": 2.88},
                }
                TERCIADO_ESTRUCTURAL = {
                    "Terciado Estructural 9mm":  {"area": 2.98},
                    "Terciado Estructural 12mm": {"area": 2.98},
                    "Terciado Estructural 15mm": {"area": 2.98},
                    "Terciado Estructural 18mm": {"area": 2.98},
                }

                def calcular_planchas(area_neta, area_plancha, desperdicio):
                    cant_exacta = area_neta / area_plancha if area_plancha > 0 else 0
                    cant_con_desp = cant_exacta * (1 + desperdicio / 100)
                    sobra_ultima = (cant_exacta % 1) * area_plancha
                    desperdicio_m2 = (cant_con_desp - cant_exacta) * area_plancha
                    return cant_exacta, cant_con_desp, sobra_ultima, desperdicio_m2

                def calcular_tornillos_plancha(cant_planchas, area_plancha, sep_borde, sep_centro):
                    perimetro = (1.22 + 2.44) * 2
                    apoyos_internos = area_plancha / (sep_centro * 1.22)
                    torn_borde = (perimetro / sep_borde) * cant_planchas
                    torn_centro = apoyos_internos * cant_planchas
                    return round(torn_borde + torn_centro)

                def seccion_muro(key_prefix, label="muro"):
                    """Helper para agregar secciones de muros con botones agregar/eliminar"""
                    key_list = f"secciones_{key_prefix}"
                    if key_list not in st.session_state:
                        st.session_state[key_list] = [{"largo": 0.0, "alto": 0.0, "cant": 0}]
                    col_add, col_del = st.columns(2)
                    with col_add:
                        if st.button("➕ Agregar sección", key=f"add_{key_prefix}"):
                            st.session_state[key_list].append({"largo": 0.0, "alto": 0.0, "cant": 0})
                    with col_del:
                        if st.button("🗑️ Eliminar última sección", key=f"del_{key_prefix}"):
                            if len(st.session_state[key_list]) > 1:
                                st.session_state[key_list].pop()
                    area_total = 0.0
                    for i, sec in enumerate(st.session_state[key_list]):
                        st.markdown(f"**Sección {i+1}**")
                        c1, c2, c3 = st.columns(3)
                        with c1:
                            sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"{key_prefix}_largo_{i}")
                        with c2:
                            sec["alto"] = st.number_input("Alto (m)", value=sec["alto"], key=f"{key_prefix}_alto_{i}")
                        with c3:
                            sec["cant"] = st.number_input("Cantidad", value=sec["cant"], step=1, key=f"{key_prefix}_cant_{i}")
                        area_sec = sec["largo"] * sec["alto"] * sec["cant"]
                        st.caption(f"Área sección {i+1}: {area_sec:.2f} m²")
                        area_total += area_sec
                    st.info(f"Área bruta total: {area_total:.2f} m²")
                    return area_total

# ============================
# INTERIOR
# ============================
                if ver(revest, "interior"):
                    with st.expander("1. Revestimiento Interior", expanded=False):

                        with st.expander("1.1 Yeso Cartón", expanded=False):
                            yeso_tipo = st.selectbox("Tipo de Yeso Cartón", list(YESO_CARTON.keys()), key="yeso_tipo")
                            if YESO_CARTON[yeso_tipo]["tipo"] == "ST":
                                st.caption("🏠 Uso: dormitorios, living, zonas secas")
                            elif YESO_CARTON[yeso_tipo]["tipo"] == "RH":
                                st.caption("💧 Uso: baños, cocinas - color verde")
                            else:
                                st.caption("🔥 Uso: shaft, bodegas, vías escape - color rojo/rosado")

                            estructura_yc = st.selectbox("Tipo de estructura", ["Metalcon", "Madera"], key="estructura_yc")
                            tornillo_yc = "Autoperforante Punta Fina (Trompeta)" if estructura_yc == "Metalcon" else "Tornillo Clavex (hilo grueso)"
                            medida_yc = "6x1\" o 6x1 1/4\""

                            caras_yc = st.selectbox("Cantidad de caras", ["1 cara", "2 caras"], key="caras_yc")
                            n_caras_yc = 1 if "1" in caras_yc else 2

                            area_bruta_yc = seccion_muro("yc") * n_caras_yc

                            st.write("**Vanos**")
                            yv1, yv2 = st.columns(2)
                            with yv1:
                                cant_puertas_yc = st.number_input("Cantidad puertas", value=0, step=1, key="cant_puertas_yc")
                                ancho_puerta_yc = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_yc")
                                alto_puerta_yc  = st.number_input("Alto puerta (m)", value=0.0, key="alto_puerta_yc")
                            with yv2:
                                cant_ventanas_yc = st.number_input("Cantidad ventanas", value=0, step=1, key="cant_ventanas_yc")
                                ancho_ventana_yc = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_yc")
                                alto_ventana_yc  = st.number_input("Alto ventana (m)", value=0.0, key="alto_ventana_yc")

                            desp_yc = st.slider("% Desperdicio", 0, 20, 10, key="desp_yc")
                            area_vanos_yc = ((cant_puertas_yc * ancho_puerta_yc * alto_puerta_yc) +
                                            (cant_ventanas_yc * ancho_ventana_yc * alto_ventana_yc)) * n_caras_yc
                            area_neta_yc = area_bruta_yc - area_vanos_yc
                            area_plancha_yc = YESO_CARTON[yeso_tipo]["area"]
                            cant_exacta_yc, cant_desp_yc, sobra_yc, desp_m2_yc = calcular_planchas(area_neta_yc, area_plancha_yc, desp_yc)
                            tornillos_yc = calcular_tornillos_plancha(cant_desp_yc, area_plancha_yc, 0.30, 0.40)
                            st.write("---")
                            st.info(f"Área neta: {area_neta_yc:.2f} m²")
                            st.info(f"Planchas exactas: {cant_exacta_yc:.1f} unidades")
                            st.success(f"Planchas con {desp_yc}% desperdicio: {cant_desp_yc:.0f} unidades")
                            st.text(f"Sobra última plancha: {sobra_yc:.2f} m²")
                            st.text(f"Desperdicio estimado: {desp_m2_yc:.2f} m²")
                            st.write("---")
                            st.info(f"🔩 Tornillo: {tornillo_yc} {medida_yc}")
                            st.info("Separación: bordes cada 30cm / centro cada 40cm")
                            st.success(f"Total tornillos: {tornillos_yc} unidades")
                            if area_neta_yc > 0:
                                registrar_pdf("Revestimientos", "Yeso Cartón", [
                                    ("Tipo", yeso_tipo),
                                    ("Área neta", f"{area_neta_yc:.2f} m²"),
                                    ("Planchas (c/desp.)", f"{cant_desp_yc:.0f} unidades"),
                                    ("Tornillos", f"{tornillos_yc} unidades"),
                                ])

                        with st.expander("1.2 Terciado Ranurado", expanded=False):
                            st.caption("Revestimiento muros y cielos interiores")
                            tr_tipo = st.selectbox("Tipo", list(TERCIADO_RANURADO.keys()), key="tr_tipo")
                            estructura_tr = st.selectbox("Tipo de estructura", ["Metalcon", "Madera"], key="estructura_tr")
                            tornillo_tr = "Clavo sin cabeza (punta) o Tornillo madera" if estructura_tr == "Madera" else "Tornillo Kover o Lenteja Punta Fina"
                            medida_tr = "2\"" if estructura_tr == "Madera" else "8x1 1/4\""

                            area_bruta_tr = seccion_muro("tr")

                            tv1, tv2 = st.columns(2)
                            with tv1:
                                cant_puertas_tr = st.number_input("Cantidad puertas", value=0, step=1, key="cant_puertas_tr")
                                ancho_puerta_tr = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_tr")
                                alto_puerta_tr  = st.number_input("Alto puerta (m)", value=0.0, key="alto_puerta_tr")
                            with tv2:
                                cant_ventanas_tr = st.number_input("Cantidad ventanas", value=0, step=1, key="cant_ventanas_tr")
                                ancho_ventana_tr = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_tr")
                                alto_ventana_tr  = st.number_input("Alto ventana (m)", value=0.0, key="alto_ventana_tr")

                            desp_tr = st.slider("% Desperdicio", 0, 20, 10, key="desp_tr")
                            area_vanos_tr = ((cant_puertas_tr * ancho_puerta_tr * alto_puerta_tr) +
                                            (cant_ventanas_tr * ancho_ventana_tr * alto_ventana_tr))
                            area_neta_tr = area_bruta_tr - area_vanos_tr
                            area_plancha_tr = TERCIADO_RANURADO[tr_tipo]["area"]
                            cant_exacta_tr, cant_desp_tr, sobra_tr, desp_m2_tr = calcular_planchas(area_neta_tr, area_plancha_tr, desp_tr)
                            tornillos_tr = calcular_tornillos_plancha(cant_desp_tr, area_plancha_tr, 0.15, 0.30)
                            st.write("---")
                            st.info(f"Área neta: {area_neta_tr:.2f} m²")
                            st.info(f"Planchas exactas: {cant_exacta_tr:.1f} unidades")
                            st.success(f"Planchas con {desp_tr}% desperdicio: {cant_desp_tr:.0f} unidades")
                            st.text(f"Sobra última plancha: {sobra_tr:.2f} m²")
                            st.text(f"Desperdicio estimado: {desp_m2_tr:.2f} m²")
                            st.write("---")
                            st.info(f"🔩 Fijación: {tornillo_tr} {medida_tr}")
                            st.info("Separación: perímetro cada 15cm / apoyos internos cada 30cm")
                            st.success(f"Total fijaciones: {tornillos_tr} unidades")
                            st.caption("⚠️ Distancia mínima al borde: 1cm")
                            if area_neta_tr > 0:
                                registrar_pdf("Revestimientos", "Terciado Ranurado", [
                                    ("Tipo", tr_tipo),
                                    ("Área neta", f"{area_neta_tr:.2f} m²"),
                                    ("Planchas (c/desp.)", f"{cant_desp_tr:.0f} unidades"),
                                    ("Fijaciones", f"{tornillos_tr} unidades"),
                                ])

# ============================
# EXTERIOR
# ============================
                if ver(revest, "exterior"):
                    with st.expander("2. Revestimiento Exterior", expanded=False):

                        with st.expander("2.1 OSB", expanded=False):
                            st.caption("Tablero estructural - 1,22x2,44m")
                            osb_tipo = st.selectbox("Espesor OSB", list(OSB_TIPOS.keys()), key="osb_tipo")
                            estructura_osb = st.selectbox("Tipo de estructura", ["Metalcon", "Madera"], key="estructura_osb")
                            tornillo_osb = "Clavo Anillado Galvanizado" if estructura_osb == "Madera" else "Tornillo Punta Broca con aletas (Wafer)"
                            medida_osb = "2 1/2\" o 3\"" if estructura_osb == "Madera" else "8x1 1/4\""

                            area_bruta_osb = seccion_muro("osb")

                            ov1, ov2 = st.columns(2)
                            with ov1:
                                cant_puertas_osb = st.number_input("Cantidad puertas", value=0, step=1, key="cant_puertas_osb")
                                ancho_puerta_osb = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_osb")
                                alto_puerta_osb  = st.number_input("Alto puerta (m)", value=0.0, key="alto_puerta_osb")
                            with ov2:
                                cant_ventanas_osb = st.number_input("Cantidad ventanas", value=0, step=1, key="cant_ventanas_osb")
                                ancho_ventana_osb = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_osb")
                                alto_ventana_osb  = st.number_input("Alto ventana (m)", value=0.0, key="alto_ventana_osb")

                            desp_osb = st.slider("% Desperdicio", 0, 20, 10, key="desp_osb")
                            fibro_encima = st.checkbox("¿Agregar fibrocemento encima del OSB?", key="fibro_encima")
                            if fibro_encima:
                                fibro_osb_tipo = st.selectbox("Tipo fibrocemento", list(FIBROCEMENTO.keys()), key="fibro_osb_tipo")

                            area_vanos_osb = ((cant_puertas_osb * ancho_puerta_osb * alto_puerta_osb) +
                                            (cant_ventanas_osb * ancho_ventana_osb * alto_ventana_osb))
                            area_neta_osb = area_bruta_osb - area_vanos_osb
                            area_plancha_osb = OSB_TIPOS[osb_tipo]["area"]
                            cant_exacta_osb, cant_desp_osb, sobra_osb, desp_m2_osb = calcular_planchas(area_neta_osb, area_plancha_osb, desp_osb)
                            tornillos_osb = calcular_tornillos_plancha(cant_desp_osb, area_plancha_osb, 0.15, 0.30)
                            st.write("---")
                            st.info(f"Área neta: {area_neta_osb:.2f} m²")
                            st.info(f"Planchas exactas: {cant_exacta_osb:.1f} unidades")
                            st.success(f"Planchas con {desp_osb}% desperdicio: {cant_desp_osb:.0f} unidades")
                            st.text(f"Sobra última plancha: {sobra_osb:.2f} m²")
                            st.write("---")
                            st.info(f"🔩 Fijación: {tornillo_osb} {medida_osb}")
                            st.info("Separación: perímetro cada 15cm / apoyos internos cada 30cm")
                            st.success(f"Total fijaciones: {tornillos_osb} unidades")
                            st.caption("⚠️ Dejar holgura de 3mm entre planchas para dilatación")
                            if fibro_encima:
                                area_plancha_fo = FIBROCEMENTO[fibro_osb_tipo]["area"]
                                cant_exacta_fo, cant_desp_fo, sobra_fo, desp_m2_fo = calcular_planchas(area_neta_osb, area_plancha_fo, desp_osb)
                                tornillos_fo = calcular_tornillos_plancha(cant_desp_fo, area_plancha_fo, 0.175, 0.275)
                                st.write("---")
                                st.subheader("Fibrocemento sobre OSB")
                                st.info(f"Planchas: {cant_desp_fo:.0f} unidades")
                                st.info("🔩 Tornillo Autoperforante Punta Broca 8x1\"")
                                st.success(f"Total tornillos fibrocemento: {tornillos_fo} unidades")
                            if area_neta_osb > 0:
                                items_osb = [
                                    ("Espesor", osb_tipo),
                                    ("Área neta", f"{area_neta_osb:.2f} m²"),
                                    ("Planchas OSB (c/desp.)", f"{cant_desp_osb:.0f} unidades"),
                                    ("Fijaciones", f"{tornillos_osb} unidades"),
                                ]
                                if fibro_encima:
                                    items_osb.append(("Planchas fibrocemento", f"{cant_desp_fo:.0f} unidades"))
                                    items_osb.append(("Tornillos fibrocemento", f"{tornillos_fo} unidades"))
                                registrar_pdf("Revestimientos", "OSB", items_osb)

                        with st.expander("2.2 Fibrocemento", expanded=False):
                            st.caption("Placa fibrocemento exterior - 1,20x2,40m")
                            fc_tipo = st.selectbox("Tipo de Fibrocemento", list(FIBROCEMENTO.keys()), key="fc_tipo")

                            area_bruta_fc = seccion_muro("fc")

                            fv1, fv2 = st.columns(2)
                            with fv1:
                                cant_puertas_fc = st.number_input("Cantidad puertas", value=0, step=1, key="cant_puertas_fc")
                                ancho_puerta_fc = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_fc")
                                alto_puerta_fc  = st.number_input("Alto puerta (m)", value=0.0, key="alto_puerta_fc")
                            with fv2:
                                cant_ventanas_fc = st.number_input("Cantidad ventanas", value=0, step=1, key="cant_ventanas_fc")
                                ancho_ventana_fc = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_fc")
                                alto_ventana_fc  = st.number_input("Alto ventana (m)", value=0.0, key="alto_ventana_fc")

                            desp_fc = st.slider("% Desperdicio", 0, 20, 10, key="desp_fc")
                            area_vanos_fc = ((cant_puertas_fc * ancho_puerta_fc * alto_puerta_fc) +
                                            (cant_ventanas_fc * ancho_ventana_fc * alto_ventana_fc))
                            area_neta_fc = area_bruta_fc - area_vanos_fc
                            area_plancha_fc = FIBROCEMENTO[fc_tipo]["area"]
                            cant_exacta_fc, cant_desp_fc, sobra_fc, desp_m2_fc = calcular_planchas(area_neta_fc, area_plancha_fc, desp_fc)
                            tornillos_fc = calcular_tornillos_plancha(cant_desp_fc, area_plancha_fc, 0.175, 0.275)
                            st.write("---")
                            st.info(f"Área neta: {area_neta_fc:.2f} m²")
                            st.info(f"Planchas exactas: {cant_exacta_fc:.1f} unidades")
                            st.success(f"Planchas con {desp_fc}% desperdicio: {cant_desp_fc:.0f} unidades")
                            st.text(f"Sobra última plancha: {sobra_fc:.2f} m²")
                            st.write("---")
                            st.info("🔩 Tornillo Autoperforante Punta Broca con aletas 8x1\"")
                            st.info("Separación: perímetro cada 15-20cm / apoyos internos cada 25-30cm")
                            st.success(f"Total tornillos: {tornillos_fc} unidades")
                            st.caption("⚠️ Distancia mínima borde horizontal: 1,5cm / vertical: 1cm")
                            if area_neta_fc > 0:
                                registrar_pdf("Revestimientos", "Fibrocemento", [
                                    ("Tipo", fc_tipo),
                                    ("Área neta", f"{area_neta_fc:.2f} m²"),
                                    ("Planchas (c/desp.)", f"{cant_desp_fc:.0f} unidades"),
                                    ("Tornillos", f"{tornillos_fc} unidades"),
                                ])

                        with st.expander("2.3 Terciado Estructural", expanded=False):
                            st.caption("Placa estructural exterior - 1,22x2,44m")
                            te_tipo = st.selectbox("Espesor", list(TERCIADO_ESTRUCTURAL.keys()), key="te_tipo")
                            estructura_te = st.selectbox("Tipo de estructura", ["Metalcon", "Madera"], key="estructura_te")
                            tornillo_te = "Clavo Anillado Galvanizado" if estructura_te == "Madera" else "Tornillo Wafer Punta Broca"
                            medida_te = "2 1/2\"" if estructura_te == "Madera" else "8x1 1/4\" o 8x1 1/2\""

                            area_bruta_te = seccion_muro("te")

                            tev1, tev2 = st.columns(2)
                            with tev1:
                                cant_puertas_te = st.number_input("Cantidad puertas", value=0, step=1, key="cant_puertas_te")
                                ancho_puerta_te = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_te")
                                alto_puerta_te  = st.number_input("Alto puerta (m)", value=0.0, key="alto_puerta_te")
                            with tev2:
                                cant_ventanas_te = st.number_input("Cantidad ventanas", value=0, step=1, key="cant_ventanas_te")
                                ancho_ventana_te = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_te")
                                alto_ventana_te  = st.number_input("Alto ventana (m)", value=0.0, key="alto_ventana_te")

                            desp_te = st.slider("% Desperdicio", 0, 20, 10, key="desp_te")
                            area_vanos_te = ((cant_puertas_te * ancho_puerta_te * alto_puerta_te) +
                                            (cant_ventanas_te * ancho_ventana_te * alto_ventana_te))
                            area_neta_te = area_bruta_te - area_vanos_te
                            area_plancha_te = TERCIADO_ESTRUCTURAL[te_tipo]["area"]
                            cant_exacta_te, cant_desp_te, sobra_te, desp_m2_te = calcular_planchas(area_neta_te, area_plancha_te, desp_te)
                            tornillos_te = calcular_tornillos_plancha(cant_desp_te, area_plancha_te, 0.15, 0.30)
                            st.write("---")
                            st.info(f"Área neta: {area_neta_te:.2f} m²")
                            st.info(f"Planchas exactas: {cant_exacta_te:.1f} unidades")
                            st.success(f"Planchas con {desp_te}% desperdicio: {cant_desp_te:.0f} unidades")
                            st.text(f"Sobra última plancha: {sobra_te:.2f} m²")
                            st.write("---")
                            st.info(f"🔩 Fijación: {tornillo_te} {medida_te}")
                            st.info("Separación: perímetro cada 15cm / apoyos internos cada 30cm")
                            st.success(f"Total fijaciones: {tornillos_te} unidades")
                            st.caption("⚠️ Instalar fibra perpendicular a los apoyos para mayor resistencia")
                            if area_neta_te > 0:
                                registrar_pdf("Revestimientos", "Terciado Estructural", [
                                    ("Espesor", te_tipo),
                                    ("Área neta", f"{area_neta_te:.2f} m²"),
                                    ("Planchas (c/desp.)", f"{cant_desp_te:.0f} unidades"),
                                    ("Fijaciones", f"{tornillos_te} unidades"),
                                ])

                        with st.expander("2.4 Siding Fibrocemento", expanded=False):
                            st.caption("Volcan, Pizarreño Cedral, Nativa - 19cm x 3,66m")
                            estructura_sid = st.selectbox("Tipo de estructura", ["Metalcon", "Madera"], key="estructura_sid")
                            tornillo_sid = "Tornillo Cabeza Lenteja (Wafer) Punta Fina" if estructura_sid == "Metalcon" else "Clavo Galvanizado Cabeza Plana"
                            medida_sid = "8x1 1/4\"" if estructura_sid == "Metalcon" else "1 1/2\" o 2\""

                            area_bruta_sid = seccion_muro("sid")

                            sv1, sv2 = st.columns(2)
                            with sv1:
                                cant_puertas_sid = st.number_input("Cantidad puertas", value=0, step=1, key="cant_puertas_sid")
                                ancho_puerta_sid = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_sid")
                                alto_puerta_sid  = st.number_input("Alto puerta (m)", value=0.0, key="alto_puerta_sid")
                            with sv2:
                                cant_ventanas_sid = st.number_input("Cantidad ventanas", value=0, step=1, key="cant_ventanas_sid")
                                ancho_ventana_sid = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_sid")
                                alto_ventana_sid  = st.number_input("Alto ventana (m)", value=0.0, key="alto_ventana_sid")

                            traslape_sid  = st.selectbox("Traslape", ["25mm", "30mm"], key="traslape_sid")
                            sep_mont_sid  = st.selectbox("Separación montantes", ["0,40m", "0,60m"], key="sep_mont_sid")
                            sep_valor_sid = 0.40 if "0,40" in sep_mont_sid else 0.60
                            desp_sid = st.slider("% Desperdicio", 0, 20, 10, key="desp_sid")
                            rend_tabla = 0.70 if "25" in traslape_sid else 0.44

                            area_vanos_sid = ((cant_puertas_sid * ancho_puerta_sid * alto_puerta_sid) +
                                            (cant_ventanas_sid * ancho_ventana_sid * alto_ventana_sid))
                            area_neta_sid = area_bruta_sid - area_vanos_sid
                            cant_tablas_sid  = area_neta_sid / rend_tabla if rend_tabla > 0 else 0
                            cant_tablas_desp = cant_tablas_sid * (1 + desp_sid / 100)
                            ml_tablas_sid    = cant_tablas_desp * 3.66
                            largo_sid_total  = sum(sec["largo"] * sec["cant"] for sec in st.session_state.get("secciones_sid", [{"largo": 0, "cant": 0}]))
                            tornillos_sid    = round((largo_sid_total / sep_valor_sid) * cant_tablas_desp * (1 + desp_sid / 100))

                            st.write("---")
                            st.info(f"Área neta: {area_neta_sid:.2f} m²")
                            st.info(f"Rendimiento por tabla: {rend_tabla} m² (traslape {traslape_sid})")
                            st.info(f"Tablas exactas: {cant_tablas_sid:.0f} unidades")
                            st.success(f"Tablas con {desp_sid}% desperdicio: {cant_tablas_desp:.0f} unidades")
                            st.text(f"Metros lineales totales: {ml_tablas_sid:.2f} ml")
                            st.write("---")
                            st.info(f"🔩 Fijación: {tornillo_sid} {medida_sid}")
                            st.info(f"1 fijación por montante cada {sep_mont_sid}")
                            st.success(f"Total fijaciones: {tornillos_sid} unidades")
                            st.caption("⚠️ Fijación a 2cm del borde superior, quedará tapada por la tabla siguiente")
                            if area_neta_sid > 0:
                                registrar_pdf("Revestimientos", "Siding Fibrocemento", [
                                    ("Traslape", traslape_sid),
                                    ("Área neta", f"{area_neta_sid:.2f} m²"),
                                    ("Tablas (c/desp.)", f"{cant_tablas_desp:.0f} unidades"),
                                    ("Metros lineales", f"{ml_tablas_sid:.2f} ml"),
                                    ("Fijaciones", f"{tornillos_sid} unidades"),
                                ])
# ============================
# PISOS Y PAVIMENTOS
# ============================
    if ver_rubro(pisos) and rubro_permitido("pisos"):
        with st.expander("Pisos y Pavimentos", expanded=False):

            # ============================
            # DATOS
            # ============================
            CERAMICO_MEDIDAS = {
                "30x30 cm":   {"largo": 0.30, "ancho": 0.30, "area": 0.09, "doble_pegado": False},
                "45x45 cm":   {"largo": 0.45, "ancho": 0.45, "area": 0.2025, "doble_pegado": False},
                "60x60 cm":   {"largo": 0.60, "ancho": 0.60, "area": 0.36, "doble_pegado": False},
                "80x80 cm":   {"largo": 0.80, "ancho": 0.80, "area": 0.64, "doble_pegado": False},
                "60x120 cm":  {"largo": 0.60, "ancho": 1.20, "area": 0.72, "doble_pegado": True},
            }

            PISO_FLOTANTE = {
                "7mm - Tránsito Residencial (AC3)":  {"espesor": 7},
                "8mm - Tránsito Residencial (AC4)":  {"espesor": 8},
                "10mm - Tránsito Comercial (AC5)":   {"espesor": 10},
                "12mm - Tránsito Comercial (AC5)":   {"espesor": 12},
            }

            MADERA_DECK = {
                "Pino Impregnado CCA":  {"precio_ref": "económico"},
                "Lenga":                {"precio_ref": "premium"},
                "Raulí":                {"precio_ref": "premium"},
                "Ipé":                  {"precio_ref": "alta gama"},
                "Coigüe":               {"precio_ref": "alta gama"},
            }

            # ============================
            # 1. CERÁMICO / PORCELANATO
            # ============================
            if ver(pisos, "ceramico"):
                with st.expander("1. Cerámico / Porcelanato", expanded=False):

                        cer_medida = st.selectbox("Medida", list(CERAMICO_MEDIDAS.keys()), key="cer_medida")
                        cer = CERAMICO_MEDIDAS[cer_medida]

                        if cer["doble_pegado"]:
                            st.warning("⚠️ Formato grande (60x120cm): Requiere doble pegado y pegamento de mayor adherencia (tipo AC o Flex)")

                        ce1, ce2 = st.columns(2)
                        with ce1:
                            largo_cer = st.number_input("Largo área (m)", value=0.0, key="largo_cer")
                        with ce2:
                            ancho_cer = st.number_input("Ancho área (m)", value=0.0, key="ancho_cer")

                        # Vanos
                        cv1, cv2 = st.columns(2)
                        with cv1:
                            cant_vanos_cer = st.number_input("Cantidad de vanos", value=0, step=1, key="cant_vanos_cer")
                        with cv2:
                            area_vano_cer = st.number_input("Área por vano (m²)", value=0.0, key="area_vano_cer")

                        junta_cer = st.selectbox("Ancho de junta (cantería)", ["2mm", "3mm", "5mm"], key="junta_cer")
                        desp_cer = st.slider("% Desperdicio", 0, 20, 10, key="desp_cer")

                        area_bruta_cer = largo_cer * ancho_cer
                        area_vanos_cer = cant_vanos_cer * area_vano_cer
                        area_neta_cer = area_bruta_cer - area_vanos_cer

                        # Cerámicos
                        ceramicos_exactos = area_neta_cer / cer["area"] if cer["area"] > 0 else 0
                        ceramicos_desp = ceramicos_exactos * (1 + desp_cer / 100)

                        # Pegamento
                        rend_pega = 4.5  # promedio 4-5 kg/m²
                        if cer["doble_pegado"]:
                            rend_pega = 8.0  # doble pegado
                        kg_pegamento = area_neta_cer * rend_pega
                        bolsas_pegamento = kg_pegamento / 25

                        # Fragüe según junta
                        junta_val = float(junta_cer.replace("mm", ""))
                        if junta_val <= 2:
                            rend_frag = 0.30
                        elif junta_val <= 3:
                            rend_frag = 0.40
                        else:
                            rend_frag = 0.50
                        kg_frag = area_neta_cer * rend_frag
                        bolsas_frag = kg_frag / 5  # bolsas de 5kg

                        st.write("---")
                        st.info(f"Área neta: {area_neta_cer:.2f} m²")
                        st.info(f"Piezas por m²: {1/cer['area']:.1f} unidades")
                        st.info(f"Cerámicos exactos: {ceramicos_exactos:.0f} unidades")
                        st.success(f"Con {desp_cer}% desperdicio: {ceramicos_desp:.0f} unidades")

                        st.write("---")
                        st.subheader("🧴 Pegamento")
                        if cer["doble_pegado"]:
                            st.caption("Doble pegado: en el piso y en la trasera de la palmeta")
                        st.info(f"Pegamento necesario: {kg_pegamento:.1f} kg")
                        st.success(f"Bolsas de 25kg: {bolsas_pegamento:.0f} bolsas")
                        st.caption(f"Rendimiento: {rend_pega} kg/m² | Rinde aprox. {25/rend_pega:.1f} m² por bolsa")

                        st.write("---")
                        st.subheader("🪣 Fragüe")
                        st.info(f"Fragüe necesario: {kg_frag:.1f} kg")
                        st.success(f"Bolsas de 5kg: {bolsas_frag:.0f} bolsas")
                        st.caption(f"Rendimiento: {rend_frag} kg/m² para junta de {junta_cer}")
                        if area_neta_cer > 0:
                            registrar_pdf("Pisos y Pavimentos", "Cerámico / Porcelanato", [
                                ("Medida", cer_medida),
                                ("Área neta", f"{area_neta_cer:.2f} m²"),
                                ("Cerámicos (c/desp.)", f"{ceramicos_desp:.0f} unidades"),
                                ("Pegamento", f"{bolsas_pegamento:.0f} bolsas de 25kg"),
                                ("Fragüe", f"{bolsas_frag:.0f} bolsas de 5kg"),
                            ])

            # ============================
            # 2. PISO FLOTANTE
            # ============================
            if ver(pisos, "flotante"):
                with st.expander("2. Piso Flotante", expanded=False):

                        pf_tipo = st.selectbox("Tipo de piso flotante", list(PISO_FLOTANTE.keys()), key="pf_tipo")

                        pf1, pf2 = st.columns(2)
                        with pf1:
                            largo_pf = st.number_input("Largo área (m)", value=0.0, key="largo_pf")
                        with pf2:
                            ancho_pf = st.number_input("Ancho área (m)", value=0.0, key="ancho_pf")

                        pfv1, pfv2 = st.columns(2)
                        with pfv1:
                            cant_vanos_pf = st.number_input("Cantidad de vanos", value=0, step=1, key="cant_vanos_pf")
                        with pfv2:
                            area_vano_pf = st.number_input("Área por vano (m²)", value=0.0, key="area_vano_pf")

                        sobre_losa = st.checkbox("¿Va sobre losa de hormigón (primer piso)?", key="sobre_losa")
                        desp_pf = st.slider("% Desperdicio", 0, 20, 10, key="desp_pf")

                        area_bruta_pf = largo_pf * ancho_pf
                        area_vanos_pf = cant_vanos_pf * area_vano_pf
                        area_neta_pf = area_bruta_pf - area_vanos_pf

                        area_pf_desp = area_neta_pf * (1 + desp_pf / 100)

                        # Cajas según comercio (rendimiento típico 2,2 m² por caja)
                        M2_POR_CAJA_PF = 2.2
                        cajas_pf = math.ceil(area_pf_desp / M2_POR_CAJA_PF) if area_pf_desp > 0 else 0

                        st.write("---")
                        st.info(f"Área neta: {area_neta_pf:.2f} m²")
                        st.success(f"Piso flotante con {desp_pf}% desperdicio: {area_pf_desp:.2f} m²")
                        st.success(f"Cajas a comprar: {cajas_pf} cajas")
                        st.caption(f"Rendimiento estimado: {M2_POR_CAJA_PF} m² por caja (verifica según marca)")

                        st.write("---")
                        st.subheader("🧸 Espuma Niveladora (Manta Polietileno)")
                        st.info(f"Espuma 2-3mm necesaria: {area_neta_pf:.2f} m²")
                        st.caption("1 m² de espuma por 1 m² de piso flotante")

                        if sobre_losa:
                            st.write("---")
                            st.subheader("💧 Barrera de Humedad (Buna/Film Polietileno)")
                            st.warning("⚠️ Instalación sobre losa: Se requiere barrera de humedad bajo la espuma")
                            st.info(f"Film polietileno necesario: {area_neta_pf:.2f} m²")
                            st.caption("Traslapar 20cm en uniones y subir 10cm por los muros")
                        if area_neta_pf > 0:
                            items_pf = [
                                ("Tipo", pf_tipo),
                                ("Área neta", f"{area_neta_pf:.2f} m²"),
                                ("Piso flotante (c/desp.)", f"{area_pf_desp:.2f} m²"),
                                ("Cajas a comprar", f"{cajas_pf} cajas (≈{M2_POR_CAJA_PF} m²/caja)"),
                                ("Espuma niveladora", f"{area_neta_pf:.2f} m²"),
                            ]
                            if sobre_losa:
                                items_pf.append(("Film polietileno", f"{area_neta_pf:.2f} m²"))
                            registrar_pdf("Pisos y Pavimentos", "Piso Flotante", items_pf)

            # ============================
            # 3. BALDOSA
            # ============================
            if ver(pisos, "baldosa"):
                with st.expander("3. Baldosa", expanded=False):
                        st.caption("Microvibrada o calcárea 2-4cm espesor - Mortero de pega tradicional")

                        ba1, ba2, ba3, ba4 = st.columns(4)
                        with ba1:
                            largo_bal = st.number_input("Largo baldosa (cm)", value=0.0, key="largo_bal")
                        with ba2:
                            ancho_bal = st.number_input("Ancho baldosa (cm)", value=0.0, key="ancho_bal")
                        with ba3:
                            largo_area_bal = st.number_input("Largo área (m)", value=0.0, key="largo_area_bal")
                        with ba4:
                            ancho_area_bal = st.number_input("Ancho área (m)", value=0.0, key="ancho_area_bal")

                        bv1, bv2 = st.columns(2)
                        with bv1:
                            cant_vanos_bal = st.number_input("Cantidad de vanos", value=0, step=1, key="cant_vanos_bal")
                        with bv2:
                            area_vano_bal = st.number_input("Área por vano (m²)", value=0.0, key="area_vano_bal")

                        espesor_bal = st.selectbox("Espesor baldosa", ["2cm", "3cm", "4cm"], key="espesor_bal")
                        junta_bal = st.selectbox("Ancho de junta", ["5mm", "8mm", "10mm"], key="junta_bal")
                        desp_bal = st.slider("% Desperdicio", 0, 20, 10, key="desp_bal")

                        area_baldosa = (largo_bal / 100) * (ancho_bal / 100)
                        area_bruta_bal = largo_area_bal * ancho_area_bal
                        area_vanos_bal = cant_vanos_bal * area_vano_bal
                        area_neta_bal = area_bruta_bal - area_vanos_bal

                        baldosas_exactas = area_neta_bal / area_baldosa if area_baldosa > 0 else 0
                        baldosas_desp = baldosas_exactas * (1 + desp_bal / 100)

                        # Mortero de pega (dosificación 1:3 arena:cemento)
                        # Espesor cama mortero aprox 3cm
                        espesor_mort = 0.03
                        vol_mortero_bal = area_neta_bal * espesor_mort
                        cemento_bal_kg = vol_mortero_bal * 400
                        cemento_bal_sacos = cemento_bal_kg / 25
                        arena_bal_m3 = vol_mortero_bal * 1.20

                        st.write("---")
                        st.info(f"Área neta: {area_neta_bal:.2f} m²")
                        st.info(f"Baldosas por m²: {1/area_baldosa:.1f} unidades" if area_baldosa > 0 else "Ingresa dimensiones")
                        st.info(f"Baldosas exactas: {baldosas_exactas:.0f} unidades")
                        st.success(f"Con {desp_bal}% desperdicio: {baldosas_desp:.0f} unidades")

                        st.write("---")
                        st.subheader("🧱 Mortero de Pega (1:3 cemento:arena)")
                        st.info(f"Cemento: {cemento_bal_sacos:.0f} sacos de 25kg")
                        st.info(f"Arena: {arena_bal_m3:.2f} m³")
                        st.caption(f"Espesor cama mortero: 3cm | Volumen mortero: {vol_mortero_bal:.3f} m³")
                        st.warning("⚠️ Baldosa puede requerir pulido o sellado posterior según tipo")
                        if area_neta_bal > 0:
                            registrar_pdf("Pisos y Pavimentos", "Baldosa", [
                                ("Área neta", f"{area_neta_bal:.2f} m²"),
                                ("Baldosas (c/desp.)", f"{baldosas_desp:.0f} unidades"),
                                ("Cemento mortero", f"{cemento_bal_sacos:.0f} sacos"),
                                ("Arena mortero", f"{arena_bal_m3:.2f} m³"),
                            ])

            # ============================
            # 4. DECK DE MADERA
            # ============================
            if ver(pisos, "deck"):
                with st.expander("4. Deck de Madera", expanded=False):

                        dk_madera = st.selectbox("Tipo de madera", list(MADERA_DECK.keys()), key="dk_madera")

                        if MADERA_DECK[dk_madera]["precio_ref"] == "económico":
                            st.caption("✅ Pino Impregnado CCA: Resistente a humedad y termitas de fábrica")
                        elif MADERA_DECK[dk_madera]["precio_ref"] == "premium":
                            st.warning("⚠️ Requiere tratamiento con impregnante protector (Cerestain o Sipasol)")
                        else:
                            st.caption("🌟 Madera de alta gama - Mayor dureza y durabilidad")

                        dk1, dk2 = st.columns(2)
                        with dk1:
                            ancho_tabla_dk = st.selectbox("Ancho tabla", ["4\" (90mm)", "5\" (120mm)"], key="ancho_tabla_dk")
                        with dk2:
                            largo_tabla_dk = st.selectbox("Largo tabla", ["3,20m", "4,00m"], key="largo_tabla_dk")

                        ancho_val_dk = 0.090 if "4\"" in ancho_tabla_dk else 0.120
                        largo_val_dk = 3.20 if "3,20" in largo_tabla_dk else 4.00

                        espesor_dk = st.selectbox("Espesor tabla", ["1\" (19mm)", "1 1/2\" (32mm)"], key="espesor_dk")
                        sep_dk = st.selectbox("Separación entre tablas", ["5mm", "8mm", "10mm"], key="sep_dk")
                        sep_val_dk = float(sep_dk.replace("mm", "")) / 1000

                        dk3, dk4 = st.columns(2)
                        with dk3:
                            largo_deck = st.number_input("Largo área deck (m)", value=0.0, key="largo_deck")
                        with dk4:
                            ancho_deck = st.number_input("Ancho área deck (m)", value=0.0, key="ancho_deck")

                        desp_dk = st.slider("% Desperdicio", 0, 20, 10, key="desp_dk")

                        area_neta_dk = largo_deck * ancho_deck

                        # Tablas necesarias
                        ancho_util_dk = ancho_val_dk + sep_val_dk
                        tablas_por_ml = 1 / ancho_util_dk
                        ml_tablas_dk = area_neta_dk * tablas_por_ml
                        cant_tablas_dk = ml_tablas_dk / largo_val_dk
                        cant_tablas_desp_dk = cant_tablas_dk * (1 + desp_dk / 100)

                        # Tornillos deck (2 por tabla por vigueta, viguetas cada 40cm)
                        viguetas_dk = largo_deck / 0.40
                        tornillos_dk = cant_tablas_desp_dk * viguetas_dk * 2

                        st.write("---")
                        st.info(f"Área deck: {area_neta_dk:.2f} m²")
                        st.info(f"Tablas exactas: {cant_tablas_dk:.0f} unidades de {largo_val_dk}m")
                        st.success(f"Con {desp_dk}% desperdicio: {cant_tablas_desp_dk:.0f} tablas")
                        st.text(f"Metros lineales totales: {cant_tablas_desp_dk * largo_val_dk:.2f} ml")

                        st.write("---")
                        st.subheader("🔩 Fijaciones")
                        st.info(f"Tornillos galvanizados: {tornillos_dk:.0f} unidades")
                        st.caption("2 tornillos por tabla en cada vigueta (viguetas cada 40cm)")

                        if dk_madera != "Pino Impregnado CCA":
                            st.write("---")
                            st.subheader("🪵 Tratamiento recomendado")
                            st.warning(f"⚠️ {dk_madera}: Aplicar impregnante protector tipo Cerestain o Sipasol antes de instalar")
                        if area_neta_dk > 0:
                            registrar_pdf("Pisos y Pavimentos", "Deck de Madera", [
                                ("Madera", dk_madera),
                                ("Área deck", f"{area_neta_dk:.2f} m²"),
                                ("Tablas (c/desp.)", f"{cant_tablas_desp_dk:.0f} de {largo_val_dk}m"),
                                ("Tornillos galvanizados", f"{tornillos_dk:.0f} unidades"),
                            ])

# ============================
# TERMINACIONES
# ============================
    if ver_rubro(term) and rubro_permitido("terminaciones"):
        with st.expander("Terminaciones", expanded=False):

            # ============================
            # DATOS
            # ============================
            PINTURAS = {
                "Látex Interior": {
                    "rendimiento": 10,  # m² por litro por mano
                    "descripcion": "Dormitorios, living, zonas secas"
                },
                "Látex Exterior": {
                    "rendimiento": 8,
                    "descripcion": "Fachadas y muros exteriores"
                },
                "Esmalte": {
                    "rendimiento": 12,
                    "descripcion": "Puertas, ventanas, metales"
                },
                "Esmalte al Agua (mate)": {
                    "rendimiento": 10,
                    "descripcion": "Cielos y zonas húmedas, acabado mate (sin reflejos)"
                },
                "Anticorrosivo": {
                    "rendimiento": 8,
                    "descripcion": "Superficies metálicas expuestas"
                },
            }

            ESTUCOS = {
                "Premezclado Exterior (Sika/Melón/Weber)": {
                    "rendimiento_saco": 1.5,  # m² por saco de 25kg a 15mm
                    "peso_saco": 25,
                    "tipo": "premezclado"
                },
                "Premezclado Antihumedad (Hidrófugo)": {
                    "rendimiento_saco": 1.5,
                    "peso_saco": 25,
                    "tipo": "premezclado"
                },
                "Tradicional (Cemento+Arena+Cal)": {
                    "rendimiento_saco": None,
                    "tipo": "tradicional"
                },
                "Fino (Maquillaje/Terminación)": {
                    "rendimiento_saco": 3.0,  # m² por saco a 5mm
                    "peso_saco": 25,
                    "tipo": "fino"
                },
            }

            CIELOS_TIPOS = {
                "Yeso Cartón ST 10mm":   {"area_plancha": 2.88, "tipo": "plancha"},
                "Yeso Cartón ST 12,5mm": {"area_plancha": 2.88, "tipo": "plancha"},
                "Volcanita 10mm":        {"area_plancha": 2.88, "tipo": "plancha"},
                "MDF 3mm":               {"area_plancha": 2.98, "tipo": "plancha"},
                "MDF 6mm":               {"area_plancha": 2.98, "tipo": "plancha"},
                "Madera Machihembrada":  {"area_plancha": None, "tipo": "madera"},
            }

            ZOCALOS = {
                "MDF 7cm":    {"alto": 0.07, "tipo": "MDF"},
                "MDF 10cm":   {"alto": 0.10, "tipo": "MDF"},
                "MDF 15cm":   {"alto": 0.15, "tipo": "MDF"},
                "Madera 7cm": {"alto": 0.07, "tipo": "Madera"},
                "Madera 10cm":{"alto": 0.10, "tipo": "Madera"},
                "Madera 15cm":{"alto": 0.15, "tipo": "Madera"},
                "PVC 7cm":    {"alto": 0.07, "tipo": "PVC"},
                "PVC 10cm":   {"alto": 0.10, "tipo": "PVC"},
                "PVC 15cm":   {"alto": 0.15, "tipo": "PVC"},
            }

            # ============================
            # 1. PINTURA
            # ============================
            if ver(term, "pintura"):
                with st.expander("1. Pintura", expanded=False):

                        # =========================================================
                        # PINTURA DE MURO
                        # =========================================================
                        st.subheader("🧱 Pintura de Muro")
                        pintura_tipo = st.selectbox("Tipo de pintura (muro)", list(PINTURAS.keys()), key="pintura_tipo")
                        st.caption(PINTURAS[pintura_tipo]["descripcion"])

                        # --- Murallas: cada una con su propio largo y alto ---
                        st.markdown("**Murallas a pintar** (agrega cada muro con su medida)")
                        if "secciones_pin" not in st.session_state:
                            st.session_state.secciones_pin = [{"largo": 0.0, "alto": 0.0}]

                        area_bruta_pin = 0.0
                        for i, sec in enumerate(st.session_state.secciones_pin):
                            cpa, cpb, cpc = st.columns([3, 3, 1])
                            with cpa:
                                sec["largo"] = st.number_input(
                                    f"Largo muro {i+1} (m)", value=sec["largo"],
                                    min_value=0.0, step=0.1, key=f"pin_largo_{i}")
                            with cpb:
                                sec["alto"] = st.number_input(
                                    f"Alto muro {i+1} (m)", value=sec["alto"],
                                    min_value=0.0, step=0.1, key=f"pin_alto_{i}")
                            with cpc:
                                st.write("")
                                st.write("")
                                if len(st.session_state.secciones_pin) > 1 and st.button("🗑️", key=f"del_pin_{i}"):
                                    st.session_state.secciones_pin.pop(i)
                                    st.rerun()
                            area_bruta_pin += sec["largo"] * sec["alto"]

                        if st.button("➕ Agregar muro", key="add_pin"):
                            st.session_state.secciones_pin.append({"largo": 0.0, "alto": 0.0})
                            st.rerun()

                        # --- Descuento de puertas / ventanas (opcional) ---
                        area_vanos_pin = 0.0
                        descontar_vanos = st.checkbox(
                            "Descontar puertas / ventanas", key="pin_descontar_vanos",
                            help="Actívalo solo si las murallas tienen puertas o ventanas.")
                        if descontar_vanos:
                            pv1, pv2 = st.columns(2)
                            with pv1:
                                cant_puertas_pin = st.number_input("Cantidad puertas", value=0, step=1, key="cant_puertas_pin")
                                ancho_puerta_pin = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_pin")
                                alto_puerta_pin = st.number_input("Alto puerta (m)", value=0.0, key="alto_puerta_pin")
                            with pv2:
                                cant_ventanas_pin = st.number_input("Cantidad ventanas", value=0, step=1, key="cant_ventanas_pin")
                                ancho_ventana_pin = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_pin")
                                alto_ventana_pin = st.number_input("Alto ventana (m)", value=0.0, key="alto_ventana_pin")
                            area_vanos_pin = ((cant_puertas_pin * ancho_puerta_pin * alto_puerta_pin) +
                                              (cant_ventanas_pin * ancho_ventana_pin * alto_ventana_pin))

                        cant_manos = st.selectbox("Cantidad de manos (muro)", ["1 mano", "2 manos", "3 manos"], index=1, key="cant_manos")
                        n_manos = int(cant_manos[0])

                        area_neta_muro = max(area_bruta_pin - area_vanos_pin, 0.0)
                        rend_muro = PINTURAS[pintura_tipo]["rendimiento"]
                        litros_muro = (area_neta_muro * n_manos) / rend_muro if rend_muro else 0.0
                        galones_muro = math.ceil(litros_muro / 4) if litros_muro > 0 else 0

                        # =========================================================
                        # PINTURA DE CIELO (opcional, se cubica por separado)
                        # =========================================================
                        st.write("---")
                        usar_cielo = st.checkbox("Incluir pintura de cielo", key="pin_usar_cielo",
                                                 help="El cielo se cubica por superficie de piso (largo x ancho).")
                        area_cielo = 0.0
                        litros_cielo = 0.0
                        galones_cielo = 0
                        pintura_tipo_cielo = ""
                        if usar_cielo:
                            st.subheader("⬜ Pintura de Cielo")
                            # Para cielo solo se usan látex o esmalte al agua (acabado mate):
                            # disimulan imperfecciones y evitan reflejos de luz.
                            tipos_cielo = ["Látex Interior", "Esmalte al Agua (mate)"]
                            pintura_tipo_cielo = st.selectbox("Tipo de pintura (cielo)", tipos_cielo, key="pintura_tipo_cielo")
                            st.caption(PINTURAS[pintura_tipo_cielo]["descripcion"])

                            st.markdown("**Cielos a pintar** (agrega cada ambiente: largo x ancho)")
                            if "secciones_pin_cielo" not in st.session_state:
                                st.session_state.secciones_pin_cielo = [{"largo": 0.0, "ancho": 0.0}]
                            for i, sec in enumerate(st.session_state.secciones_pin_cielo):
                                cca, ccb, ccc = st.columns([3, 3, 1])
                                with cca:
                                    sec["largo"] = st.number_input(
                                        f"Largo ambiente {i+1} (m)", value=sec["largo"],
                                        min_value=0.0, step=0.1, key=f"pin_cielo_largo_{i}")
                                with ccb:
                                    sec["ancho"] = st.number_input(
                                        f"Ancho ambiente {i+1} (m)", value=sec["ancho"],
                                        min_value=0.0, step=0.1, key=f"pin_cielo_ancho_{i}")
                                with ccc:
                                    st.write("")
                                    st.write("")
                                    if len(st.session_state.secciones_pin_cielo) > 1 and st.button("🗑️", key=f"del_pin_cielo_{i}"):
                                        st.session_state.secciones_pin_cielo.pop(i)
                                        st.rerun()
                                area_cielo += sec["largo"] * sec["ancho"]

                            if st.button("➕ Agregar ambiente", key="add_pin_cielo"):
                                st.session_state.secciones_pin_cielo.append({"largo": 0.0, "ancho": 0.0})
                                st.rerun()

                            cant_manos_cielo = st.selectbox("Cantidad de manos (cielo)", ["1 mano", "2 manos", "3 manos"], index=1, key="cant_manos_cielo")
                            n_manos_cielo = int(cant_manos_cielo[0])
                            rend_cielo = PINTURAS[pintura_tipo_cielo]["rendimiento"]
                            litros_cielo = (area_cielo * n_manos_cielo) / rend_cielo if rend_cielo else 0.0
                            galones_cielo = math.ceil(litros_cielo / 4) if litros_cielo > 0 else 0

                        # Superficie total a pintar (muro + cielo) para complementos
                        area_total_pintar = area_neta_muro + area_cielo

                        # --- Complementos opcionales (cada uno por selector) ---
                        st.write("---")
                        st.markdown("**Complementos (opcional)** — sobre la superficie total a pintar")
                        usar_pasta = st.checkbox("Pasta Muro (masilla para juntas)", key="pin_usar_pasta")
                        usar_cinta = st.checkbox("Cinta para Juntas", key="pin_usar_cinta")
                        usar_sellador = st.checkbox("Sellador / Imprimante Base", key="pin_usar_sellador")

                        kg_pasta = 0.0
                        tarros_pasta = 0
                        ml_cinta = 0.0
                        rollos_cinta = 0
                        litros_sellador = 0.0
                        galones_sellador = 0

                        if usar_pasta:
                            st.write("---")
                            st.subheader("🪣 Pasta Muro (Masilla/Compuesto para Juntas)")
                            st.caption("Se aplica antes del sellador para rellenar juntas e imperfecciones")
                            kg_pasta = area_total_pintar * 0.30  # 300 g por m²
                            tarros_pasta = math.ceil(kg_pasta / 25) if kg_pasta > 0 else 0
                            st.info(f"Pasta muro necesaria: {kg_pasta:.1f} kg")
                            st.success(f"Tarros de 25kg: {tarros_pasta} tarros")

                        if usar_cinta:
                            st.write("---")
                            st.subheader("📏 Cinta para Juntas")
                            st.caption("Se aplica sobre las juntas entre planchas antes de la pasta")
                            ml_cinta = area_total_pintar / 2.88 * 4.84  # metros lineales de juntas por plancha
                            rollos_cinta = math.ceil(ml_cinta / 75) if ml_cinta > 0 else 0
                            st.info(f"Metros lineales de cinta: {ml_cinta:.1f} ml")
                            st.success(f"Rollos de 75m: {rollos_cinta} rollos")

                        if usar_sellador:
                            st.write("---")
                            st.subheader("🖌️ Sellador / Imprimante Base")
                            st.caption("Se aplica antes de la pintura para mejorar adherencia")
                            litros_sellador = area_total_pintar / 10  # 1 litro por 10 m²
                            galones_sellador = math.ceil(litros_sellador / 4) if litros_sellador > 0 else 0
                            st.info(f"Sellador necesario: {litros_sellador:.1f} litros")
                            st.success(f"Galones de 4 litros: {galones_sellador} galones")

                        # --- Resultados de pintura (muro y cielo por separado) ---
                        st.write("---")
                        st.subheader("🎨 Resumen de Pintura")
                        st.markdown(f"**Muro — {pintura_tipo}**")
                        st.info(f"Área muro: {area_neta_muro:.2f} m²  |  Manos: {n_manos}")
                        st.success(f"Pintura muro: {litros_muro:.1f} litros ({galones_muro} galones de 4 L)")
                        if usar_cielo:
                            st.markdown(f"**Cielo — {pintura_tipo_cielo}**")
                            st.info(f"Área cielo: {area_cielo:.2f} m²  |  Manos: {n_manos_cielo}")
                            st.success(f"Pintura cielo: {litros_cielo:.1f} litros ({galones_cielo} galones de 4 L)")

                        if area_neta_muro > 0 or area_cielo > 0:
                            items_pin = []
                            if area_neta_muro > 0:
                                items_pin += [
                                    ("Tipo muro", pintura_tipo),
                                    ("Área muro", f"{area_neta_muro:.2f} m²"),
                                    ("Pintura muro", f"{litros_muro:.1f} litros ({galones_muro} galones)"),
                                ]
                            if usar_cielo and area_cielo > 0:
                                items_pin += [
                                    ("Tipo cielo", pintura_tipo_cielo),
                                    ("Área cielo", f"{area_cielo:.2f} m²"),
                                    ("Pintura cielo", f"{litros_cielo:.1f} litros ({galones_cielo} galones)"),
                                ]
                            if usar_pasta:
                                items_pin.append(("Pasta muro", f"{tarros_pasta} tarros de 25kg"))
                            if usar_cinta:
                                items_pin.append(("Cinta juntas", f"{rollos_cinta} rollos de 75m"))
                            if usar_sellador:
                                items_pin.append(("Sellador", f"{galones_sellador} galones"))
                            registrar_pdf("Terminaciones", "Pintura", items_pin)
                        else:
                            quitar_pdf("Terminaciones", "Pintura")

            # ============================
            # 2. ESTUCO / REVOQUE
            # ============================
            if ver(term, "estuco"):
                with st.expander("2. Estuco / Revoque", expanded=False):

                        estuco_tipo = st.selectbox("Tipo de estuco", list(ESTUCOS.keys()), key="estuco_tipo")
                        est = ESTUCOS[estuco_tipo]

                        # --- Murallas: cada una con su propio largo y alto ---
                        st.markdown("**Murallas a estucar** (agrega cada muro con su medida)")
                        if "secciones_est" not in st.session_state:
                            st.session_state.secciones_est = [{"largo": 0.0, "alto": 0.0}]

                        area_bruta_est = 0.0
                        for i, sec in enumerate(st.session_state.secciones_est):
                            cea, ceb, cec = st.columns([3, 3, 1])
                            with cea:
                                sec["largo"] = st.number_input(
                                    f"Largo muro {i+1} (m)", value=sec["largo"],
                                    min_value=0.0, step=0.1, key=f"est_largo_{i}")
                            with ceb:
                                sec["alto"] = st.number_input(
                                    f"Alto muro {i+1} (m)", value=sec["alto"],
                                    min_value=0.0, step=0.1, key=f"est_alto_{i}")
                            with cec:
                                st.write("")
                                st.write("")
                                if len(st.session_state.secciones_est) > 1 and st.button("🗑️", key=f"del_est_{i}"):
                                    st.session_state.secciones_est.pop(i)
                                    st.rerun()
                            area_bruta_est += sec["largo"] * sec["alto"]

                        if st.button("➕ Agregar muro", key="add_est"):
                            st.session_state.secciones_est.append({"largo": 0.0, "alto": 0.0})
                            st.rerun()

                        # --- Descuento de puertas / ventanas (opcional) ---
                        area_vanos_est = 0.0
                        descontar_vanos_est = st.checkbox(
                            "Descontar puertas / ventanas", key="est_descontar_vanos",
                            help="Actívalo solo si las murallas tienen puertas o ventanas.")
                        if descontar_vanos_est:
                            ev1, ev2 = st.columns(2)
                            with ev1:
                                cant_puertas_est = st.number_input("Cantidad puertas", value=0, step=1, key="cant_puertas_est")
                                ancho_puerta_est = st.number_input("Ancho puerta (m)", value=0.0, key="ancho_puerta_est")
                                alto_puerta_est = st.number_input("Alto puerta (m)", value=0.0, key="alto_puerta_est")
                            with ev2:
                                cant_ventanas_est = st.number_input("Cantidad ventanas", value=0, step=1, key="cant_ventanas_est")
                                ancho_ventana_est = st.number_input("Ancho ventana (m)", value=0.0, key="ancho_ventana_est")
                                alto_ventana_est = st.number_input("Alto ventana (m)", value=0.0, key="alto_ventana_est")
                            area_vanos_est = ((cant_puertas_est * ancho_puerta_est * alto_puerta_est) +
                                              (cant_ventanas_est * ancho_ventana_est * alto_ventana_est))

                        espesor_est = st.selectbox("Espesor de aplicación",
                                                    ["5mm (fino)", "10mm (estándar)", "15mm (máximo por capa)", "20mm (2 capas)", "25mm (2 capas)"],
                                                    key="espesor_est")
                        espesor_val = float(espesor_est.split("mm")[0]) / 1000

                        area_neta_est = max(area_bruta_est - area_vanos_est, 0.0)

                        if espesor_val > 0.015:
                            n_capas = 2
                            st.warning(f"⚠️ Espesor mayor a 15mm: Se requieren {n_capas} capas. Esperar secado entre capas.")
                        else:
                            n_capas = 1

                        if espesor_val >= 0.030:
                            st.error("⚠️ Espesor mayor a 30mm: Requiere malla de refuerzo obligatoriamente")

                        st.write("---")
                        st.info(f"Área neta: {area_neta_est:.2f} m²")
                        st.info(f"Capas necesarias: {n_capas}")

                        if est["tipo"] == "premezclado" or est["tipo"] == "fino":
                            sacos_est = (area_neta_est / est["rendimiento_saco"]) * n_capas
                            st.success(f"Sacos de {est['peso_saco']}kg: {sacos_est:.0f} sacos")
                            st.caption(f"Rendimiento: {est['rendimiento_saco']} m² por saco a {espesor_est}")
                            if area_neta_est > 0:
                                registrar_pdf("Terminaciones", "Estuco / Revoque", [
                                    ("Tipo", estuco_tipo),
                                    ("Área neta", f"{area_neta_est:.2f} m²"),
                                    ("Capas", str(n_capas)),
                                    ("Material", f"{sacos_est:.0f} sacos de {est['peso_saco']}kg"),
                                ])

                        elif est["tipo"] == "tradicional":
                            vol_est = area_neta_est * espesor_val * n_capas
                            cemento_est = vol_est * 300  # kg cemento por m³
                            arena_est = vol_est * 1.20
                            cal_est = vol_est * 100  # kg cal por m³
                            sacos_cemento_est = cemento_est / 25
                            sacos_cal_est = cal_est / 25

                            st.info(f"Volumen mortero: {vol_est:.3f} m³")
                            st.success(f"Cemento: {sacos_cemento_est:.0f} sacos de 25kg")
                            st.success(f"Arena: {arena_est:.2f} m³")
                            st.success(f"Cal: {sacos_cal_est:.0f} sacos de 25kg")
                            st.caption("Dosificación: 1 cemento : 4 arena : 1 cal")
                            if area_neta_est > 0:
                                registrar_pdf("Terminaciones", "Estuco / Revoque", [
                                    ("Tipo", estuco_tipo),
                                    ("Área neta", f"{area_neta_est:.2f} m²"),
                                    ("Capas", str(n_capas)),
                                    ("Cemento", f"{sacos_cemento_est:.0f} sacos de 25kg"),
                                    ("Arena", f"{arena_est:.2f} m³"),
                                    ("Cal", f"{sacos_cal_est:.0f} sacos de 25kg"),
                                ])

            # ============================
            # 3. CIELOS
            # ============================
            if ver(term, "cielos"):
                with st.expander("3. Cielos", expanded=False):

                        cielo_tipo = st.selectbox("Tipo de cielo", list(CIELOS_TIPOS.keys()), key="cielo_tipo")
                        ci = CIELOS_TIPOS[cielo_tipo]

                        st.markdown("**Habitaciones** (agrega cada una con su medida)")
                        _sec_cielo = secciones_input(
                            "cielo", [("largo", "Largo (m)"), ("ancho", "Ancho (m)")], "habitación")
                        desp_ci = st.slider("% Desperdicio", 0, 20, 10, key="desp_ci")

                        # Acumulados por habitación (cada una con su propia medida)
                        area_ci = sum(s["largo"] * s["ancho"] for s in _sec_cielo)
                        perimetro_ci = sum((s["largo"] + s["ancho"]) * 2 for s in _sec_cielo)
                        cant_largueros = sum(s["ancho"] / 0.40 for s in _sec_cielo)
                        cant_conectores = sum((s["ancho"] / 0.40) * (s["largo"] / 1.20) for s in _sec_cielo)

                        st.write("---")
                        st.info(f"Área total cielo: {area_ci:.2f} m²")

                        if ci["tipo"] == "plancha":
                            cant_planchas_ci = area_ci / ci["area_plancha"]
                            cant_planchas_desp_ci = cant_planchas_ci * (1 + desp_ci / 100)
                            st.info(f"Planchas exactas: {cant_planchas_ci:.1f} unidades")
                            st.success(f"Con {desp_ci}% desperdicio: {cant_planchas_desp_ci:.0f} planchas")

                        elif ci["tipo"] == "madera":
                            ancho_tabla_ci = st.selectbox("Ancho tabla machihembrada", ["10cm", "15cm", "20cm"], key="ancho_tabla_ci")
                            ancho_val_ci = float(ancho_tabla_ci.replace("cm", "")) / 100
                            largo_tabla_ci = st.selectbox("Largo tabla", ["3,20m", "4,00m"], key="largo_tabla_ci")
                            largo_val_ci = 3.20 if "3,20" in largo_tabla_ci else 4.00
                            ml_ci = area_ci / ancho_val_ci
                            tablas_ci = ml_ci / largo_val_ci
                            tablas_desp_ci = tablas_ci * (1 + desp_ci / 100)
                            st.info(f"Metros lineales: {ml_ci:.1f} ml")
                            st.success(f"Tablas con {desp_ci}% desperdicio: {tablas_desp_ci:.0f} tablas de {largo_tabla_ci}")

                        # Estructura de cielo (Perfiles AT)
                        st.write("---")
                        st.subheader("🔩 Estructura de Cielo (Perfiles AT)")
                        st.caption("Perfil AT en todo el perímetro + largueros cada 40cm")

                        # Perfil AT perímetro (perímetro ya acumulado por habitación)
                        largo_perfil_at = st.selectbox("Largo perfil AT", ["2,40m", "3,00m"], key="largo_at")
                        largo_val_at = 2.40 if "2,40" in largo_perfil_at else 3.00

                        cant_perfiles_at = perimetro_ci / largo_val_at if largo_val_at > 0 else 0
                        cant_perfiles_at_desp = cant_perfiles_at * 1.10

                        # Largueros Portante 40R cada 40cm (ya acumulados por habitación)
                        largo_larguero = st.selectbox("Largo Portante 40R", ["2,40m", "3,00m"], key="largo_larguero")
                        largo_val_larg = 2.40 if "2,40" in largo_larguero else 3.00
                        cant_largueros_desp = cant_largueros * 1.10

                        # Tornillos
                        tornillos_ci = (cant_perfiles_at_desp + cant_largueros_desp) * 4

                        st.info(f"Perfiles AT perímetro: {cant_perfiles_at_desp:.0f} piezas de {largo_val_at}m")
                        st.info(f"Portante 40R: {cant_largueros_desp:.0f} piezas de {largo_val_larg}m")
                        st.info(f"Conectores TF: {cant_conectores:.0f} unidades")
                        st.success(f"Tornillos: {tornillos_ci:.0f} unidades")
                        st.caption("Largueros cada 40cm | Conectores cada 1,20m")
                        if area_ci > 0:
                            if ci["tipo"] == "plancha":
                                item_rev_ci = ("Planchas (c/desp.)", f"{cant_planchas_desp_ci:.0f} unidades")
                            else:
                                item_rev_ci = ("Tablas (c/desp.)", f"{tablas_desp_ci:.0f} de {largo_tabla_ci}")
                            registrar_pdf("Terminaciones", "Cielos", [
                                ("Tipo", cielo_tipo),
                                ("Área total", f"{area_ci:.2f} m²"),
                                item_rev_ci,
                                ("Perfiles AT", f"{cant_perfiles_at_desp:.0f} de {largo_val_at}m"),
                                ("Portante 40R", f"{cant_largueros_desp:.0f} de {largo_val_larg}m"),
                                ("Tornillos", f"{tornillos_ci:.0f} unidades"),
                            ])

            # ============================
            # 4. ZÓCALOS Y GUARDAPOLVOS
            # ============================
            if ver(term, "zocalos"):
                with st.expander("4. Zócalos y Guardapolvos", expanded=False):

                        zocalo_tipo = st.selectbox("Tipo de zócalo", list(ZOCALOS.keys()), key="zocalo_tipo")
                        zoc = ZOCALOS[zocalo_tipo]

                        zo1, zo2 = st.columns(2)
                        with zo1:
                            largo_zoc = st.number_input("Metros lineales totales (m)", value=0.0, key="largo_zoc")
                        with zo2:
                            cant_vanos_zoc = st.number_input("Cantidad de vanos/puertas", value=0, step=1, key="cant_vanos_zoc")

                        ancho_vano_zoc = st.number_input("Ancho vano promedio (m)", value=0.90, key="ancho_vano_zoc")
                        largo_pieza_zoc = st.selectbox("Largo por pieza", ["2,40m", "3,00m", "3,20m"], key="largo_pieza_zoc")
                        largo_val_zoc = {"2,40m": 2.40, "3,00m": 3.00, "3,20m": 3.20}[largo_pieza_zoc]
                        desp_zoc = st.slider("% Desperdicio", 0, 20, 10, key="desp_zoc")

                        # Descuento vanos
                        ml_vanos_zoc = cant_vanos_zoc * ancho_vano_zoc
                        ml_neto_zoc = largo_zoc - ml_vanos_zoc

                        # Piezas
                        cant_piezas_zoc = ml_neto_zoc / largo_val_zoc
                        cant_piezas_desp_zoc = cant_piezas_zoc * (1 + desp_zoc / 100)

                        # Fijaciones según material
                        if zoc["tipo"] == "MDF":
                            fijacion_zoc = "Pegamento para MDF + Clavo sin cabeza"
                            cant_fijaciones_zoc = round(ml_neto_zoc / 0.40)  # cada 40cm
                            medida_fijacion = "Clavo 1 1/2\""
                        elif zoc["tipo"] == "Madera":
                            fijacion_zoc = "Clavo sin cabeza galvanizado"
                            cant_fijaciones_zoc = round(ml_neto_zoc / 0.40)
                            medida_fijacion = "Clavo 1 1/2\" o 2\""
                        else:  # PVC
                            fijacion_zoc = "Pegamento PVC o clip de fijación"
                            cant_fijaciones_zoc = round(ml_neto_zoc / 0.50)
                            medida_fijacion = "Clip cada 50cm"

                        st.write("---")
                        st.info(f"Metros lineales netos: {ml_neto_zoc:.2f} ml")
                        st.info(f"Piezas exactas: {cant_piezas_zoc:.1f} unidades")
                        st.success(f"Con {desp_zoc}% desperdicio: {cant_piezas_desp_zoc:.0f} piezas de {largo_val_zoc}m")

                        st.write("---")
                        st.subheader("🔩 Fijaciones")
                        st.info(f"Tipo: {fijacion_zoc}")
                        st.info(f"Medida: {medida_fijacion}")
                        st.success(f"Cantidad: {cant_fijaciones_zoc} fijaciones")
                        st.caption("Distancia mínima al borde: 2cm")         
                        if ml_neto_zoc > 0:
                            registrar_pdf("Terminaciones", "Zócalos y Guardapolvos", [
                                ("Tipo", zocalo_tipo),
                                ("Metros lineales netos", f"{ml_neto_zoc:.2f} ml"),
                                ("Piezas (c/desp.)", f"{cant_piezas_desp_zoc:.0f} de {largo_val_zoc}m"),
                                ("Fijaciones", f"{cant_fijaciones_zoc} unidades"),
                            ])

# ============================
# CUBIERTA / TECHUMBRE
# ============================
    if ver_rubro(cubierta) and rubro_permitido("cubierta"):
        with st.expander("Cubierta / Techumbre", expanded=False):

            def area_cubierta(key_prefix):
                """Secciones de faldones (largo x ancho real de la pendiente)."""
                key_list = f"secciones_{key_prefix}"
                if key_list not in st.session_state:
                    st.session_state[key_list] = [{"largo": 0.0, "ancho": 0.0}]
                col_add, col_del = st.columns(2)
                with col_add:
                    if st.button("➕ Agregar faldón", key=f"add_{key_prefix}"):
                        st.session_state[key_list].append({"largo": 0.0, "ancho": 0.0})
                with col_del:
                    if st.button("🗑️ Eliminar último faldón", key=f"del_{key_prefix}"):
                        if len(st.session_state[key_list]) > 1:
                            st.session_state[key_list].pop()
                area_total = 0.0
                for i, sec in enumerate(st.session_state[key_list]):
                    st.markdown(f"**Faldón {i+1}** (medida real de la pendiente)")
                    c1, c2 = st.columns(2)
                    with c1:
                        sec["largo"] = st.number_input("Largo (m)", value=sec["largo"], key=f"{key_prefix}_largo_{i}")
                    with c2:
                        sec["ancho"] = st.number_input("Ancho (m)", value=sec["ancho"], key=f"{key_prefix}_ancho_{i}")
                    area_sec = sec["largo"] * sec["ancho"]
                    st.caption(f"Área faldón {i+1}: {area_sec:.2f} m²")
                    area_total += area_sec
                st.info(f"Área total de cubierta: {area_total:.2f} m²")
                return area_total

            MADERA_COSTANERA = {
                "2x2 Pino 3,20m":  {"largo": 3.20},
                "2x3 Pino 3,20m":  {"largo": 3.20},
                "2x4 Pino 3,20m":  {"largo": 3.20},
            }
            COSTANERA_METALICA = {
                "Perfil Omega 92x50x2,0mm - 6,00m": {"largo": 6.00},
                "Perfil Omega 70x40x1,5mm - 6,00m": {"largo": 6.00},
                "Costanera 80x40x15x2,0mm - 6,00m": {"largo": 6.00},
                "Costanera 100x50x15x2,0mm - 6,00m": {"largo": 6.00},
                "Perfil C 100x50x2,0mm - 6,00m": {"largo": 6.00},
            }
            PERFIL_CERCHA_MET = {
                "Perfil Omega 92x50x2,0mm - 6,00m": {"largo": 6.00},
                "Perfil Omega 70x40x1,5mm - 6,00m": {"largo": 6.00},
                "Perfil C 100x50x2,0mm - 6,00m": {"largo": 6.00},
                "Perfil C 150x50x2,0mm - 6,00m": {"largo": 6.00},
            }
            PLANCHAS_CUBIERTA = {
                "Zinc acanalado 5V (ancho útil 0,80m)": {"tipo": "plancha", "ancho_util": 0.80,
                    "largos": ["2,00m", "2,50m", "3,00m", "3,66m"]},
                "Teja asfáltica (paquete 3,1 m²)": {"tipo": "paquete", "rend_paquete": 3.1},
                "Teja arcilla/cerámica (24 u/m²)": {"tipo": "unidad", "u_m2": 24},
                "Panel Sándwich PV (ancho útil 1,00m)": {"tipo": "plancha", "ancho_util": 1.00,
                    "largos": ["3,00m", "4,00m", "5,00m", "6,00m"]},
            }

    # --- 1. Estructura de Madera ---
            if ver(cubierta, "estructura_madera"):
                with st.expander("1. Estructura de Madera (costaneras + cerchas)", expanded=False):
                    st.caption("Costaneras de madera y cálculo de cerchas")

                    # --- Costaneras ---
                    st.markdown("### 🪵 Costaneras")
                    area_cm = area_cubierta("cub_est_mad")
                    mad_tipo = st.selectbox("Tipo de costanera", list(MADERA_COSTANERA.keys()), key="cub_mad_tipo")
                    sep_cm = st.selectbox("Separación entre costaneras (m)", ["0,40", "0,50", "0,60"], key="cub_mad_sep")
                    sep_cm_val = float(sep_cm.replace(",", "."))
                    desp_cm = st.slider("% Desperdicio costaneras", 0, 20, 10, key="cub_mad_desp")

                    largo_pieza_cm = MADERA_COSTANERA[mad_tipo]["largo"]
                    ml_costaneras = (area_cm / sep_cm_val) if sep_cm_val > 0 else 0
                    ml_costaneras_desp = ml_costaneras * (1 + desp_cm / 100)
                    piezas_cm = math.ceil(ml_costaneras_desp / largo_pieza_cm) if largo_pieza_cm > 0 else 0
                    clavos_cm = math.ceil(piezas_cm * 6)

                    st.info(f"Metros lineales de costanera: {ml_costaneras:.2f} ml")
                    st.success(f"Costaneras (c/desp.): {piezas_cm} piezas de {largo_pieza_cm}m")

                    # --- Cerchas ---
                    st.write("---")
                    st.markdown("### 📐 Cerchas")
                    calc_cerchas = st.checkbox("Calcular cerchas", key="cub_calc_cerchas")

                    # Escuadrías con sección real (m) para volumen
                    ESCUADRIA_CERCHA = {
                        "2x4 Pino 3,20m": {"largo": 3.20, "esp": 0.0508, "ancho": 0.1016},
                        "2x5 Pino 3,20m": {"largo": 3.20, "esp": 0.0508, "ancho": 0.1270},
                        "2x6 Pino 3,20m": {"largo": 3.20, "esp": 0.0508, "ancho": 0.1524},
                    }

                    cerchas_n = 0
                    ml_total_cerchas = 0.0
                    piezas_cercha = 0
                    vol_cerchas = 0.0
                    h_cumbrera = 0.0
                    grados_pend = 0.0
                    par_largo = pend_largo = diag_largo = luz_cercha = 0.0

                    if calc_cerchas:
                        ce1, ce2 = st.columns(2)
                        with ce1:
                            luz_cercha = st.number_input("Luz / ancho a cubrir (m)", value=0.0, key="cub_luz")
                        with ce2:
                            pend_pct = st.number_input("Pendiente (%)", value=30.0, key="cub_pend_pct")

                        ce3, ce4 = st.columns(2)
                        with ce3:
                            largo_techo_cer = st.number_input("Largo del techo (m)", value=0.0, key="cub_largo_techo")
                        with ce4:
                            sep_cerchas = st.number_input("Separación entre cerchas (m)", value=1.00, key="cub_sep_cerchas")

                        # Cantidad de cerchas = (largo / separación) + 1
                        cerchas_n = (math.floor(largo_techo_cer / sep_cerchas) + 1) if sep_cerchas > 0 and largo_techo_cer > 0 else 0

                        esc_cercha = st.selectbox("Escuadría de la madera", list(ESCUADRIA_CERCHA.keys()), key="cub_esc_cercha")
                        esc_d = ESCUADRIA_CERCHA[esc_cercha]
                        largo_esc_cercha = esc_d["largo"]
                        desp_cercha = st.slider("% Desperdicio cerchas", 0, 15, 10, key="cub_desp_cercha")

                        # Geometría (cercha tipo pendolón con tornapuntas)
                        h_cumbrera = (luz_cercha / 2) * (pend_pct / 100)        # altura de cumbrera
                        grados_pend = math.degrees(math.atan(pend_pct / 100))   # pendiente en grados
                        cordon_inf = luz_cercha                                 # 1 pieza horizontal
                        par_largo = math.sqrt((luz_cercha / 2) ** 2 + h_cumbrera ** 2)  # cada par inclinado (×2)
                        pend_largo = h_cumbrera                                 # pendolón / pie derecho central
                        diag_largo = math.sqrt((luz_cercha / 4) ** 2 + (h_cumbrera / 2) ** 2)  # cada diagonal (×2)

                        ml_una_cercha = cordon_inf + (2 * par_largo) + pend_largo + (2 * diag_largo)
                        ml_total_cerchas = ml_una_cercha * cerchas_n
                        ml_total_cerchas_desp = ml_total_cerchas * (1 + desp_cercha / 100)
                        piezas_cercha = math.ceil(ml_total_cerchas_desp / largo_esc_cercha) if largo_esc_cercha > 0 else 0
                        # Volumen en m³ (sección × ml total con desperdicio)
                        vol_cerchas = ml_total_cerchas_desp * esc_d["esp"] * esc_d["ancho"]

                        if luz_cercha > 0 and cerchas_n > 0:
                            st.info(f"Cantidad de cerchas: {cerchas_n}  (largo {largo_techo_cer}m ÷ {sep_cerchas}m + 1)")
                            st.info(f"Altura de cumbrera: {h_cumbrera:.2f} m  |  Pendiente: {grados_pend:.1f}°")
                            st.markdown("**Longitudes por cercha:**")
                            st.text(f"Cordón inferior: {cordon_inf:.2f} m (×1)")
                            st.text(f"Pares inclinados: {par_largo:.2f} m (×2)")
                            st.text(f"Pendolón / pie derecho: {pend_largo:.2f} m (×1)")
                            st.text(f"Diagonales / tornapuntas: {diag_largo:.2f} m (×2)")
                            st.text(f"Metros lineales por cercha: {ml_una_cercha:.2f} ml")
                            st.success(f"Madera total cerchas: {ml_total_cerchas:.2f} ml → {piezas_cercha} piezas de {largo_esc_cercha}m ({esc_cercha.split()[0]})")
                            st.success(f"Volumen de madera: {vol_cerchas:.3f} m³")
                            st.caption("Modelo: cercha tipo pendolón con tornapuntas. Verifique con cálculo estructural.")

                    st.write("---")
                    st.info(f"Clavos aprox.: {clavos_cm} unidades")

                    if area_cm > 0 or (calc_cerchas and ml_total_cerchas > 0):
                        items_cm = [
                            ("Costanera", mad_tipo),
                            ("Área cubierta", f"{area_cm:.2f} m²"),
                            ("Costaneras (c/desp.)", f"{piezas_cm} de {largo_pieza_cm}m"),
                        ]
                        if calc_cerchas and ml_total_cerchas > 0:
                            items_cm.append(("Cerchas", f"{cerchas_n} unidades"))
                            items_cm.append(("Altura cumbrera", f"{h_cumbrera:.2f} m ({grados_pend:.1f}°)"))
                            items_cm.append(("Par inclinado", f"{par_largo:.2f} m (×2 por cercha)"))
                            items_cm.append(("Pendolón", f"{pend_largo:.2f} m (×1 por cercha)"))
                            items_cm.append(("Diagonales", f"{diag_largo:.2f} m (×2 por cercha)"))
                            items_cm.append(("Madera cerchas (c/desp.)", f"{piezas_cercha} de {largo_esc_cercha}m"))
                            items_cm.append(("Volumen madera", f"{vol_cerchas:.3f} m³"))
                        items_cm.append(("Clavos", f"{clavos_cm} unidades"))
                        registrar_pdf("Cubierta / Techumbre", "Estructura de Madera", items_cm)


    # --- 2. Estructura Metálica ---
            if ver(cubierta, "estructura_metalica"):
                with st.expander("2. Estructura Metálica (costaneras + cerchas)", expanded=False):
                    st.caption("Costaneras (Omega / C) y cálculo de cerchas metálicas")

                    # --- Costaneras ---
                    st.markdown("### 🔩 Costaneras")
                    area_met = area_cubierta("cub_est_met")
                    met_tipo = st.selectbox("Tipo de costanera/perfil", list(COSTANERA_METALICA.keys()), key="cub_met_tipo")
                    sep_met = st.selectbox("Separación entre costaneras (m)", ["0,60", "0,80", "1,00", "1,20"], key="cub_met_sep")
                    sep_met_val = float(sep_met.replace(",", "."))
                    desp_met = st.slider("% Desperdicio costaneras", 0, 20, 10, key="cub_met_desp")

                    largo_pieza_met = COSTANERA_METALICA[met_tipo]["largo"]
                    ml_met = (area_met / sep_met_val) if sep_met_val > 0 else 0
                    ml_met_desp = ml_met * (1 + desp_met / 100)
                    piezas_met = math.ceil(ml_met_desp / largo_pieza_met) if largo_pieza_met > 0 else 0
                    tornillos_met = math.ceil(piezas_met * 8)

                    st.info(f"Metros lineales: {ml_met:.2f} ml")
                    st.success(f"Costaneras (c/desp.): {piezas_met} piezas de {largo_pieza_met}m")
                    st.info(f"Tornillos autoperforantes: {tornillos_met} unidades")

                    # --- Cerchas metálicas ---
                    st.write("---")
                    st.markdown("### 📐 Cerchas metálicas")
                    calc_cer_met = st.checkbox("Calcular cerchas metálicas", key="cub_calc_cer_met")

                    cer_met_n = 0
                    ml_total_cer_met = 0.0
                    piezas_cer_met = 0
                    h_cum_met = 0.0
                    grados_met = 0.0
                    par_met = pend_met = diag_met = luz_met = 0.0

                    if calc_cer_met:
                        cm1, cm2 = st.columns(2)
                        with cm1:
                            luz_met = st.number_input("Luz / ancho a cubrir (m)", value=0.0, key="cub_luz_met")
                        with cm2:
                            pend_pct_met = st.number_input("Pendiente (%)", value=30.0, key="cub_pend_met")

                        cm3, cm4 = st.columns(2)
                        with cm3:
                            largo_techo_met = st.number_input("Largo del techo (m)", value=0.0, key="cub_largo_techo_met")
                        with cm4:
                            sep_cer_met = st.number_input("Separación entre cerchas (m)", value=1.20, key="cub_sep_cer_met")

                        cer_met_n = (math.floor(largo_techo_met / sep_cer_met) + 1) if sep_cer_met > 0 and largo_techo_met > 0 else 0

                        perfil_cer = st.selectbox("Perfil de la cercha", list(PERFIL_CERCHA_MET.keys()), key="cub_perfil_cer")
                        largo_perfil_cer = PERFIL_CERCHA_MET[perfil_cer]["largo"]
                        desp_cer_met = st.slider("% Desperdicio cerchas", 0, 15, 10, key="cub_desp_cer_met")

                        # Geometría (igual que madera: pendolón con tornapuntas)
                        h_cum_met = (luz_met / 2) * (pend_pct_met / 100)
                        grados_met = math.degrees(math.atan(pend_pct_met / 100))
                        cordon_inf_met = luz_met
                        par_met = math.sqrt((luz_met / 2) ** 2 + h_cum_met ** 2)
                        pend_met = h_cum_met
                        diag_met = math.sqrt((luz_met / 4) ** 2 + (h_cum_met / 2) ** 2)

                        ml_una_cer_met = cordon_inf_met + (2 * par_met) + pend_met + (2 * diag_met)
                        ml_total_cer_met = ml_una_cer_met * cer_met_n
                        ml_total_cer_met_desp = ml_total_cer_met * (1 + desp_cer_met / 100)
                        piezas_cer_met = math.ceil(ml_total_cer_met_desp / largo_perfil_cer) if largo_perfil_cer > 0 else 0

                        if luz_met > 0 and cer_met_n > 0:
                            st.info(f"Cantidad de cerchas: {cer_met_n}  (largo {largo_techo_met}m ÷ {sep_cer_met}m + 1)")
                            st.info(f"Altura de cumbrera: {h_cum_met:.2f} m  |  Pendiente: {grados_met:.1f}°")
                            st.markdown("**Longitudes por cercha:**")
                            st.text(f"Cordón inferior: {cordon_inf_met:.2f} m (×1)")
                            st.text(f"Pares inclinados: {par_met:.2f} m (×2)")
                            st.text(f"Pendolón / pie derecho: {pend_met:.2f} m (×1)")
                            st.text(f"Diagonales: {diag_met:.2f} m (×2)")
                            st.text(f"Metros lineales por cercha: {ml_una_cer_met:.2f} ml")
                            st.success(f"Perfil total cerchas: {ml_total_cer_met:.2f} ml → {piezas_cer_met} barras de {largo_perfil_cer}m")
                            st.caption("Modelo: cercha tipo pendolón con tornapuntas. Verifique con cálculo estructural.")

                    if area_met > 0 or (calc_cer_met and ml_total_cer_met > 0):
                        items_met = [
                            ("Costanera/perfil", met_tipo),
                            ("Área cubierta", f"{area_met:.2f} m²"),
                            ("Costaneras (c/desp.)", f"{piezas_met} de {largo_pieza_met}m"),
                            ("Tornillos", f"{tornillos_met} unidades"),
                        ]
                        if calc_cer_met and ml_total_cer_met > 0:
                            items_met.append(("Cerchas", f"{cer_met_n} unidades"))
                            items_met.append(("Perfil cercha", perfil_cer))
                            items_met.append(("Altura cumbrera", f"{h_cum_met:.2f} m ({grados_met:.1f}°)"))
                            items_met.append(("Par inclinado", f"{par_met:.2f} m (×2 por cercha)"))
                            items_met.append(("Pendolón", f"{pend_met:.2f} m (×1 por cercha)"))
                            items_met.append(("Diagonales", f"{diag_met:.2f} m (×2 por cercha)"))
                            items_met.append(("Perfil cerchas (c/desp.)", f"{piezas_cer_met} barras de {largo_perfil_cer}m"))
                        registrar_pdf("Cubierta / Techumbre", "Estructura Metálica", items_met)

    # --- 3. Cubierta (planchas) ---
            if ver(cubierta, "planchas"):
                with st.expander("3. Cubierta (planchas / tejas)", expanded=False):
                    area_pl = area_cubierta("cub_planchas")
                    plancha_tipo = st.selectbox("Tipo de cubierta", list(PLANCHAS_CUBIERTA.keys()), key="cub_plancha_tipo")
                    pl = PLANCHAS_CUBIERTA[plancha_tipo]
                    desp_pl = st.slider("% Desperdicio (traslapes y cortes)", 0, 25, 15, key="cub_plancha_desp")
                    area_pl_desp = area_pl * (1 + desp_pl / 100)

                    st.write("---")
                    st.info(f"Área cubierta: {area_pl:.2f} m²")
                    st.success(f"Área con {desp_pl}% (traslapes/cortes): {area_pl_desp:.2f} m²")

                    items_pl = [("Tipo", plancha_tipo), ("Área cubierta", f"{area_pl:.2f} m²")]

                    if pl["tipo"] == "plancha":
                        largo_pl = st.selectbox("Largo de plancha", pl["largos"], key="cub_plancha_largo")
                        largo_pl_val = float(largo_pl.replace("m", "").replace(",", "."))
                        area_util_plancha = pl["ancho_util"] * largo_pl_val
                        cant_planchas = math.ceil(area_pl_desp / area_util_plancha) if area_util_plancha > 0 else 0
                        st.success(f"Planchas {largo_pl}: {cant_planchas} unidades")
                        st.caption(f"Ancho útil {pl['ancho_util']}m × {largo_pl} = {area_util_plancha:.2f} m²/plancha")
                        items_pl.append(("Planchas", f"{cant_planchas} de {largo_pl}"))
                        # Tornillos techo (caballete): aprox 6 por m²
                        tornillos_techo = math.ceil(area_pl * 6)
                        st.info(f"Tornillos techo (autoperf. con golilla): {tornillos_techo} unidades")
                        items_pl.append(("Tornillos techo", f"{tornillos_techo} unidades"))

                    elif pl["tipo"] == "paquete":
                        paquetes = math.ceil(area_pl_desp / pl["rend_paquete"]) if pl["rend_paquete"] > 0 else 0
                        st.success(f"Paquetes de teja asfáltica: {paquetes} paquetes")
                        st.caption(f"Rendimiento: {pl['rend_paquete']} m² por paquete")
                        clavos_teja = math.ceil(area_pl * 8)
                        st.info(f"Clavos para teja: {clavos_teja} unidades")
                        items_pl.append(("Paquetes (c/desp.)", f"{paquetes} paquetes"))
                        items_pl.append(("Clavos", f"{clavos_teja} unidades"))

                    elif pl["tipo"] == "unidad":
                        tejas = math.ceil(area_pl_desp * pl["u_m2"])
                        st.success(f"Tejas: {tejas} unidades")
                        st.caption(f"Rendimiento: {pl['u_m2']} tejas por m²")
                        items_pl.append(("Tejas (c/desp.)", f"{tejas} unidades"))

                    if area_pl > 0:
                        registrar_pdf("Cubierta / Techumbre", "Cubierta (planchas)", items_pl)

    # --- 4. Aislación y Fieltro ---
            if ver(cubierta, "aislacion"):
                with st.expander("4. Aislación y Fieltro", expanded=False):
                    area_ais = area_cubierta("cub_aislacion")
                    st.write("**Fieltro asfáltico (barrera)**")
                    usar_fieltro = st.checkbox("Incluir fieltro asfáltico", value=True, key="cub_fieltro")
                    st.write("**Aislación térmica**")
                    usar_aislante = st.checkbox("Incluir aislación térmica (lana/poliestireno)", value=True, key="cub_aislante")
                    desp_ais = st.slider("% Desperdicio", 0, 20, 10, key="cub_ais_desp")

                    st.write("---")
                    items_ais = [("Área cubierta", f"{area_ais:.2f} m²")]

                    if usar_fieltro:
                        # Rollo de fieltro 15 lb cubre ~40 m² (1m x 40m, con traslape ~36 m² útiles)
                        m2_fieltro = area_ais * (1 + desp_ais / 100)
                        rollos_fieltro = math.ceil(m2_fieltro / 36)
                        st.info(f"Fieltro asfáltico: {m2_fieltro:.2f} m²")
                        st.success(f"Rollos de fieltro (≈36 m² útiles): {rollos_fieltro} rollos")
                        items_ais.append(("Fieltro asfáltico", f"{rollos_fieltro} rollos (≈40 m²/rollo)"))

                    if usar_aislante:
                        m2_aislante = area_ais * (1 + desp_ais / 100)
                        st.info(f"Aislación térmica: {m2_aislante:.2f} m²")
                        items_ais.append(("Aislación térmica", f"{m2_aislante:.2f} m²"))

                    if area_ais > 0 and (usar_fieltro or usar_aislante):
                        registrar_pdf("Cubierta / Techumbre", "Aislación y Fieltro", items_ais)

# ============================
# PRESUPUESTO (Premium)
# ============================
if option == "Presupuesto":
    st.title("💰 Presupuesto")

    if not st.session_state.get("usuario"):
        st.info("Inicia sesión para usar el presupuesto.")
        st.caption("El presupuesto es una función del Plan Premium.")
        st.stop()

    if not puede_presupuesto():
        aviso_premium("El presupuesto con precios")
        st.write("---")
        st.markdown(
            "Con el **Plan Pro** podrás:\n"
            "- Ponerle **precio a cada material** y obtener el costo total\n"
            "- Sumar **mano de obra** y **margen de ganancia**\n"
            "- Ver el total **con IVA y sin IVA**\n"
            "- Llevar el presupuesto a un **PDF profesional**"
        )
        st.stop()

    # --- Usuario premium: presupuesto real ---
    usuario = st.session_state.get("usuario")
    # Leer materiales del acumulador persistente (no se borra al cambiar de sección)
    persistente = st.session_state.get("materiales_persistente", {})
    pdf_extra = list(persistente.values())

    if not pdf_extra:
        st.info("Primero realiza una cubicación en la sección **Cubicacion**. "
                "Aquí aparecerán los materiales para ponerles precio.")
        st.stop()

    st.caption("Elige las partidas y materiales que quieres incluir, ponles precio y arma tu presupuesto.")

    # Densidades estándar para convertir kg → m³ (solo en el presupuesto)
    DENSIDADES_M3 = {"gravilla": 1500, "arena": 1600}

    import math as _math
    CLAVOS_POR_KILO = 75       # estimado para clavos de 3"
    TORNILLOS_POR_CAJA = 1000  # caja estándar Metalcon

    # Etiquetas que NO son materiales presupuestables (datos informativos)
    NO_MATERIALES = [
        "área", "area", "superficie", "volumen", "dosificación", "dosificacion",
        "espesor", "medida", "tipo madera", "manos", "capas", "traslape",
        "dirección x", "dirección y", "direccion x", "direccion y",
        "altura cumbrera", "pendiente", "par inclinado", "pendolón", "pendolon",
        "metros lineales netos", "acero total",
    ]

    # Construir lista de materiales presupuestables desde pdf_extra
    materiales_disponibles = []
    for bloque in pdf_extra:
        rubro = bloque.get("rubro", "")
        partida = bloque.get("partida", "")
        for etiqueta, valor in bloque.get("items", []):
            etiqueta_baja = etiqueta.lower()
            # Saltar datos informativos que no son materiales
            if any(palabra in etiqueta_baja for palabra in NO_MATERIALES):
                continue
            cant, uni = parsear_cantidad(valor)
            if cant is not None and cant > 0:
                # Arena y gravilla: kg → m³
                if uni and "kg" in uni.lower():
                    densidad = None
                    if "gravilla" in etiqueta_baja:
                        densidad = DENSIDADES_M3["gravilla"]
                    elif "arena" in etiqueta_baja:
                        densidad = DENSIDADES_M3["arena"]
                    if densidad:
                        cant = cant / densidad
                        uni = "m³"
                # Tornillos: unidades → cajas de 1.000 (redondeo hacia arriba)
                if "tornillo" in etiqueta_baja and uni and "unidad" in uni.lower():
                    cant = _math.ceil(cant / TORNILLOS_POR_CAJA)
                    uni = "caja(s) de 1.000"
                # Clavos: unidades → kilos estimados
                elif "clavo" in etiqueta_baja and uni and "unidad" in uni.lower():
                    cant = cant / CLAVOS_POR_KILO
                    uni = "kg"
                materiales_disponibles.append({
                    "rubro": rubro, "partida": partida,
                    "material": etiqueta, "cantidad": cant, "unidad": uni,
                })

    if not materiales_disponibles:
        st.warning("No se encontraron materiales con cantidad en la cubicación actual.")
        st.stop()

    st.write("---")
    st.subheader("1. Materiales")
    st.caption("Marca los que quieras incluir y ponles precio unitario (referencial, ajústalo a tu proveedor).")
    st.info("💡 Los precios que aparecen son **valores referenciales netos (sin IVA)** de ejemplo. "
            "Ajústalos a los de tu proveedor y tu región. El IVA se agrega al final del presupuesto.")

    # Cargar precios personales guardados del usuario (una vez por sesión)
    if "precios_usuario_cargados" not in st.session_state:
        st.session_state["precios_usuario_dict"] = cargar_precios_usuario(usuario["id"])
        st.session_state["precios_usuario_cargados"] = True
    precios_guardados = st.session_state.get("precios_usuario_dict", {})

    if precios_guardados:
        st.success(f"✓ Tienes {len(precios_guardados)} precios personales guardados. Se aplican automáticamente.")

    with st.popover("⚙️ Opciones"):
        if st.button("🗑️ Reiniciar materiales", help="Borra los materiales acumulados y vuelve a tomarlos de la cubicación actual"):
            st.session_state["materiales_persistente"] = {}
            st.rerun()
        if st.button("🔄 Recargar mis precios guardados"):
            st.session_state["precios_usuario_dict"] = cargar_precios_usuario(usuario["id"])
            st.rerun()

    # Agrupar por rubro
    rubros_unicos = []
    for m in materiales_disponibles:
        if m["rubro"] not in rubros_unicos:
            rubros_unicos.append(m["rubro"])

    subtotal_materiales = 0
    items_presupuesto = []

    for rubro in rubros_unicos:
        mats_rubro = [m for m in materiales_disponibles if m["rubro"] == rubro]
        with st.expander(f"📦 {rubro}", expanded=True):
            # Agrupar por partida dentro del rubro
            partidas_rubro = []
            for m in mats_rubro:
                if m["partida"] not in partidas_rubro:
                    partidas_rubro.append(m["partida"])

            for partida in partidas_rubro:
                mats_part = [m for m in mats_rubro if m["partida"] == partida]
                pc1, pc2 = st.columns([4, 1])
                with pc1:
                    st.markdown(f"**{partida}**")
                with pc2:
                    if st.button("✕ Quitar", key=f"quitar_{rubro}_{partida}".replace(" ", "_"),
                                 help=f"Quitar {partida} del presupuesto"):
                        clave = f"{rubro}||{partida}"
                        st.session_state.get("materiales_persistente", {}).pop(clave, None)
                        st.rerun()

                for idx, m in enumerate(mats_part):
                    key_base = f"pres_{rubro}_{m['partida']}_{m['material']}_{idx}".replace(" ", "_")
                    c_incluir, c_info = st.columns([1, 5])
                    with c_incluir:
                        incluir = st.checkbox("Incluir", value=True, key=f"inc_{key_base}", label_visibility="collapsed")
                    with c_info:
                        st.markdown(f"{m['material']}")
                        if m['unidad'] in ("m³", "kg"):
                            cant_fmt = f"{m['cantidad']:.2f}"
                        else:
                            cant_fmt = f"{m['cantidad']:.0f}"
                        st.caption(f"{cant_fmt} {m['unidad']}")

                    # Menú de opciones (marca/tipo) si existen para este material
                    opciones = opciones_para(m["material"])
                    precio_sugerido = precio_referencial(m["material"])
                    sufijo_precio = ""
                    if opciones:
                        etiquetas = [f"{nombre} — {fmt_clp(p)}" for nombre, p in opciones] + ["Otro (precio libre)"]
                        sel = st.selectbox("Tipo / marca", etiquetas, key=f"opt_{key_base}")
                        if sel != "Otro (precio libre)":
                            i_sel = etiquetas.index(sel)
                            precio_sugerido = opciones[i_sel][1]
                            sufijo_precio = f"_{i_sel}"

                    if m["material"] in precios_guardados:
                        precio_sugerido = precios_guardados[m["material"]]

                    precio = st.number_input(
                        f"Precio unitario ({m['unidad']}) — neto sin IVA",
                        min_value=0, value=precio_sugerido, step=100,
                        key=f"precio_{key_base}{sufijo_precio}",
                    )

                    if incluir and precio > 0:
                        sub = m["cantidad"] * precio
                        subtotal_materiales += sub
                        items_presupuesto.append({
                            "rubro": rubro, "partida": partida,
                            "material": m["material"], "cantidad": m["cantidad"],
                            "unidad": m["unidad"], "precio": precio, "subtotal": sub,
                        })
                    st.session_state.setdefault("_precios_actuales", {})
                    st.session_state["_precios_actuales"][m["material"]] = precio
                st.divider()

    st.success(f"**Subtotal materiales: {fmt_clp(subtotal_materiales)}**")

    # Guardar precios como lista personal
    with st.expander("💾 Guardar mis precios"):
        st.caption("Guarda los precios actuales en tu cuenta. Se aplicarán automáticamente en tus próximos presupuestos.")
        if st.button("💾 Guardar mis precios", type="primary"):
            precios_a_guardar = st.session_state.get("_precios_actuales", {})
            try:
                guardar_precios_usuario(usuario["id"], precios_a_guardar)
                # Actualizar el dict en memoria
                st.session_state["precios_usuario_dict"] = cargar_precios_usuario(usuario["id"])
                st.success("¡Precios guardados! Se aplicarán en tus próximos presupuestos.")
            except Exception:
                st.error("No se pudieron guardar los precios. Intenta de nuevo.")

    # --- Mano de obra ---
    st.write("---")
    st.subheader("2. Mano de obra")
    metodo_mo = st.radio(
        "¿Cómo quieres calcular la mano de obra?",
        ["Monto total", "% sobre materiales", "Por medida (m²/m³/ml/global)"],
        key="metodo_mano_obra",
    )

    costo_mano_obra = 0
    if metodo_mo == "Monto total":
        costo_mano_obra = st.number_input("Monto de mano de obra ($)", min_value=0, value=0, step=10000, key="mo_monto")
    elif metodo_mo == "% sobre materiales":
        pct_mo = st.number_input("Porcentaje sobre materiales (%)", min_value=0.0, value=30.0, step=5.0, key="mo_pct")
        costo_mano_obra = subtotal_materiales * (pct_mo / 100)
        st.caption(f"{pct_mo:.0f}% de {fmt_clp(subtotal_materiales)} = {fmt_clp(costo_mano_obra)}")
    else:  # Por medida
        cm0, cm1, cm2 = st.columns([1.2, 1, 1.3])
        with cm0:
            unidad_mo = st.selectbox("Unidad", ["m²", "m³", "ml", "global"], key="mo_unidad")
        if unidad_mo == "global":
            with cm1:
                st.caption("Monto global")
            with cm2:
                costo_mano_obra = st.number_input("Monto global ($)", min_value=0, value=0, step=10000, key="mo_global")
            st.caption(f"Mano de obra global: {fmt_clp(costo_mano_obra)}")
        else:
            with cm1:
                cantidad_mo = st.number_input(f"Cantidad ({unidad_mo})", min_value=0.0, value=0.0, step=1.0, key="mo_cantidad")
            with cm2:
                valor_unidad_mo = st.number_input(f"Valor por {unidad_mo} ($)", min_value=0, value=0, step=1000, key="mo_valor_unidad")
            costo_mano_obra = cantidad_mo * valor_unidad_mo
            st.caption(f"{cantidad_mo:.1f} {unidad_mo} × {fmt_clp(valor_unidad_mo)} = {fmt_clp(costo_mano_obra)}")

    st.success(f"**Mano de obra: {fmt_clp(costo_mano_obra)}**")

    # --- Margen de ganancia ---
    st.write("---")
    st.subheader("3. Margen de ganancia")
    pct_margen = st.number_input("Margen de ganancia (%)", min_value=0.0, value=15.0, step=5.0, key="margen_pct")
    base_margen = subtotal_materiales + costo_mano_obra
    monto_margen = base_margen * (pct_margen / 100)
    st.caption(f"{pct_margen:.0f}% sobre (materiales + mano de obra) = {fmt_clp(monto_margen)}")

    # --- Totales con IVA ---
    neto = subtotal_materiales + costo_mano_obra + monto_margen
    iva = neto * 0.19
    total_con_iva = neto + iva

    st.write("---")
    st.subheader("Resumen del presupuesto")

    res1, res2 = st.columns(2)
    with res1:
        st.markdown(f"Subtotal materiales: **{fmt_clp(subtotal_materiales)}**")
        st.markdown(f"Mano de obra: **{fmt_clp(costo_mano_obra)}**")
        st.markdown(f"Margen ({pct_margen:.0f}%): **{fmt_clp(monto_margen)}**")
    with res2:
        st.markdown(f"Neto (sin IVA): **{fmt_clp(neto)}**")
        st.markdown(f"IVA (19%): **{fmt_clp(iva)}**")

    st.markdown(
        f"<div style='background:#1a3a1a;border-radius:8px;padding:14px 18px;margin-top:10px;"
        f"display:flex;justify-content:space-between;align-items:center;'>"
        f"<span style='font-size:16px;color:#7CFC7C;'>Total con IVA</span>"
        f"<span style='font-size:22px;font-weight:600;color:#7CFC7C;'>{fmt_clp(total_con_iva)}</span></div>",
        unsafe_allow_html=True,
    )

    st.caption("Los precios son referenciales y definidos por ti. Ajusta a tu proveedor y a tu realidad de obra.")

    # Guardar el presupuesto en session_state para el PDF (paso siguiente)
    st.session_state["presupuesto_actual"] = {
        "items": items_presupuesto,
        "subtotal_materiales": subtotal_materiales,
        "mano_obra": costo_mano_obra,
        "margen_pct": pct_margen,
        "margen": monto_margen,
        "neto": neto,
        "iva": iva,
        "total": total_con_iva,
    }

    # --- Exportar presupuesto a PDF ---
    st.write("---")
    st.subheader("📄 Exportar presupuesto")
    pcol1, pcol2 = st.columns(2)
    with pcol1:
        nombre_pres = st.text_input(
            "Nombre del proyecto",
            value=st.session_state.get("proyecto", {}).get("nombre", ""),
            placeholder="Ej: Casa Don Pedro - Angol",
            key="pres_nombre_proyecto",
        )
    with pcol2:
        cliente_pres = st.text_input("Cliente (opcional)", placeholder="Nombre del cliente", key="pres_cliente")

    if subtotal_materiales > 0 or costo_mano_obra > 0:
        if st.button("📄 Generar PDF del presupuesto", type="primary", use_container_width=True):
            try:
                pdf_buffer = generar_pdf_presupuesto(
                    nombre_pres, st.session_state["presupuesto_actual"], cliente_pres or None,
                    datos_usuario=datos_usuario_pdf()
                )
                nombre_archivo = f"Presupuesto_{nombre_pres or 'obracubic'}.pdf".replace(" ", "_")
                st.download_button(
                    label="⬇️ Descargar presupuesto PDF",
                    data=pdf_buffer,
                    file_name=nombre_archivo,
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception:
                import traceback
                print("ERROR generando PDF de presupuesto:", traceback.format_exc())
                st.error("No se pudo generar el PDF. Intenta de nuevo.")

        # --- Exportar a Excel (.xlsx) — solo Plan Pro Élite ---
        if puede_exportar_excel():
            if st.button("📊 Exportar a Excel (.xlsx)", use_container_width=True):
                try:
                    xlsx_buffer = generar_excel_presupuesto(
                        nombre_pres, st.session_state["presupuesto_actual"], cliente_pres or None,
                        datos_usuario=datos_usuario_pdf()
                    )
                    nombre_xlsx = f"Presupuesto_{nombre_pres or 'obracubic'}.xlsx".replace(" ", "_")
                    st.download_button(
                        label="⬇️ Descargar presupuesto Excel",
                        data=xlsx_buffer,
                        file_name=nombre_xlsx,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception:
                    import traceback
                    print("ERROR generando Excel de presupuesto:", traceback.format_exc())
                    st.error("No se pudo generar el Excel. Intenta de nuevo.")
        else:
            st.caption("📊 La exportación a Excel está disponible en el Plan Pro Élite.")
    else:
        st.caption("Agrega al menos un material con precio o mano de obra para generar el PDF.")


# ============================
# BITÁCORA DE OBRA (Pro Élite)
# ============================
if option == "Bitácora":
    st.header("📓 Bitácora de Obra")

    if not puede_bitacora():
        st.warning("La Bitácora de obra está disponible en el **Plan Pro Élite**.")
    else:
        usuario = st.session_state.get("usuario") or {}
        usuario_id = usuario.get("id")

        # Nombre del proyecto al que se asocia la bitácora
        proyecto_bit = st.text_input(
            "Proyecto",
            value=st.session_state.get("proyecto", {}).get("nombre", ""),
            placeholder="Ej: Casa Don Pedro - Angol",
            key="bitacora_proyecto",
        )

        if not proyecto_bit:
            st.info("Escribe el nombre del proyecto para ver y agregar entradas a su bitácora.")
        else:
            st.caption("Registra el avance de la obra: anota lo realizado y adjunta una foto.")

            # --- Nueva entrada ---
            with st.expander("➕ Nueva entrada", expanded=True):
                nota_bit = st.text_area(
                    "Nota / avance del día",
                    placeholder="Ej: Se hormigonó el radier del living. Faltan terminaciones de borde.",
                    key="bitacora_nota",
                )
                foto_bit = st.file_uploader(
                    "Foto (opcional)", type=["jpg", "jpeg", "png"], key="bitacora_foto",
                    help="Adjunta una foto del avance. Se reduce automáticamente para ahorrar espacio.",
                )
                if st.button("💾 Guardar entrada", type="primary"):
                    if not nota_bit.strip() and not foto_bit:
                        st.error("Escribe una nota o adjunta una foto.")
                    else:
                        foto_b64 = None
                        try:
                            if foto_bit is not None:
                                import base64
                                from PIL import Image
                                import io as _io
                                img = Image.open(foto_bit)
                                img = img.convert("RGB")
                                # Reducir a un máximo de 1024px de ancho para ahorrar espacio
                                if img.width > 1024:
                                    nuevo_alto = int(img.height * 1024 / img.width)
                                    img = img.resize((1024, nuevo_alto))
                                buf_img = _io.BytesIO()
                                img.save(buf_img, format="JPEG", quality=70)
                                foto_b64 = base64.b64encode(buf_img.getvalue()).decode("utf-8")
                            guardar_bitacora(usuario_id, proyecto_bit, nota_bit.strip(), foto_b64)
                            st.success("Entrada guardada en la bitácora.")
                            st.rerun()
                        except Exception:
                            import traceback
                            print("ERROR guardando bitácora:", traceback.format_exc())
                            st.error("No se pudo guardar la entrada. Intenta de nuevo.")

            # --- Historial de entradas ---
            st.write("---")
            st.subheader(f"Historial — {proyecto_bit}")
            entradas = listar_bitacora(usuario_id, proyecto_bit)
            if not entradas:
                st.caption("Aún no hay entradas para este proyecto.")
            for e in entradas:
                fecha_txt = (e.get("creado_en") or "")[:16].replace("T", " ")
                with st.container(border=True):
                    cols = st.columns([5, 1])
                    with cols[0]:
                        st.markdown(f"**🗓️ {fecha_txt}**")
                    with cols[1]:
                        if st.button("🗑️", key=f"del_bit_{e['id']}", help="Eliminar entrada"):
                            try:
                                eliminar_bitacora(e["id"])
                                st.rerun()
                            except Exception:
                                st.error("No se pudo eliminar.")
                    if e.get("nota"):
                        st.write(e["nota"])
                    if e.get("foto"):
                        try:
                            import base64
                            st.image(base64.b64decode(e["foto"]), use_container_width=True)
                        except Exception:
                            st.caption("(No se pudo mostrar la foto)")


# ============================
# APU — ANÁLISIS DE PRECIOS UNITARIOS (Pro Élite)
# ============================
if option == "APU":
    st.header("🔬 Análisis de Precios Unitarios (APU)")

    if not puede_apu():
        st.warning("El Análisis de Precios Unitarios está disponible en el **Plan Pro Élite**.")
    else:
        st.caption(
            "Arma el precio de una partida automáticamente desde tu cubicación: "
            "se cargan solos los materiales y precios; tú agregas la mano de obra y los recargos."
        )
        import pandas as pd

        st.session_state.setdefault("apu_unidad", "m³")
        st.session_state.setdefault("apu_cantidad", 0.0)
        st.session_state.setdefault("apu_load_n", 0)

        # --- Generar automático desde una partida cubicada ---
        persistente_apu = st.session_state.get("materiales_persistente", {})
        if persistente_apu:
            with st.expander("⚡ Generar automático desde una partida cubicada", expanded=True):
                claves = list(persistente_apu.keys())
                etiquetas = [f"{persistente_apu[k].get('rubro','')} — {persistente_apu[k].get('partida','')}"
                             for k in claves]
                sel_idx = st.selectbox("Partida cubicada", range(len(claves)),
                                       format_func=lambda i: etiquetas[i], key="apu_sel_partida")
                if st.button("⚡ Cargar materiales de esta partida", type="primary"):
                    bloque = persistente_apu[claves[sel_idx]]
                    mats = materiales_de_partida_apu(bloque)
                    medida, uni_med = medida_de_partida_apu(bloque)
                    nombre_partida = bloque.get("partida", "")
                    st.session_state["apu_mat_data"] = [
                        {"Descripción": mat, "Unidad": uni, "Cantidad": cant,
                         "Precio unitario": precio_referencial(mat)}
                        for (mat, cant, uni) in mats
                    ] or [{"Descripción": "", "Unidad": "", "Cantidad": 0.0, "Precio unitario": 0}]
                    # Mano de obra referencial automática (según la partida y su medida)
                    mo_rate = mano_obra_referencial(nombre_partida)
                    if mo_rate > 0:
                        st.session_state["apu_mo_data"] = [
                            {"Descripción": f"Mano de obra ({nombre_partida})",
                             "Cantidad": medida if medida > 0 else 0.0,
                             "Precio unitario": mo_rate}
                        ]
                    else:
                        st.session_state["apu_mo_data"] = [
                            {"Descripción": "", "Cantidad": 0.0, "Precio unitario": 0}]
                    st.session_state["apu_partida"] = nombre_partida
                    if medida > 0:
                        st.session_state["apu_cantidad"] = medida
                        st.session_state["apu_unidad"] = uni_med
                    st.session_state["apu_load_n"] += 1  # refrescar las tablas
                    mo_msg = " y mano de obra referencial" if mo_rate > 0 else ""
                    st.success(f"Cargados {len(mats)} materiales{mo_msg}. Revisa y ajusta los valores.")
                    st.rerun()
        else:
            st.info("💡 Cubica una partida en la sección **Cubicacion** y luego vuelve aquí para generar su APU automático. También puedes llenarlo a mano abajo.")

        # --- Datos de la partida ---
        ac1, ac2, ac3 = st.columns([3, 1, 2])
        with ac1:
            apu_partida = st.text_input("Partida", placeholder="Ej: Hormigón G-25 en radier", key="apu_partida")
        with ac2:
            apu_unidad = st.text_input("Unidad", key="apu_unidad")
        with ac3:
            apu_cantidad = st.number_input("Medida de la partida (para el precio unitario)",
                                           min_value=0.0, step=1.0, key="apu_cantidad",
                                           help="La cantidad total (m³, m², ml). El precio unitario = total ÷ esta medida.")

        n = st.session_state["apu_load_n"]

        # --- Materiales (cantidades totales de la partida) ---
        st.subheader("1. Materiales")
        df_mat = pd.DataFrame(st.session_state.get("apu_mat_data",
                              [{"Descripción": "", "Unidad": "", "Cantidad": 0.0, "Precio unitario": 0}]))
        df_mat = st.data_editor(
            df_mat, num_rows="dynamic", use_container_width=True, key=f"apu_mat_{n}",
            column_config={
                "Cantidad": st.column_config.NumberColumn(format="%.2f", help="Cantidad total de la partida"),
                "Precio unitario": st.column_config.NumberColumn(format="%d", help="Precio neto (sin IVA)"),
            },
        )
        st.session_state["apu_mat_data"] = df_mat.to_dict("records")
        sub_mat = float((df_mat["Cantidad"].fillna(0) * df_mat["Precio unitario"].fillna(0)).sum())
        st.info(f"Subtotal materiales: {fmt_clp(sub_mat)}")

        # --- Mano de obra (total de la partida) ---
        st.subheader("2. Mano de obra")
        df_mo = pd.DataFrame(st.session_state.get("apu_mo_data",
                             [{"Descripción": "", "Cantidad": 0.0, "Precio unitario": 0}]))
        df_mo = st.data_editor(
            df_mo, num_rows="dynamic", use_container_width=True, key=f"apu_mo_{n}",
            column_config={
                "Cantidad": st.column_config.NumberColumn(format="%.2f", help="HH o jornadas totales"),
                "Precio unitario": st.column_config.NumberColumn(format="%d"),
            },
        )
        st.session_state["apu_mo_data"] = df_mo.to_dict("records")
        sub_mo = float((df_mo["Cantidad"].fillna(0) * df_mo["Precio unitario"].fillna(0)).sum())
        st.info(f"Subtotal mano de obra: {fmt_clp(sub_mo)}")

        # --- Porcentajes ---
        st.subheader("3. Recargos")
        rc1, rc2, rc3 = st.columns(3)
        with rc1:
            pct_leyes = st.number_input("Leyes sociales (% sobre M.O.)", min_value=0.0, value=0.0, step=1.0, key="apu_leyes")
        with rc2:
            pct_herr = st.number_input("Herramientas (% sobre M.O.)", min_value=0.0, value=5.0, step=1.0, key="apu_herr")
        with rc3:
            pct_ggu = st.number_input("Gastos grales. + utilidad (%)", min_value=0.0, value=25.0, step=1.0, key="apu_ggu")

        monto_leyes = sub_mo * (pct_leyes / 100)
        monto_herr = sub_mo * (pct_herr / 100)
        costo_directo = sub_mat + sub_mo + monto_leyes + monto_herr
        monto_ggu = costo_directo * (pct_ggu / 100)
        total_partida = costo_directo + monto_ggu
        precio_unitario = (total_partida / apu_cantidad) if apu_cantidad > 0 else 0.0

        # --- Resultado ---
        st.write("---")
        st.subheader("Resultado")
        with st.container(border=True):
            st.markdown(f"Materiales: **{fmt_clp(sub_mat)}**")
            st.markdown(f"Mano de obra: **{fmt_clp(sub_mo)}**")
            st.markdown(f"Leyes sociales ({pct_leyes:.0f}%): **{fmt_clp(monto_leyes)}**")
            st.markdown(f"Herramientas ({pct_herr:.0f}%): **{fmt_clp(monto_herr)}**")
            st.markdown(f"Costo directo: **{fmt_clp(costo_directo)}**")
            st.markdown(f"Gastos grales. + utilidad ({pct_ggu:.0f}%): **{fmt_clp(monto_ggu)}**")
            st.markdown(
                f"<div style='margin-top:8px;padding:10px;background:#1E1E1E;border-radius:8px;'>"
                f"<span style='color:#fff;'>Total de la partida: </span>"
                f"<span style='font-size:22px;font-weight:700;color:#1E8E3E;'>{fmt_clp(total_partida)}</span></div>",
                unsafe_allow_html=True,
            )
            if apu_cantidad > 0:
                st.markdown(
                    f"<div style='margin-top:6px;padding:10px;background:#1E1E1E;border-radius:8px;'>"
                    f"<span style='color:#fff;'>Precio unitario ({apu_unidad or 'unidad'}): </span>"
                    f"<span style='font-size:22px;font-weight:700;color:#FF6B00;'>{fmt_clp(precio_unitario)}</span></div>",
                    unsafe_allow_html=True,
                )
            else:
                st.caption("💡 Ingresa la medida de la partida arriba para ver el precio unitario.")
        st.caption("Valores netos (sin IVA). Estimación referencial.")

        # --- Exportar APU a PDF ---
        if st.button("📄 Exportar APU a PDF", use_container_width=True):
            try:
                datos_apu = {
                    "partida": apu_partida, "unidad": apu_unidad, "cantidad": apu_cantidad,
                    "materiales": st.session_state.get("apu_mat_data", []),
                    "mano_obra": st.session_state.get("apu_mo_data", []),
                    "sub_mat": sub_mat, "sub_mo": sub_mo,
                    "leyes": monto_leyes, "herr": monto_herr,
                    "costo_directo": costo_directo, "ggu": monto_ggu,
                    "total": total_partida, "precio_unitario": precio_unitario,
                }
                pdf_apu = generar_pdf_apu(datos_apu, datos_usuario=datos_usuario_pdf())
                nombre_apu = f"APU_{apu_partida or 'partida'}.pdf".replace(" ", "_")
                st.download_button(
                    "⬇️ Descargar APU en PDF", data=pdf_apu, file_name=nombre_apu,
                    mime="application/pdf", use_container_width=True,
                )
            except Exception:
                import traceback
                print("ERROR generando PDF de APU:", traceback.format_exc())
                st.error("No se pudo generar el PDF del APU. Intenta de nuevo.")


# ============================
# EXPORTAR A PDF
# ============================
# Solo mostrar la exportación de PDF en la sección Cubicación
if option == "Cubicacion" and not st.session_state.get("usuario"):
    # Usuario sin cuenta: no genera informes (según plan)
    st.write("---")
    st.info("📄 **¿Quieres generar el PDF de tu cubicación?** "
            "Crea una cuenta **gratis** para descargar tus informes en PDF.")

if option == "Cubicacion" and st.session_state.get("usuario"):
    st.write("---")
    st.subheader("📄 Exportar Cubicación")
    nombre_proyecto = st.text_input(
        "Nombre del proyecto (opcional)",
        placeholder="Ej: Casa Don Pedro - Angol",
        key="nombre_proyecto"
    )
    # Cada partida usa su volumen guardado (_xxx_vol), que solo se llena al cubicar de verdad.
    # Así, una partida no cubicada queda en 0 y no aparece en el PDF.
    _vol_emp = st.session_state.get("_emp_vol", 0)
    _perd_emp = st.session_state.get("emp_perdida", 5)
    _dos_emp = st.session_state.get("dos_emp", "G-15")
    vol_emp_final = _vol_emp * (1 + _perd_emp / 100) if _vol_emp else 0
    st.session_state["mat_emp"] = calcular_materiales(vol_emp_final, _dos_emp)
    st.session_state["vol_emp"] = vol_emp_final

    _vol_cim = st.session_state.get("_cim_vol", 0)
    _dos_cim = st.session_state.get("dos_cim", "G-20")
    st.session_state["mat_cim"] = calcular_materiales(_vol_cim, _dos_cim)
    st.session_state["vol_pilares"] = _vol_cim

    _vol_sc = st.session_state.get("_sc_vol", 0)
    _dos_sc = st.session_state.get("dos_sc", "G-20")
    st.session_state["mat_sc"] = calcular_materiales(_vol_sc, _dos_sc)
    st.session_state["vol_sc_neto"] = _vol_sc

    _vol_rad = st.session_state.get("_rad_vol", 0)
    _perd_rad = st.session_state.get("radier_perdida", 5)
    _dos_rad = st.session_state.get("dos_rad", "G-20")
    vol_radier_final = _vol_rad * (1 + _perd_rad / 100) if _vol_rad else 0
    st.session_state["mat_rad"] = calcular_materiales(vol_radier_final, _dos_rad)
    st.session_state["vol_radier"] = vol_radier_final

    #st.write("DEBUG vol_rad:", sum(s["largo"] * s["ancho"] * s["espesor"] for s in st.session_state.get("secciones_rad", [])))
    #st.write("DEBUG dos_rad:", st.session_state.get("dos_rad", "no existe"))
    #st.write("DEBUG radier_perdida:", st.session_state.get("radier_perdida", "no existe"))#
    # --- Totales recalculados para el PDF ---
    total_hormigon = (
        st.session_state.get("vol_emp", 0) +
        st.session_state.get("vol_pilares", 0) +
        st.session_state.get("vol_sc_neto", 0) +
        st.session_state.get("vol_radier", 0)
    )
    total_sacos = (
        st.session_state.get("mat_emp", {}).get("cemento_sacos", 0) +
        st.session_state.get("mat_cim", {}).get("cemento_sacos", 0) +
        st.session_state.get("mat_sc",  {}).get("cemento_sacos", 0) +
        st.session_state.get("mat_rad", {}).get("cemento_sacos", 0)
    )
    total_gravilla = (
        st.session_state.get("mat_emp", {}).get("gravilla_kg", 0) +
        st.session_state.get("mat_cim", {}).get("gravilla_kg", 0) +
        st.session_state.get("mat_sc",  {}).get("gravilla_kg", 0) +
        st.session_state.get("mat_rad", {}).get("gravilla_kg", 0)
    )
    total_arena = (
        st.session_state.get("mat_emp", {}).get("arena_kg", 0) +
        st.session_state.get("mat_cim", {}).get("arena_kg", 0) +
        st.session_state.get("mat_sc",  {}).get("arena_kg", 0) +
        st.session_state.get("mat_rad", {}).get("arena_kg", 0)
    )
    total_agua = (
        st.session_state.get("mat_emp", {}).get("agua_lt", 0) +
        st.session_state.get("mat_cim", {}).get("agua_lt", 0) +
        st.session_state.get("mat_sc",  {}).get("agua_lt", 0) +
        st.session_state.get("mat_rad", {}).get("agua_lt", 0)
    )
    if st.button("📄 Generar PDF", type="primary"):
        try:
            pdf_buffer = generar_pdf_cubicacion(
                nombre_proyecto=nombre_proyecto,
                vol_emp=st.session_state.get("vol_emp", 0),
                dos_emp=st.session_state.get("dos_emp", "G-15"),
                mat_emp=st.session_state.get("mat_emp", {"cemento_sacos": 0, "gravilla_kg": 0, "arena_kg": 0, "agua_lt": 0}),
                vol_pilares=st.session_state.get("vol_pilares", 0),
                dos_cim=st.session_state.get("dos_cim", "G-20"),
                mat_cim=st.session_state.get("mat_cim", {"cemento_sacos": 0, "gravilla_kg": 0, "arena_kg": 0, "agua_lt": 0}),
                vol_sc_neto=st.session_state.get("vol_sc_neto", 0),
                dos_sc=st.session_state.get("dos_sc", "G-20"),
                mat_sc=st.session_state.get("mat_sc", {"cemento_sacos": 0, "gravilla_kg": 0, "arena_kg": 0, "agua_lt": 0}),
                vol_radier=st.session_state.get("vol_radier", 0),
                dos_rad=st.session_state.get("dos_rad", "G-20"),
                mat_rad=st.session_state.get("mat_rad", {"cemento_sacos": 0, "gravilla_kg": 0, "arena_kg": 0, "agua_lt": 0}),
                total_hormigon=total_hormigon,
                total_sacos=total_sacos,
                total_gravilla=total_gravilla,
                total_arena=total_arena,
                total_agua=total_agua,
                canal_tipo=st.session_state.get("pdf_canal_tipo", ""),
                cant_piezas_canal=st.session_state.get("pdf_cant_piezas_canal", 0),
                ml_canal=st.session_state.get("pdf_ml_canal", 0),
                largo_canal=st.session_state.get("pdf_largo_canal", 0),
                montante_tipo=st.session_state.get("pdf_montante_tipo", ""),
                total_montantes=st.session_state.get("pdf_total_montantes", 0),
                largo_montante=st.session_state.get("pdf_largo_montante", 0),
                esq_tipo=st.session_state.get("pdf_esq_tipo", ""),
                cant_esquinas=st.session_state.get("pdf_cant_esquinas", 0),
                largo_esq=st.session_state.get("pdf_largo_esq", 0),
                pdf_extra=list(st.session_state.get("materiales_persistente", {}).values()),
                con_marca_agua=not puede_pdf_con_logo(),
                datos_usuario=datos_usuario_pdf(),
            )
            nombre_archivo = f"ObraCubic_{nombre_proyecto or 'cubicacion'}.pdf".replace(" ", "_")
            st.download_button(
                label="⬇️ Descargar PDF",
                data=pdf_buffer,
                file_name=nombre_archivo,
                mime="application/pdf",
            )
        except Exception:
            import traceback
            print("ERROR generando PDF de cubicación:", traceback.format_exc())
            st.error("No se pudo generar el PDF. Revisa que las partidas tengan datos válidos e inténtalo de nuevo.")

    # ------------------------------------------------------------------
    # Botón: Compartir resumen por WhatsApp
    # Arma un mensaje limpio con TODAS las partidas cubicadas (no solo
    # hormigón) y lo abre en WhatsApp. No interfiere con el botón de PDF:
    # usa un enlace <a>, no st.button.
    # ------------------------------------------------------------------
    _persistente = st.session_state.get("materiales_persistente", {})
    if total_hormigon > 0 or _persistente:
        import urllib.parse as _wsp_parse

        _nombre_p = st.session_state.get("nombre_proyecto", "").strip()
        _cabecera = (f"*Cubicación: {_nombre_p}*" if _nombre_p
                     else "*Resumen de Cubicación*")

        _lineas = [_cabecera]

        # 1) Resumen agregado de hormigón (si hay volumen)
        _RUBRO_HORM = "Hormigón y Movimiento de tierra"
        if total_hormigon > 0:
            _lineas += [
                "",
                "*HORMIGÓN (total)*",
                f"*Volumen:* {total_hormigon:.2f} m3",
                f"*Cemento:* {total_sacos} sacos",
                f"*Gravilla:* {total_gravilla} kg",
                f"*Arena:* {total_arena} kg",
                f"*Agua:* {total_agua} lt",
            ]

        # 2) Resto de partidas cubicadas: SOLO materiales comprables
        #    (se descartan filas informativas como tipo, largo, volumen, etc.,
        #     igual que hace el presupuesto)
        _INFO_WORDS = (
            "área", "area", "superficie", "volumen", "dosificación", "dosificacion",
            "espesor", "medida", "tipo", "manos", "capas", "traslape",
            "dirección", "direccion", "altura", "pendiente", "par inclinado",
            "pendolón", "pendolon", "metros lineales netos", "acero total",
            "largo", "orientación", "orientacion", "separación", "separacion",
            "estructura",
        )

        def _es_comprable(_etq, _val):
            if any(_w in _etq.lower() for _w in _INFO_WORDS):
                return False
            _c, _ = parsear_cantidad(_val)
            return _c is not None and _c > 0

        _rubros = {}
        for _bloque in _persistente.values():
            _rubro = _bloque.get("rubro", "")
            if _rubro == _RUBRO_HORM:
                continue  # ya está en el agregado de arriba
            _rubros.setdefault(_rubro, []).append(_bloque)

        for _rubro, _bloques in _rubros.items():
            # Conservar solo partidas que tengan al menos un material comprable
            _bloques_con_mat = []
            for _bloque in _bloques:
                _mats = [(_e, _v) for (_e, _v) in _bloque.get("items", [])
                         if _es_comprable(_e, _v)]
                if _mats:
                    _bloques_con_mat.append((_bloque.get("partida", ""), _mats))
            if not _bloques_con_mat:
                continue  # rubro sin materiales comprables -> no se muestra
            _lineas.append("")
            _lineas.append(f"*{_rubro.upper()}*")
            for _partida, _mats in _bloques_con_mat:
                if _partida:
                    _lineas.append(f"_{_partida}_")
                for _e, _v in _mats:
                    _lineas.append(f"- {_e}: {_v}")

        _lineas += ["", "_Generado con ObraCubic_", "obracubic.streamlit.app"]
        _mensaje_wsp = "\n".join(_lineas)

        _url_wsp = "https://wa.me/?text=" + _wsp_parse.quote(_mensaje_wsp)

        st.markdown(
            f"""
            <a href="{_url_wsp}" target="_blank" style="
                display:block;
                width:100%;
                box-sizing:border-box;
                background-color:#25D366;
                color:#FFFFFF;
                text-align:center;
                padding:13px 20px;
                border-radius:10px;
                font-weight:bold;
                font-size:16px;
                text-decoration:none;
                margin-top:10px;
                box-shadow:0 2px 6px rgba(37,211,102,0.4);
            ">Compartir resumen por WhatsApp</a>
            """,
            unsafe_allow_html=True,
        )
        st.caption("Se abrirá WhatsApp con el resumen de todas las partidas listo para enviar a tu cliente o barraca.")



# ============================
# SECCIÓN PLANES
# ============================
if option == "Planes":
    st.subheader("💎 Planes de ObraCubic")
    st.caption("Elige el plan que mejor se adapte a tu trabajo. Estamos en fase beta: "
               "los planes de pago estarán disponibles muy pronto.")

    # Refrescar plan desde la base (aplica vencimiento automático)
    refrescar_plan_usuario()
    _plan_user = plan_actual()

    col_g, col_b, col_e = st.columns(3)

    # --- Plan Gratis ---
    with col_g:
        with st.container(border=True):
            st.markdown("### ☕ Plan Gratis")
            st.markdown("## $0")
            st.caption("Para empezar a cubicar")
            st.write("---")
            st.markdown(
                "✅ Todas las partidas de cubicación\n\n"
                "✅ Hasta **5 cubicaciones** guardadas\n\n"
                "✅ PDF estándar (con marca ObraCubic)\n\n"
                "❌ Presupuestos\n\n"
                "❌ PDF con tu logo\n\n"
                "❌ Exportar a Excel"
            )
            st.write("")
            if _plan_user == "gratis":
                st.success("✓ Tu plan actual")
            else:
                st.button("Plan Gratis", disabled=True, use_container_width=True, key="plan_g")

    # --- Plan Pro Básico ---
    with col_b:
        with st.container(border=True):
            st.markdown("### 🚀 Plan Pro Básico")
            st.markdown("## $5.990 <span style='font-size:0.5em; color:gray;'>/ mes</span>", unsafe_allow_html=True)
            st.caption("Para profesionales independientes")
            st.write("---")
            st.markdown(
                "✅ Todo lo del Plan Gratis\n\n"
                "✅ **Cubicaciones ilimitadas**\n\n"
                "✅ Historial completo en la nube\n\n"
                "✅ **PDF con tu logo y nombre**\n\n"
                "✅ **Presupuestos** (precios + IVA)\n\n"
                "✅ Lista personal de precios"
            )
            st.write("")
            if _plan_user == "pro_basico":
                st.success("✓ Tu plan actual")
            else:
                st.button("✉️ Probar gratis 7 días", disabled=True, use_container_width=True, key="plan_b",
                          help="Escríbenos a contacto.obracubic@gmail.com para activar tu prueba")

    # --- Plan Pro Élite ---
    with col_e:
        with st.container(border=True):
            st.markdown("### 👑 Plan Pro Élite")
            st.markdown("## $14.990 <span style='font-size:0.5em; color:gray;'>/ mes</span>", unsafe_allow_html=True)
            st.caption("Para oficinas técnicas")
            st.write("---")
            st.markdown(
                "✅ Todo lo del Plan Pro Básico\n\n"
                "✅ Presupuesto avanzado (**APU**)\n\n"
                "✅ **Exportar a Excel** (.xlsx)\n\n"
                "✅ PDF + formatos ejecutivos\n\n"
                "✅ **Bitácora de obra** (notas + fotos)\n\n"
                "✅ Soporte prioritario"
            )
            st.write("")
            if _plan_user == "pro_elite":
                st.success("✓ Tu plan actual")
            else:
                st.button("✉️ Probar gratis 7 días", disabled=True, use_container_width=True, key="plan_e",
                          help="Escríbenos a contacto.obracubic@gmail.com para activar tu prueba")

    st.write("")
    st.info(
        "🚧 **ObraCubic está en fase beta.** Los planes de pago aún no tienen cobro automático, "
        "pero **ya puedes probar los planes Pro gratis**.\n\n"
        "**¿Cómo probar el Plan Pro Básico o Pro Élite?**\n\n"
        "1. Escríbenos a **contacto.obracubic@gmail.com** indicando el correo de tu cuenta "
        "y qué plan quieres probar.\n"
        "2. Te activamos una **prueba gratuita de 7 días** para que uses todas las funciones Pro.\n"
        "3. Si te sirve, puedes mantener el plan mediante una membresía (por ahora, pago por transferencia).\n\n"
        "Estamos afinando la app con usuarios reales antes de habilitar el cobro automático. "
        "¡Gracias por acompañarnos en esta etapa! 🙌"
    )

    st.markdown(
        "<div style='text-align:center; margin-top:8px;'>"
        "📧 <strong>Contacto:</strong> contacto.obracubic@gmail.com"
        "</div>", unsafe_allow_html=True
    )

    # Botones que abren el correo con asunto y mensaje predeterminados
    import urllib.parse as _urlparse
    correo_destino = "contacto.obracubic@gmail.com"
    correo_usuario = ""
    if st.session_state.get("usuario"):
        correo_usuario = st.session_state["usuario"].get("email", "")

    def _link_correo(nombre_plan):
        asunto = f"Solicitud de prueba {nombre_plan} - ObraCubic"
        cuerpo = (
            "Hola equipo de ObraCubic,\n\n"
            f"Quiero probar el plan: {nombre_plan}\n"
            f"El correo de mi cuenta es: {correo_usuario}\n\n"
            "¡Gracias!"
        )
        asunto_q = _urlparse.quote(asunto)
        cuerpo_q = _urlparse.quote(cuerpo)
        return f"mailto:{correo_destino}?subject={asunto_q}&body={cuerpo_q}"

    st.write("")
    bcol1, bcol2 = st.columns(2)
    with bcol1:
        st.markdown(
            f"<a href='{_link_correo('Pro Básico')}' target='_blank' "
            "style='display:block; text-align:center; background:#FF6B00; color:white; "
            "padding:10px; border-radius:8px; text-decoration:none; font-weight:bold;'>"
            "🚀 Solicitar prueba Pro Básico</a>", unsafe_allow_html=True
        )
    with bcol2:
        st.markdown(
            f"<a href='{_link_correo('Pro Élite')}' target='_blank' "
            "style='display:block; text-align:center; background:#1E1E1E; color:white; "
            "padding:10px; border-radius:8px; text-decoration:none; font-weight:bold;'>"
            "👑 Solicitar prueba Pro Élite</a>", unsafe_allow_html=True
        )
    st.caption("Si el botón no abre tu correo, escríbenos manualmente a contacto.obracubic@gmail.com")

    if _plan_user == "sin_cuenta":
        st.caption("Crea una cuenta gratis para empezar a usar ObraCubic.")